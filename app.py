"""
Streamlit app for qPCR data analysis using MAK2 model.
"""

import sys
import os

# Add current directory to Python path for imports
if os.path.dirname(__file__):
    sys.path.insert(0, os.path.dirname(__file__))

import streamlit as st
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from mak2_model import MAK2Model, calculate_amplification_efficiency, pre_estimate_background
from optimizer import MAK2Optimizer
from bootstrap import bootstrap_parameter_uncertainty, BootstrapAnalyzer
from example_data_loader import ExampleDataLoader
from qpcr_data_converter import QPCRDataConverter, load_abi_results_csv
from sample_selector_ui import create_sample_selector_widget
from replicate_analysis import (
    parse_sample_groups,
    calculate_replicate_stats,
    analyze_dilution_series,
    compare_precision,
    plot_dilution_series_comparison
)
from calibration import (
    build_standard_curve,
    build_ct_standard_curve,
    apply_calibration,
    apply_ct_calibration,
    plot_calibration,
    plot_ct_calibration,
    plot_replicate_overlay,
    calculate_replicate_param_summary,
    build_limited_dilution_calibration,
    plot_limited_dilution_diagnostics,
)

# Page config
st.set_page_config(
    page_title="qPCR MAK2 Analyzer",
    page_icon="🧬",
    layout="wide"
)

# ============================================================================
# PERSISTENT RESULTS CACHE
# ============================================================================
# Streamlit Cloud drops the websocket when the browser tab is backgrounded,
# which wipes session state.  We use TWO layers of persistence:
#
# 1. @st.cache_resource — in-process memory, survives websocket drops and
#    session resets within the same Python process.  Most reliable.
# 2. Pickle file on disk — survives app module reloads within the same
#    container.  Backup for layer 1.
#
# Neither survives a full container restart on Streamlit Cloud free tier
# (~7–15 min idle timeout).  That is a platform limitation.

import tempfile, pickle, hashlib

@st.cache_resource
def _get_results_store():
    """Singleton dict shared across all sessions in this process."""
    return {}

_results_store = _get_results_store()

_RESULTS_CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.cache')
os.makedirs(_RESULTS_CACHE_DIR, exist_ok=True)

def _results_cache_path():
    return os.path.join(_RESULTS_CACHE_DIR, 'last_batch_results.pkl')

# Standard curve session-state key prefixes (per-channel suffixed)
_STD_CURVE_KEY_PREFIXES = (
    '_std_curve_d0_df', '_std_curve_d0_summary',
    '_std_curve_ct_df', '_std_curve_ct_summary',
    '_std_curve_variance_df',
)

def _clear_std_curve_keys():
    """Remove all per-channel standard curve session state keys."""
    to_delete = [k for k in st.session_state
                 if any(k.startswith(pfx) for pfx in _STD_CURVE_KEY_PREFIXES)]
    to_delete.append('_std_curve_channels')
    for k in to_delete:
        st.session_state.pop(k, None)

def _checkpoint_cache_path():
    return os.path.join(_RESULTS_CACHE_DIR, 'batch_checkpoint.pkl')

def _save_checkpoint(checkpoint_data):
    """Atomic-write checkpoint to disk after each well. Called per-well."""
    try:
        _results_store['checkpoint'] = checkpoint_data
        path = _checkpoint_cache_path()
        tmp_path = path + '.tmp'
        with open(tmp_path, 'wb') as f:
            pickle.dump(checkpoint_data, f, protocol=pickle.HIGHEST_PROTOCOL)
        os.replace(tmp_path, path)  # atomic on POSIX
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

def _restore_checkpoint():
    """Try to load an incomplete batch checkpoint. Returns dict or None."""
    if 'checkpoint' in _results_store:
        return _results_store['checkpoint']
    path = _checkpoint_cache_path()
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'rb') as f:
            cp = pickle.load(f)
        if cp.get('version') != 1:
            return None
        _results_store['checkpoint'] = cp
        return cp
    except Exception:
        return None

def _clear_checkpoint():
    """Remove checkpoint after batch completes successfully."""
    _results_store.pop('checkpoint', None)
    try:
        path = _checkpoint_cache_path()
        if os.path.exists(path):
            os.unlink(path)
    except OSError:
        pass

def _clear_all():
    """Nuclear reset: clear ALL session state, caches, and checkpoints."""
    # 1. Clear in-memory results store (survives session resets)
    _results_store.clear()
    # 2. Clear disk caches
    _clear_checkpoint()
    try:
        path = _results_cache_path()
        if os.path.exists(path):
            os.unlink(path)
    except OSError:
        pass
    # 3. Clear ALL session state except Streamlit internals
    for key in list(st.session_state.keys()):
        if not key.startswith('_streamlit'):
            del st.session_state[key]

def _save_results_to_disk(results_df, results_list, all_samples, no_signal_samples, cycles, settings, no_signal_fluor=None):
    """Persist batch results to both in-memory store and disk."""
    try:
        slim_list = []
        for r in results_list:
            slim = {k: v for k, v in r.items() if k != 'fluor_data'}
            slim_list.append(slim)
        payload = {
            'results_df': results_df,
            'results_list': slim_list,
            'all_samples': {k: v.tolist() if hasattr(v, 'tolist') else v for k, v in all_samples.items()},
            'no_signal_samples': no_signal_samples,
            'no_signal_fluor': {k: v.tolist() if hasattr(v, 'tolist') else v for k, v in (no_signal_fluor or {}).items()},
            'cycles': cycles.tolist() if hasattr(cycles, 'tolist') else cycles,
            'settings': settings,
        }
        # Layer 1: in-process store (most reliable)
        _results_store['payload'] = payload

        # Layer 2: pickle to disk (backup)
        try:
            with open(_results_cache_path(), 'wb') as f:
                pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)
        except Exception as e:
            st.toast(f"Warning: disk cache save failed: {e}", icon="⚠️")
    except Exception as e:
        st.toast(f"Warning: result cache save failed: {e}", icon="⚠️")

def _restore_results_from_cache():
    """Try to restore batch results from in-memory store or disk."""
    payload = None

    # Layer 1: in-process store
    if 'payload' in _results_store:
        payload = _results_store['payload']

    # Layer 2: disk fallback
    if payload is None:
        try:
            path = _results_cache_path()
            if os.path.exists(path):
                with open(path, 'rb') as f:
                    payload = pickle.load(f)
                # Promote to in-memory store for future restores
                _results_store['payload'] = payload
        except Exception:
            pass

    if payload is None:
        return False

    try:
        st.session_state['batch_results'] = payload['results_df']
        st.session_state['batch_results_list'] = payload['results_list']
        st.session_state['batch_all_samples'] = {
            k: np.array(v) for k, v in payload['all_samples'].items()
        }
        st.session_state['batch_no_signal_samples'] = payload['no_signal_samples']
        _ns_fluor_payload = payload.get('no_signal_fluor', {})
        if _ns_fluor_payload:
            st.session_state['batch_no_signal_fluor'] = {
                k: np.array(v) for k, v in _ns_fluor_payload.items()
            }
        st.session_state['batch_cycles'] = np.array(payload['cycles'])
        st.session_state['batch_settings'] = payload['settings']
        return True
    except Exception as e:
        st.toast(f"Warning: result restore failed: {e}", icon="⚠️")
        return False

# On app startup: if session state has no batch results, try to restore
# from cache.  This covers websocket drops during long computations.
_cache_has_memory = 'payload' in _results_store
_cache_has_disk = os.path.exists(_results_cache_path())
_session_has_results = 'batch_results' in st.session_state

if not _session_has_results:
    if _restore_results_from_cache():
        st.toast("Restored previous batch results", icon="✅")
        _session_has_results = True

# Detect incomplete checkpoint (batch interrupted mid-fit)
_checkpoint_data = _restore_checkpoint()
_has_incomplete_checkpoint = (_checkpoint_data is not None and not _session_has_results)

# Diagnostic sidebar — always visible so we can debug persistence
_has_checkpoint_file = os.path.exists(_checkpoint_cache_path())
with st.sidebar.expander("🔧 Cache diagnostics", expanded=False):
    st.caption(
        f"Session has results: **{_session_has_results}**\n\n"
        f"Memory cache: **{'yes' if _cache_has_memory else 'no'}**\n\n"
        f"Disk cache: **{'yes' if _cache_has_disk else 'no'}**\n\n"
        f"Checkpoint: **{'yes' if _has_checkpoint_file else 'no'}**"
    )

# ============================================================================
# EXCEL EXPORT
# ============================================================================

def _build_excel_download(results_df, extra_sheets=None, chart_sheets=None):
    """Build a multi-sheet Excel file with all available results.

    Args:
        results_df: Main batch results DataFrame (always included).
        extra_sheets: dict of {sheet_name: DataFrame} for additional tabs.
        chart_sheets: dict of {sheet_name: {data: DataFrame, summary: dict,
                      chart_type: str}} for tabs with native Excel charts.
    """
    import io
    from openpyxl.chart import ScatterChart, Reference, Series as XlSeries
    from openpyxl.utils import get_column_letter

    if extra_sheets is None:
        extra_sheets = {}
    if chart_sheets is None:
        chart_sheets = {}
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Batch Results', index=False)
        for sheet_name, df in extra_sheets.items():
            if df is not None and len(df) > 0:
                safe_name = sheet_name[:31]
                df.to_excel(writer, sheet_name=safe_name, index=False)

        # Chart sheets: data + summary stats + scatter chart
        for sheet_name, spec in chart_sheets.items():
            safe_name = sheet_name[:31]
            df = spec.get('data')
            summary = spec.get('summary', {})
            chart_type = spec.get('chart_type', '')

            if df is None or len(df) == 0:
                continue

            df.to_excel(writer, sheet_name=safe_name, index=False)
            ws = writer.sheets[safe_name]
            data_rows = len(df)

            # Write summary stats below data
            summary_start_row = data_rows + 3  # leave a blank row
            row = summary_start_row
            ws.cell(row=row, column=1, value='--- Summary ---')
            row += 1
            for k, v in summary.items():
                ws.cell(row=row, column=1, value=k)
                ws.cell(row=row, column=2, value=v)
                row += 1

            # Create scatter chart based on chart_type
            if chart_type == 'std_curve_d0':
                # log10(D0) vs log10(Copies) scatter
                _add_std_curve_d0_chart(ws, df, data_rows)
            elif chart_type == 'std_curve_ct':
                # Ct vs log10(Copies) scatter
                _add_std_curve_ct_chart(ws, df, data_rows)
            elif chart_type == 'dilution_series':
                _add_dilution_chart(ws, df, data_rows)

    buf.seek(0)
    return buf.getvalue()


def _make_scatter_series(ws, x_col, y_col, data_rows, title="Data",
                         color="4472C4"):
    """Create a scatter series with uniform-color circle markers, no connecting line."""
    from openpyxl.chart import Reference, Series as XlSeries
    from openpyxl.chart.marker import Marker
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

    x_vals = Reference(ws, min_col=x_col, min_row=2, max_row=data_rows + 1)
    y_vals = Reference(ws, min_col=y_col, min_row=2, max_row=data_rows + 1)
    series = XlSeries(y_vals, x_vals, title=title)
    series.marker = Marker(symbol='circle', size=7)
    series.marker.graphicalProperties.solidFill = color
    series.graphicalProperties.line.noFill = True  # no connecting line
    return series


def _style_axis(axis, title, num_fmt="General", log=False):
    """Configure an axis with visible labels, tick marks, and no gridlines."""
    axis.title = title
    axis.delete = False          # force tick labels visible
    axis.numFmt = num_fmt
    axis.majorTickMark = "out"
    axis.majorGridlines = None   # remove gridlines
    if log:
        axis.scaling.logBase = 10


def _add_std_curve_d0_chart(ws, df, data_rows):
    """Add D0 standard curve scatter chart to worksheet."""
    from openpyxl.chart import ScatterChart
    from openpyxl.chart.trendline import Trendline
    from openpyxl.utils import get_column_letter

    cols = list(df.columns)
    x_col = cols.index('log10_D0') + 1 if 'log10_D0' in cols else None
    y_col = cols.index('log10_Copies') + 1 if 'log10_Copies' in cols else None
    if not (x_col and y_col):
        return

    chart = ScatterChart()
    chart.title = "D0 Standard Curve (log-log)"
    chart.legend.position = 'b'
    _style_axis(chart.x_axis, "log10(D0)", "0.00")
    _style_axis(chart.y_axis, "log10(Known Copies)", "0.00")
    chart.width = 18
    chart.height = 12

    series = _make_scatter_series(ws, x_col, y_col, data_rows, "Standards")
    series.trendline = Trendline(trendlineType='linear',
                                 dispRSqr=True, dispEq=True)
    chart.series.append(series)
    ws.add_chart(chart, f"{get_column_letter(len(cols) + 2)}2")


def _add_std_curve_ct_chart(ws, df, data_rows):
    """Add Ct standard curve scatter chart to worksheet."""
    from openpyxl.chart import ScatterChart
    from openpyxl.chart.trendline import Trendline
    from openpyxl.utils import get_column_letter

    cols = list(df.columns)
    x_col = cols.index('Ct') + 1 if 'Ct' in cols else None
    y_col = cols.index('log10_Copies') + 1 if 'log10_Copies' in cols else None
    if not (x_col and y_col):
        return

    chart = ScatterChart()
    chart.title = "Ct Standard Curve"
    chart.legend.position = 'b'
    _style_axis(chart.x_axis, "Ct", "0.0")
    _style_axis(chart.y_axis, "log10(Known Copies)", "0.00")
    chart.width = 18
    chart.height = 12

    series = _make_scatter_series(ws, x_col, y_col, data_rows, "Standards")
    series.trendline = Trendline(trendlineType='linear',
                                 dispRSqr=True, dispEq=True)
    chart.series.append(series)
    ws.add_chart(chart, f"{get_column_letter(len(cols) + 2)}2")


def _add_error_bars(series, sheet_name, sd_col, data_rows):
    """Add y-axis error bars from a standard deviation column."""
    from openpyxl.chart.error_bar import ErrorBars
    from openpyxl.chart.series import NumDataSource, NumRef
    from openpyxl.utils import get_column_letter

    col_letter = get_column_letter(sd_col)
    ref = f"'{sheet_name}'!${col_letter}$2:${col_letter}${data_rows + 1}"
    src = NumDataSource(numRef=NumRef(f=ref))
    series.errBars = ErrorBars(errDir='y', errBarType='both',
                               errValType='cust', plus=src, minus=src)


def _add_dilution_chart(ws, df, data_rows):
    """Add dilution series scatter charts using log10-transformed data.

    Plots log10(Dilution) on x and Ct or log10(D0) on y, both on plain
    linear axes, so the linear trendline equation is y = mx + b.
    """
    from openpyxl.chart import ScatterChart
    from openpyxl.chart.trendline import Trendline
    from openpyxl.utils import get_column_letter

    cols = list(df.columns)

    def _col(name):
        return cols.index(name) + 1 if name in cols else None

    log_dil = _col('log10_Dilution')
    ct_col = _col('Ct_Mean')
    log_d0 = _col('log10_D0_Mean')
    ct_sd_col = _col('Ct_SD')
    sheet_name = ws.title

    if not log_dil:
        return  # need log10 columns

    # Chart 1: Ct vs log10(Dilution) — linear axes, linear trendline
    if ct_col:
        chart1 = ScatterChart()
        chart1.title = "Ct vs log10(Dilution)"
        chart1.legend.position = 'b'
        _style_axis(chart1.x_axis, "log10(Dilution)", "0.0")
        _style_axis(chart1.y_axis, "Mean Ct", "0.0")
        chart1.width = 16
        chart1.height = 10
        series = _make_scatter_series(ws, log_dil, ct_col, data_rows,
                                      "Ct Mean", color="4472C4")
        if ct_sd_col:
            _add_error_bars(series, sheet_name, ct_sd_col, data_rows)
        series.trendline = Trendline(trendlineType='linear',
                                     dispRSqr=True, dispEq=True)
        chart1.series.append(series)
        ws.add_chart(chart1, f"{get_column_letter(len(cols) + 2)}2")

    # Chart 2: log10(D0) vs log10(Dilution) — linear axes, linear trendline
    if log_d0:
        chart2 = ScatterChart()
        chart2.title = "log10(D0) vs log10(Dilution)"
        chart2.legend.position = 'b'
        _style_axis(chart2.x_axis, "log10(Dilution)", "0.0")
        _style_axis(chart2.y_axis, "log10(D0)", "0.0")
        chart2.width = 16
        chart2.height = 10
        series = _make_scatter_series(ws, log_dil, log_d0, data_rows,
                                      "D0 Mean", color="ED7D31")
        series.trendline = Trendline(trendlineType='linear',
                                     dispRSqr=True, dispEq=True)
        chart2.series.append(series)
        ws.add_chart(chart2, f"{get_column_letter(len(cols) + 2)}15")

# ============================================================================
# HELPERS
# ============================================================================

def _ch(name):
    """Return the channel prefix of a sample name.
    'FAM_A1' → 'FAM', 'target::well' → 'target', else 'default'."""
    if '::' in name: return name.split('::')[0]
    if '_'  in name: return name.split('_')[0]
    return 'default'

def _get_well_pos(name):
    """Extract bare well position from a sample name.
    'FAM_A1' → 'A1', 'target::A1' → 'A1', 'A1' → 'A1'."""
    if '::' in name: return name.split('::')[1] if len(name.split('::')) > 1 else name
    if '_'  in name: return '_'.join(name.split('_')[1:]) if len(name.split('_')) > 1 else name
    return name


# ============================================================================
# MAIN APP
# ============================================================================

st.title("🧬 qPCR Model Fitting with MAK2+")

# Community signup banner
st.info("""
👋 **MAK2+ is in open beta!** Free to use for research and education.
💬 [GitHub Discussions](https://github.com/gboggy2/MAK2-plus/discussions) for questions | 🐛 [Report issues](https://github.com/gboggy2/MAK2-plus/issues)
""", icon="🧬")
st.markdown("""
This tool fits the MAK2 mechanistic model to qPCR data, including primer depletion effects.

**Smart Truncation:** Data is automatically truncated at the max derivative + N cycles (default: 3),
avoiding deep plateau artifacts (enzyme degradation, dNTP depletion) not modeled by primer depletion.

Based on [Boggy & Woolf (2010)](https://doi.org/10.1371/journal.pone.0012355) with extensions for primer concentration tracking.
""")

# Sidebar for data input
st.sidebar.header("Data Input")

if st.sidebar.button("🗑️ Clear All & Start Over", help="Clear all data, results, metadata, and caches"):
    _clear_all()
    st.rerun()

# Initialize example data loader (once)
if 'example_loader' not in st.session_state:
    st.session_state.example_loader = ExampleDataLoader(data_dir="example_data")

data_source = st.sidebar.radio(
    "Choose data source:",
    ["Example Data", "Upload File", "Manual Entry", "Load Previous Results"]
)

cycles = None
fluorescence = None
all_samples = {}  # For batch processing
batch_mode = False

if data_source == "Example Data":
    # Get available datasets
    display_names, filename_map = st.session_state.example_loader.create_streamlit_selector()
    
    # Create dropdown selector
    selected_display = st.sidebar.selectbox(
        "Select example dataset:",
        display_names,
        key="example_dataset_selector"
    )
    
    # Get the actual filename
    selected_filename = filename_map[selected_display]
    
    # Show dataset info in an expander
    dataset_info = st.session_state.example_loader.get_dataset_info(selected_filename)
    
    with st.sidebar.expander("ℹ️ About this dataset", expanded=False):
        st.markdown(f"**{dataset_info['display_name']}**")
        st.write(dataset_info['description'])
        
        col1, col2 = st.sidebar.columns(2)
        col1.metric("Samples", dataset_info['n_samples'])
        col2.metric("Cycles", dataset_info['n_cycles'])
        
        if 'characteristics' in dataset_info:
            st.write("**Characteristics:**")
            for char in dataset_info['characteristics']:
                st.write(f"• {char}")
        
        if 'expected_results' in dataset_info:
            st.info(f"**Expected:** {dataset_info['expected_results']}")
    
    # Load button
    if st.sidebar.button("📥 Load Example Data", type="primary"):
        try:
            # Load the dataset
            loaded_cycles, fluorescence_df, metadata = st.session_state.example_loader.load_dataset(
                selected_filename
            )
            
            # Clear stale results from previous dataset
            for key in ['batch_results', 'batch_results_list', 'fitted_params',
                        'optimizer', 'bootstrap_results',
                        '_no_signal_df', '_replicate_stats_df',
                        '_precision_comparison_df',
                        '_limited_dilution_df', '_dilution_series_df',
                        '_dilution_series_summary']:
                if key in st.session_state:
                    del st.session_state[key]
            _clear_std_curve_keys()

            # Store in session state immediately
            st.session_state.loaded_cycles = loaded_cycles
            st.session_state.fluorescence_df = fluorescence_df
            st.session_state.dataset_name = selected_display
            st.session_state.data_loaded = True
            
            # Check if multiple samples (batch mode)
            if fluorescence_df.shape[1] > 1:
                st.session_state.batch_mode_available = True
            else:
                st.session_state.batch_mode_available = False
            
            st.sidebar.success(f"✅ Loaded {selected_display}")
            st.rerun()
            
        except Exception as e:
            st.sidebar.error(f"Error loading dataset: {str(e)}")
    
    # After loading, handle batch mode and sample selection
    if st.session_state.get('data_loaded', False):
        loaded_cycles = st.session_state.loaded_cycles
        fluorescence_df = st.session_state.fluorescence_df
        
        if st.session_state.get('batch_mode_available', False):
            st.sidebar.info(f"📊 Loaded {fluorescence_df.shape[1]} samples")
            batch_mode = st.sidebar.checkbox("Batch fit all samples", value=True, key="example_batch_mode")
            
            if batch_mode:
                # Store all samples
                cycles = loaded_cycles
                for col in fluorescence_df.columns:
                    all_samples[col] = fluorescence_df[col].values
                st.sidebar.success(f"Loaded {len(all_samples)} samples")
                
                # Select one to preview
                preview_sample = st.sidebar.selectbox("Preview sample:", list(all_samples.keys()))
                fluorescence = all_samples[preview_sample]
            else:
                # Single sample mode - let user select
                sample_col = st.sidebar.selectbox("Select sample column:", fluorescence_df.columns)
                cycles = loaded_cycles
                fluorescence = fluorescence_df[sample_col].values
        else:
            # Single sample dataset
            cycles = loaded_cycles
            fluorescence = fluorescence_df.iloc[:, 0].values
    elif data_source == "Example Data":
        # Show instruction if no data loaded yet
        st.sidebar.info("👆 Click 'Load Example Data' to begin")

elif data_source == "Upload File":
    uploaded_file = st.sidebar.file_uploader(
        "Upload CSV or Excel file",
        type=['csv', 'xlsx', 'xls'],
        help="Supports various qPCR formats: simple, wide, Bio-Rad CFX, QuantStudio, ABI Multicomponent"
    )

    meta_file = st.sidebar.file_uploader(
        "Upload plate metadata CSV (optional)",
        type=['csv'],
        help=(
            "ABI results table CSV with sample names, known quantities (standard curve), "
            "instrument Ct values, and per-channel Ct thresholds. "
            "When provided, instrument thresholds replace auto-computed ones."
        ),
        key="meta_file_uploader"
    )

    # Parse metadata file and store in session state
    if meta_file is not None:
        meta_key = f"{meta_file.name}_{meta_file.size}"
        if st.session_state.get('last_meta_file_key') != meta_key:
            try:
                parsed_meta = load_abi_results_csv(meta_file)
                st.session_state['abi_results_meta'] = parsed_meta
                st.session_state['sample_metadata'] = parsed_meta['sample_metadata']
                st.session_state['last_meta_file_key'] = meta_key
                # Clear fit results so thresholds are re-applied on next run
                for key in ['batch_results', 'batch_results_list', 'fitted_params',
                            'optimizer', 'bootstrap_results',
                            '_no_signal_df', '_replicate_stats_df', '_precision_comparison_df',
                            '_limited_dilution_df',
                            '_dilution_series_df', '_dilution_series_summary']:
                    if key in st.session_state:
                        del st.session_state[key]
                _clear_std_curve_keys()
                n = parsed_meta['n_wells']
                chs = ', '.join(f"{ch}={v}" for ch, v in parsed_meta['channel_thresholds'].items())
                st.sidebar.success(f"✅ Metadata: {n} wells loaded  \nThresholds: {chs}")
            except Exception as e:
                st.sidebar.error(f"Could not parse metadata file: {e}")
    elif st.session_state.get('last_meta_file_key') is not None:
        # Metadata file was removed — clear it
        for key in ['abi_results_meta', 'sample_metadata', 'last_meta_file_key']:
            if key in st.session_state:
                del st.session_state[key]

    if uploaded_file is not None:
        # Clear fitted results when new file uploaded (use name+size to detect re-uploads)
        # Note: id(uploaded_file) changes on every rerun, so we use name+size instead
        file_key = f"{uploaded_file.name}_{uploaded_file.size}"
        if 'last_uploaded_file_key' not in st.session_state:
            st.session_state['last_uploaded_file_key'] = file_key
        elif st.session_state['last_uploaded_file_key'] != file_key:
            st.session_state['last_uploaded_file_key'] = file_key
            # Clear all previous results for the new file
            for key in ['fitted_params', 'optimizer', 'bootstrap_results', 'batch_results',
                       'batch_results_list',
                       'uploaded_cycles', 'uploaded_samples', 'uploaded_metadata',
                       'selected_target', 'target_confirmed', 'selected_channels', 'all_targets',
                       'abi_extra_info', 'abi_passive_reference', 'rox_normalized',
                       'abi_results_meta', 'sample_metadata', 'last_meta_file_key',
                       # Excel export data
                       '_no_signal_df', '_replicate_stats_df', '_precision_comparison_df',
                       '_limited_dilution_df',
                       '_dilution_series_df', '_dilution_series_summary']:
                if key in st.session_state:
                    del st.session_state[key]
            _clear_std_curve_keys()
        
        # Process file with enhanced converter on first load
        if 'uploaded_cycles' not in st.session_state:
            try:
                with st.sidebar.status("📂 Processing file...", expanded=True) as status:
                    st.write("Reading file...")
                    
                    # Use enhanced converter
                    converter = QPCRDataConverter(add_offset=True, offset_value=1e-5)
                    loaded_cycles, loaded_samples, metadata = converter.load_from_file(uploaded_file)
                    
                    st.write(f"✅ Detected format: **{metadata['format']}**")
                    st.write(f"✅ Loaded {metadata['n_samples']} samples")
                    st.write(f"✅ {metadata['n_cycles']} cycles per sample")
                    
                    # Store in session state
                    st.session_state.uploaded_cycles = loaded_cycles
                    st.session_state.uploaded_samples = loaded_samples
                    st.session_state.uploaded_metadata = metadata
                    
                    status.update(label="✅ File processed!", state="complete")
                
            except Exception as e:
                st.sidebar.error(f"Error loading file: {str(e)}")
                st.sidebar.info("💡 Expected format: First column = cycles, other columns = samples")
                import traceback
                with st.sidebar.expander("🔍 Error Details"):
                    st.code(traceback.format_exc())
        
        # After loading, show sample selector in main area
        if 'uploaded_cycles' in st.session_state:
            loaded_cycles = st.session_state.uploaded_cycles
            loaded_samples = st.session_state.uploaded_samples
            metadata = st.session_state.uploaded_metadata
            
            
            # Auto-load all targets for multiplexed files
            if metadata.get('requires_target_selection', False) and 'selected_target' not in st.session_state:
                extra_info = metadata['extra_info']
                targets = extra_info['targets']

                try:
                    with st.spinner(f"Loading {len(targets)} targets..."):
                        converter = QPCRDataConverter(add_offset=True, offset_value=1e-5)
                        all_target_data = converter.filter_all_targets(extra_info)

                        # Combine all targets — prefix well names with target
                        combined_samples = {}
                        combined_metadata = {}
                        ref_cycles = None

                        for target_name, (t_cycles, t_samples, t_meta) in all_target_data.items():
                            if ref_cycles is None:
                                ref_cycles = t_cycles
                            for well, fluor in t_samples.items():
                                key = f"{target_name}::{well}"
                                combined_samples[key] = fluor
                                if t_meta and well in t_meta:
                                    combined_metadata[key] = {**t_meta[well], '_target': target_name, '_well': well}
                                else:
                                    combined_metadata[key] = {'_target': target_name, '_well': well}

                        # Store combined data
                        st.session_state.selected_target = 'All'
                        st.session_state.all_targets = targets
                        st.session_state.uploaded_cycles = ref_cycles
                        st.session_state.uploaded_samples = combined_samples
                        st.session_state.sample_metadata = combined_metadata

                        st.session_state.uploaded_metadata = {
                            'format': metadata['format'],
                            'n_samples': len(combined_samples),
                            'n_cycles': len(ref_cycles),
                            'sample_names': list(combined_samples.keys()),
                            'cycle_range': (ref_cycles.min(), ref_cycles.max()),
                            'requires_target_selection': False
                        }

                        st.success(f"✅ Loaded {len(combined_samples)} samples across {len(targets)} targets: {', '.join(targets)}")
                        st.rerun()
                except Exception as e:
                    st.error(f"Error loading targets: {e}")
                    import traceback
                    st.code(traceback.format_exc())

                st.stop()
            
            # Auto-load all dye channels for ABI Multicomponent files
            if metadata.get('requires_channel_selection', False) and 'selected_channels' not in st.session_state:
                extra_info = metadata['extra_info']
                channels = extra_info['channels']
                passive_reference = extra_info.get('passive_reference')

                # Store extra_info for later re-normalization
                st.session_state.abi_extra_info = extra_info
                st.session_state.abi_passive_reference = passive_reference

                # Default: normalize by ROX if passive reference is available
                normalize_rox = passive_reference is not None

                try:
                    with st.spinner(f"Loading {len(channels)} dye channels..."):
                        converter = QPCRDataConverter(add_offset=True, offset_value=1e-5)

                        combined_samples = {}
                        ref_cycles = None

                        for channel in channels:
                            # Always load raw RFU — ROX normalization is applied only
                            # inside calculate_ct() via rox_by_well, never to the
                            # fluorescence arrays used for MAK2+ fitting.
                            c_cycles, c_samples, _ = converter.filter_by_channel(
                                extra_info, channel, normalize_by_reference=False
                            )
                            if ref_cycles is None:
                                ref_cycles = c_cycles
                            for well_pos, fluor in c_samples.items():
                                key = f"{channel}_{well_pos}"
                                combined_samples[key] = fluor

                        st.session_state.selected_channels = channels
                        st.session_state.uploaded_cycles = ref_cycles
                        st.session_state.uploaded_samples = combined_samples
                        st.session_state.rox_normalized = normalize_rox

                        # Store ROX per-well for Ct normalization
                        if passive_reference and 'samples_by_channel' in extra_info:
                            rox_data = extra_info['samples_by_channel'].get(passive_reference, {})
                            st.session_state.rox_by_well = {well: np.array(arr) for well, arr in rox_data.items()}
                        else:
                            st.session_state.rox_by_well = {}

                        st.session_state.uploaded_metadata = {
                            'format': metadata['format'],
                            'n_samples': len(combined_samples),
                            'n_cycles': len(ref_cycles),
                            'sample_names': list(combined_samples.keys()),
                            'cycle_range': (ref_cycles.min(), ref_cycles.max()),
                            'requires_channel_selection': False,
                            'requires_target_selection': False,
                        }

                        ref_note = (
                            f" (passive reference **{passive_reference}** used for Ct normalization)" if normalize_rox and passive_reference
                            else f" (passive reference **{passive_reference}** excluded)" if passive_reference
                            else ""
                        )
                        st.success(
                            f"✅ Loaded {len(combined_samples)} samples across "
                            f"{len(channels)} channels: {', '.join(channels)}{ref_note}"
                        )
                        st.rerun()
                except Exception as e:
                    st.error(f"Error loading channels: {e}")
                    import traceback
                    st.code(traceback.format_exc())

                st.stop()

            # Show file info in sidebar
            all_targets = st.session_state.get('all_targets', [])
            all_channels = st.session_state.get('selected_channels', [])
            if all_targets:
                target_label = f" ({len(all_targets)} targets)"
            elif all_channels:
                target_label = f" ({len(all_channels)} channels)"
            elif 'selected_target' in st.session_state:
                target_label = f" ({st.session_state.selected_target})"
            else:
                target_label = ""
            st.sidebar.success(f"✅ {metadata['n_samples']} samples loaded{target_label}")

            # ROX normalization toggle (ABI Multicomponent files only)
            abi_extra_info = st.session_state.get('abi_extra_info')
            abi_passive_ref = st.session_state.get('abi_passive_reference')
            if abi_extra_info is not None and abi_passive_ref is not None:
                current_rox_norm = st.session_state.get('rox_normalized', True)
                with st.sidebar.expander(f"🧪 ROX Normalization ({abi_passive_ref})", expanded=False):
                    new_rox_norm = st.checkbox(
                        f"Use {abi_passive_ref} for Ct normalization",
                        value=current_rox_norm,
                        help=(
                            f"When enabled, Ct is computed on Rn = raw / {abi_passive_ref} "
                            "to match the instrument's ΔRn threshold units. "
                            "MAK2+ fitting always uses raw RFU regardless of this setting."
                        ),
                        key="rox_norm_toggle"
                    )
                    if new_rox_norm != current_rox_norm:
                        # Re-load channels with new normalization setting
                        try:
                            channels = abi_extra_info['channels']
                            converter = QPCRDataConverter(add_offset=True, offset_value=1e-5)
                            combined_samples = {}
                            ref_cycles = None
                            for channel in channels:
                                # Always raw RFU — ROX is applied only in calculate_ct()
                                c_cycles, c_samples, _ = converter.filter_by_channel(
                                    abi_extra_info, channel, normalize_by_reference=False
                                )
                                if ref_cycles is None:
                                    ref_cycles = c_cycles
                                for well_pos, fluor in c_samples.items():
                                    combined_samples[f"{channel}_{well_pos}"] = fluor

                            # Update session state — clear fit results since scale changed
                            st.session_state.uploaded_cycles = ref_cycles
                            st.session_state.uploaded_samples = combined_samples
                            st.session_state.rox_normalized = new_rox_norm

                            # Store ROX per-well for Ct normalization
                            passive_ref = st.session_state.get('abi_passive_reference', None)
                            if passive_ref and 'samples_by_channel' in abi_extra_info:
                                rox_data = abi_extra_info['samples_by_channel'].get(passive_ref, {})
                                st.session_state.rox_by_well = {well: np.array(arr) for well, arr in rox_data.items()}
                            else:
                                st.session_state.rox_by_well = {}
                            st.session_state.uploaded_metadata = {
                                **st.session_state.uploaded_metadata,
                                'n_samples': len(combined_samples),
                                'n_cycles': len(ref_cycles),
                                'sample_names': list(combined_samples.keys()),
                                'cycle_range': (ref_cycles.min(), ref_cycles.max()),
                            }
                            for key in ['fitted_params', 'optimizer', 'bootstrap_results',
                                        'batch_results', 'batch_results_list',
                                        'upload_ready_batch',
                                        'upload_ready_single', 'upload_batch_samples',
                                        '_no_signal_df', '_replicate_stats_df',
                                        '_precision_comparison_df',
                                        '_limited_dilution_df', '_dilution_series_df',
                                        '_dilution_series_summary']:
                                if key in st.session_state:
                                    del st.session_state[key]
                            _clear_std_curve_keys()
                            norm_word = "with" if new_rox_norm else "without"
                            st.success(f"Reloaded {norm_word} {abi_passive_ref} normalization")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error reloading: {e}")

            with st.sidebar.expander("📋 File Details"):
                st.write(f"**Format:** {metadata['format']}")
                st.write(f"**Cycles:** {metadata['n_cycles']}")
                st.write(f"**Cycle range:** {metadata['cycle_range'][0]} - {metadata['cycle_range'][1]}")
                if all_targets:
                    st.write(f"**Targets:** {', '.join(all_targets)}")
                elif all_channels:
                    st.write(f"**Channels:** {', '.join(all_channels)}")
                elif 'selected_target' in st.session_state:
                    st.write(f"**Target:** {st.session_state.selected_target}")
                
                # Debug info
                with st.expander("🔍 Debug Info"):
                    st.write(f"loaded_samples dict size: {len(loaded_samples)}")
                    st.write(f"metadata n_samples: {metadata['n_samples']}")
                    st.write(f"requires_target_selection: {metadata.get('requires_target_selection', 'N/A')}")
                    if len(loaded_samples) > 0:
                        st.write(f"Sample names: {list(loaded_samples.keys())[:5]}")
            
            # Show sample selector in main area if no data fitted yet
            if cycles is None or fluorescence is None:
                # Check if we have samples loaded
                if len(loaded_samples) == 0:
                    st.warning("⚠️ No samples loaded. Please select a target first.")
                    st.info("👆 Use the sidebar to upload a file and select a target")
                else:
                    st.markdown("## 📂 Select Samples to Analyze")
                    st.info(f"📊 Loaded **{metadata['n_samples']} samples** from `{uploaded_file.name}`")
                    
                    # Create sample selector interface
                    from sample_selector_ui import SampleSelector
                    selector = SampleSelector(loaded_cycles, loaded_samples)
                    
                    # Show full interface
                    mode_key, selected_samples = selector.create_full_interface(
                        key_prefix="upload",
                        show_preview=True
                    )
                    
                    # Add "Ready to Fit" button
                    if selected_samples:
                        st.markdown("---")
                        col1, col2, col3 = st.columns([1, 2, 1])
                        with col2:
                            if st.button("🚀 Ready to Fit Selected Samples", type="primary", use_container_width=True):
                                # Store selected samples for fitting
                                if mode_key == "batch":
                                    # Batch mode
                                    cycles = loaded_cycles
                                    batch_mode = True
                                    all_samples = {name: loaded_samples[name] for name in selected_samples}
                                    # Set first sample as preview
                                    fluorescence = loaded_samples[selected_samples[0]]
                                    st.session_state.upload_ready_batch = True
                                    st.session_state.upload_batch_samples = all_samples
                                    st.session_state.upload_preview_sample = selected_samples[0]
                                else:
                                    # Single mode
                                    cycles = loaded_cycles
                                    fluorescence = loaded_samples[selected_samples[0]]
                                    batch_mode = False
                                    st.session_state.upload_ready_single = True
                                
                                st.session_state.upload_cycles = cycles
                                st.session_state.upload_fluorescence = fluorescence
                                st.session_state.upload_batch_mode = batch_mode
                                st.rerun()
            
            # If already ready to fit, load from session state
            if st.session_state.get('upload_ready_batch', False):
                cycles = st.session_state.upload_cycles
                all_samples = st.session_state.upload_batch_samples
                batch_mode = True
                preview_sample_name = st.session_state.get('upload_preview_sample')
                fluorescence = all_samples[preview_sample_name]

                # Show info in sidebar
                st.sidebar.info(f"🎯 Ready to fit {len(all_samples)} samples (batch mode)")
                # Allow changing preview sample
                preview_sample_name = st.sidebar.selectbox(
                    "Preview sample:",
                    list(all_samples.keys()),
                    index=list(all_samples.keys()).index(preview_sample_name) if preview_sample_name in all_samples else 0
                )
                fluorescence = all_samples[preview_sample_name]
                st.session_state.upload_preview_sample = preview_sample_name

                # Button to change selection/mode
                if st.sidebar.button("↩️ Change Selection", key="change_selection_batch"):
                    for key in ['upload_ready_batch', 'upload_ready_single', 'upload_batch_samples',
                               'upload_preview_sample', 'upload_cycles', 'upload_fluorescence',
                               'upload_batch_mode', 'fitted_params', 'optimizer', 'batch_results']:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.rerun()

            elif st.session_state.get('upload_ready_single', False):
                cycles = st.session_state.upload_cycles
                fluorescence = st.session_state.upload_fluorescence
                batch_mode = False

                st.sidebar.success("🎯 Ready to fit single sample")

                # Button to change selection/mode
                if st.sidebar.button("↩️ Change Selection", key="change_selection_single"):
                    for key in ['upload_ready_batch', 'upload_ready_single', 'upload_batch_samples',
                               'upload_preview_sample', 'upload_cycles', 'upload_fluorescence',
                               'upload_batch_mode', 'fitted_params', 'optimizer', 'batch_results']:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.rerun()

elif data_source == "Manual Entry":
    manual_input = st.sidebar.text_area(
        "Enter cycle,fluorescence pairs (one per line):",
        "1,0.1\n2,0.15\n3,0.25\n4,0.45\n5,0.85",
        key="manual_input_text"
    )
    if manual_input:
        try:
            lines = [line.strip() for line in manual_input.split('\n') if line.strip()]
            data_list = [list(map(float, line.split(','))) for line in lines]
            cycles = np.array([d[0] for d in data_list])
            fluorescence = np.array([d[1] for d in data_list])
            st.sidebar.success(f"Parsed {len(cycles)} data points")
            
            # Clear fitted results if input changed
            if 'last_manual_input' not in st.session_state:
                st.session_state['last_manual_input'] = manual_input
            elif st.session_state['last_manual_input'] != manual_input:
                st.session_state['last_manual_input'] = manual_input
                if 'fitted_params' in st.session_state:
                    del st.session_state['fitted_params']
                if 'optimizer' in st.session_state:
                    del st.session_state['optimizer']
                if 'bootstrap_results' in st.session_state:
                    del st.session_state['bootstrap_results']
        except:
            st.sidebar.error("Invalid format. Use: cycle,fluorescence")

elif data_source == "Load Previous Results":
    _prev_file = st.sidebar.file_uploader(
        "Upload a previous batch_fit_results .xlsx",
        type=['xlsx'],
        help="Upload an Excel file exported from a previous MAK2+ batch fit to restore all results.",
        key="prev_results_uploader",
    )
    if _prev_file is not None:
        _prev_key = f"{_prev_file.name}_{_prev_file.size}"
        if st.session_state.get('_prev_results_key') != _prev_key:
            try:
                _prev_sheets = pd.read_excel(_prev_file, sheet_name=None)

                # ── Batch Results ──
                if 'Batch Results' not in _prev_sheets:
                    st.sidebar.error("No 'Batch Results' sheet found in file.")
                else:
                    _prev_results_df = _prev_sheets['Batch Results']

                    # ── Input Data → cycles + all_samples + fluor_data ──
                    _prev_cycles = None
                    _prev_all_samples = {}
                    _prev_no_signal = {}
                    if 'Input Data' in _prev_sheets:
                        _input_df = _prev_sheets['Input Data']
                        if 'Cycle' in _input_df.columns:
                            _prev_cycles = _input_df['Cycle'].values.astype(float)
                            for _col in _input_df.columns:
                                if _col != 'Cycle':
                                    _prev_all_samples[_col] = _input_df[_col].values.astype(float)

                    # ── Settings → batch_settings ──
                    _prev_settings = {}
                    if 'Settings' in _prev_sheets:
                        _sdf = _prev_sheets['Settings']
                        _flat = {}
                        for _, _sr in _sdf.iterrows():
                            _flat[str(_sr['Setting'])] = str(_sr['Value']) if pd.notna(_sr['Value']) else ''
                        # Reconstruct typed dict
                        _nested_keys = {}  # parent → {child: val}
                        for _fk, _fv in _flat.items():
                            if '.' in _fk:
                                _parent, _child = _fk.split('.', 1)
                                _nested_keys.setdefault(_parent, {})[_child] = _fv
                            else:
                                _prev_settings[_fk] = _fv
                        # Convert types
                        for _k in ['first_fit_cycle', 'truncate_cycle',
                                    'global_threshold', 'global_baseline_mean']:
                            if _k in _prev_settings:
                                try:
                                    _prev_settings[_k] = float(_prev_settings[_k]) if _prev_settings[_k] not in ('', 'None') else None
                                except ValueError:
                                    _prev_settings[_k] = None
                        for _k in ['cycles_before_max', 'cycles_after_max']:
                            if _k in _prev_settings:
                                try:
                                    _prev_settings[_k] = int(float(_prev_settings[_k]))
                                except (ValueError, TypeError):
                                    _prev_settings[_k] = 3
                        if 'auto_truncate' in _prev_settings:
                            _prev_settings['auto_truncate'] = _prev_settings['auto_truncate'] in ('True', 'true', '1')
                        # Nested dicts
                        for _nk, _nv in _nested_keys.items():
                            _typed = {}
                            for _ck, _cv in _nv.items():
                                try:
                                    _typed[_ck] = float(_cv) if _cv not in ('', 'None') else None
                                except ValueError:
                                    _typed[_ck] = _cv
                            _prev_settings[_nk] = _typed
                        # Defaults
                        _prev_settings.setdefault('custom_bounds_dict', None)

                    # ── Metadata → sample_metadata ──
                    _prev_meta = {}
                    if 'Metadata' in _prev_sheets:
                        _mdf = _prev_sheets['Metadata']
                        for _, _mr in _mdf.iterrows():
                            _wk = str(_mr.get('Well_Key', ''))
                            if _wk:
                                _md = {c: _mr[c] for c in _mdf.columns if c != 'Well_Key' and pd.notna(_mr[c])}
                                _prev_meta[_wk] = _md

                    # ── No Signal samples ──
                    _prev_no_signal_fluor = {}
                    if 'No Signal Samples' in _prev_sheets:
                        _ns_df = _prev_sheets['No Signal Samples']
                        if 'Sample' in _ns_df.columns:
                            for _, _nsr in _ns_df.iterrows():
                                _ns_name = str(_nsr['Sample'])
                                # Store fluor data separately
                                if _ns_name in _prev_all_samples:
                                    _prev_no_signal_fluor[_ns_name] = _prev_all_samples.pop(_ns_name)
                                # Build metadata dict from the No Signal sheet columns
                                _ns_meta = {}
                                if 'Reason' in _ns_df.columns and pd.notna(_nsr.get('Reason')):
                                    _ns_meta['reason'] = str(_nsr['Reason'])
                                if 'Fluorescence Range' in _ns_df.columns and pd.notna(_nsr.get('Fluorescence Range')):
                                    try:
                                        _ns_meta['F_range'] = float(_nsr['Fluorescence Range'])
                                    except (ValueError, TypeError):
                                        _ns_meta['F_range'] = 0.0
                                if '% of Max on Plate' in _ns_df.columns and pd.notna(_nsr.get('% of Max on Plate')):
                                    try:
                                        _ns_meta['F_range_pct'] = float(str(_nsr['% of Max on Plate']).rstrip('%'))
                                    except (ValueError, TypeError):
                                        _ns_meta['F_range_pct'] = 0.0
                                # Also check alternate column names (Range, % of Max) from older exports
                                if 'F_range' not in _ns_meta and 'Range' in _ns_df.columns:
                                    try:
                                        _ns_meta['F_range'] = float(_nsr.get('Range', 0))
                                    except (ValueError, TypeError):
                                        pass
                                if 'F_range_pct' not in _ns_meta and '% of Max' in _ns_df.columns:
                                    try:
                                        _ns_meta['F_range_pct'] = float(str(_nsr.get('% of Max', '0')).rstrip('%'))
                                    except (ValueError, TypeError):
                                        pass
                                _ns_meta.setdefault('reason', 'No signal detected')
                                _ns_meta.setdefault('F_range', 0.0)
                                _ns_meta.setdefault('F_range_pct', 0.0)
                                _prev_no_signal[_ns_name] = _ns_meta

                    # ── Rebuild results_list with fluor_data ──
                    _prev_results_list = []
                    for _, _rr in _prev_results_df.iterrows():
                        _rd = _rr.to_dict()
                        _sn = _rd.get('Sample', '')
                        if _sn in _prev_all_samples:
                            _rd['fluor_data'] = _prev_all_samples[_sn]
                        elif _sn in _prev_no_signal_fluor:
                            _rd['fluor_data'] = _prev_no_signal_fluor[_sn]
                        else:
                            _rd['fluor_data'] = None
                        _prev_results_list.append(_rd)

                    # ── Set session state ──
                    st.session_state['batch_results'] = _prev_results_df
                    st.session_state['batch_results_list'] = _prev_results_list
                    st.session_state['batch_all_samples'] = _prev_all_samples
                    st.session_state['batch_no_signal_samples'] = _prev_no_signal
                    st.session_state['batch_no_signal_fluor'] = _prev_no_signal_fluor
                    if _prev_cycles is not None:
                        st.session_state['batch_cycles'] = _prev_cycles
                    if _prev_settings:
                        st.session_state['batch_settings'] = _prev_settings
                    if _prev_meta:
                        st.session_state['sample_metadata'] = _prev_meta

                    # Store no-signal DataFrame for display
                    if 'No Signal Samples' in _prev_sheets:
                        st.session_state['_no_signal_df'] = _prev_sheets['No Signal Samples']

                    st.session_state['_prev_results_key'] = _prev_key
                    _n_wells = len(_prev_results_list)
                    _n_pass = sum(1 for r in _prev_results_list
                                 if r.get('Success') and str(r.get('Success', '')).startswith('✓'))
                    st.sidebar.success(
                        f"✅ Loaded {_n_wells} wells ({_n_pass} passed) from previous results"
                    )
                    st.rerun()
            except Exception as _e:
                st.sidebar.error(f"Failed to load results: {_e}")

# ── Resume bootstrap: if _resume_checkpoint triggered a rerun, restore
# batch_mode / all_samples / cycles / fluorescence from session state so
# the main content path can reach the batch fit code.
if '_resume_checkpoint' in st.session_state and cycles is None:
    if st.session_state.get('upload_ready_batch', False):
        cycles = st.session_state.get('upload_cycles')
        all_samples = st.session_state.get('upload_batch_samples', {})
        batch_mode = True
        _preview = st.session_state.get('upload_preview_sample')
        if _preview and _preview in all_samples:
            fluorescence = all_samples[_preview]
        elif all_samples:
            fluorescence = next(iter(all_samples.values()))

# Main content
if cycles is not None and fluorescence is not None:
    
    # Show D0 estimation details
    with st.expander("📊 D₀ Bounds Estimation (from Exponential Fits)", expanded=False):
        from mak2_model import estimate_D0_bounds, estimate_MAK2_params_from_exponential

        # Get bounds and estimates with fit info (D0 is now in fluorescence units)
        D0_lower, D0_upper, F_bg_estimate, fit_info = estimate_D0_bounds(
            cycles, fluorescence
        )

        # Display results - top row for D0
        col1, col2, col3 = st.columns(3)
        col1.metric("D₀ Lower Bound", f"{D0_lower:.2e}", help="From perfect doubling fit (2^n) - in fluorescence units")
        col2.metric("D₀ Upper Bound", f"{D0_upper:.2e}",
                   help=f"From efficiency fit (E^n, E={fit_info.get('efficiency', 1.8):.2f}) - in fluorescence units")
        col3.metric("Background Est.", f"{F_bg_estimate:.4f}", help="Estimated from exponential fits")

        # Also get analytical estimates for k and P0
        try:
            analytical_estimates, analytical_bounds = estimate_MAK2_params_from_exponential(
                cycles, fluorescence, P0_assumed=1.0, verbose=False
            )
            k_estimate = analytical_estimates['k']
            P0_estimate = analytical_estimates['P0']

            # Display k and P0 initial guesses in a second row
            st.markdown("**Initial Parameter Guesses:**")
            col4, col5 = st.columns(2)
            col4.metric("k Initial Guess", f"{k_estimate:.4f}", help="Analytical estimate of primer depletion rate from exponential growth")
            col5.metric("P₀ Initial Guess", f"{P0_estimate:.4f}", help="Estimated initial primer concentration (= F_max)")
        except Exception as e:
            st.warning(f"Could not calculate analytical k and P0 estimates: {str(e)}")
        
        # Show exponential phase region and fits
        if fit_info:
            exp_cycles_lower = fit_info['exp_cycles_lower']
            exp_fluor_lower = fit_info['exp_fluorescence_lower']
            fit_lower = fit_info['fit_lower']
            fit_lower_extended = fit_info['fit_lower_extended']  # Extended to efficiency range
            
            exp_cycles_upper = fit_info['exp_cycles_upper']
            exp_fluor_upper = fit_info['exp_fluorescence_upper']
            fit_upper = fit_info['fit_upper']
            
            threshold_cycle = fit_info.get('threshold_cycle', exp_cycles_lower[-1])
            r2_lower = fit_info.get('r2_lower', 0.0)
            r2_upper = fit_info.get('r2_upper', 0.0)
            
            st.info(f"🔬 **Perfect doubling fit**: cycle {exp_cycles_lower[0]:.0f} to {exp_cycles_lower[-1]:.0f}, R² = {r2_lower:.4f}  \n"
                   f"🔬 **Efficiency fit**: cycle {exp_cycles_upper[0]:.0f} to {exp_cycles_upper[-1]:.0f}, R² = {r2_upper:.4f}, E = {fit_info.get('efficiency', 1.8):.2f}")
            
            # Create visualization of exponential fits
            fig_exp = go.Figure()
            
            # Plot all data
            fig_exp.add_trace(go.Scatter(
                x=cycles, y=fluorescence,
                mode='markers',
                name='All Data',
                marker=dict(color='lightgray', size=5, opacity=0.5)
            ))
            
            # Highlight perfect doubling region data points
            fig_exp.add_trace(go.Scatter(
                x=exp_cycles_lower, 
                y=exp_fluor_lower,
                mode='markers',
                name='Perfect Doubling Region',
                marker=dict(color='blue', size=8, symbol='circle')
            ))
            
            # Highlight efficiency region data points
            fig_exp.add_trace(go.Scatter(
                x=exp_cycles_upper, 
                y=exp_fluor_upper,
                mode='markers',
                name='Efficiency Region',
                marker=dict(color='red', size=8, symbol='diamond')
            ))
            
            # Add perfect doubling fitted line (clipped to its own fitted region)
            fig_exp.add_trace(go.Scatter(
                x=exp_cycles_lower,
                y=fit_lower,
                mode='lines',
                name='Perfect Doubling Fit (2^n)',
                line=dict(color='blue', width=3)
            ))
            
            # Add efficiency fitted line
            fig_exp.add_trace(go.Scatter(
                x=exp_cycles_upper,
                y=fit_upper,
                mode='lines',
                name=f'Efficiency Fit (E^n, E={fit_info["efficiency"]:.2f})',
                line=dict(color='red', width=3, dash='dash')
            ))
            
            # Add vertical line at baseline end
            fig_exp.add_vline(
                x=threshold_cycle,
                line_dash="dash",
                line_color="blue",
                annotation_text="Baseline End",
                annotation_position="top right"
            )
            
            # Add background estimate line (sloped if slope is present)
            if 'bg_slope' in fit_info and 'bg_intercept' in fit_info:
                # Plot sloped background line across the data range
                bg_intercept = fit_info['bg_intercept']
                bg_slope = fit_info['bg_slope']
                bg_line_x = np.array([cycles.min(), cycles.max()])
                bg_line_y = bg_intercept + bg_slope * bg_line_x
                
                fig_exp.add_trace(go.Scatter(
                    x=bg_line_x,
                    y=bg_line_y,
                    mode='lines',
                    line=dict(color='purple', dash='dot', width=2),
                    name='Background Est.',
                    showlegend=True
                ))
            else:
                # Fallback to horizontal line if slope not available
                fig_exp.add_hline(
                    y=F_bg_estimate, 
                    line_dash="dot", 
                    line_color="purple",
                    annotation_text="Background Est."
                )
            
            # Add threshold line
            if 'threshold' in fit_info:
                fig_exp.add_hline(
                    y=fit_info['threshold'],
                    line_dash="dot",
                    line_color="orange",
                    annotation_text="Baseline End Level"
                )
            
            # Y-axis range: tight around the data, not forced to zero
            _y_min = fluorescence.min()
            _y_max = fluorescence.max()
            _y_pad = (_y_max - _y_min) * 0.08
            fig_exp.update_layout(
                title="Exponential Phase Detection and Fits for D₀ Estimation",
                xaxis_title="Cycle",
                yaxis_title="Fluorescence",
                yaxis=dict(range=[_y_min - _y_pad, _y_max + _y_pad]),
                height=500,
                showlegend=True,
                legend=dict(
                    yanchor="bottom",
                    y=0.01,
                    xanchor="right",
                    x=0.99,
                    bgcolor="rgba(255,255,255,0.8)",
                    bordercolor="lightgrey",
                    borderwidth=1,
                )
            )
            
            st.plotly_chart(fig_exp, use_container_width=True)
            
            st.caption(f"📈 **Blue** (cycles {exp_cycles_lower[0]:.0f}-{exp_cycles_lower[-1]:.0f}): "
                      f"Perfect doubling fit (2^n) → lower D₀ bound. "
                      f"**Red** (cycles {exp_cycles_upper[0]:.0f}-{exp_cycles_upper[-1]:.0f}): "
                      f"Efficiency fit (E={fit_info['efficiency']:.2f}^n) → upper D₀ bound. "
                      f"Both start at cycle 1, but end at different points based on data characteristics.")
    
    # Fitting options
    st.sidebar.header("Fitting Options")
    
    st.sidebar.info(
        "💡 **Smart Truncation**\n\n"
        "Avoids deep plateau effects (enzyme degradation, dNTP depletion) "
        "not modeled by primer depletion alone."
    )
    
    first_fit_cycle = st.sidebar.slider(
        "First cycle to fit (floor)",
        min_value=1,
        max_value=10,
        value=3,
        step=1,
        help="Hard minimum: never fit cycles before this number. "
             "Cycles 1–2 often show a nonlinear thermal transient. "
             "Default: 3 (matches instrument baseline start)."
    )

    cycles_before_max = st.sidebar.slider(
        "Cycles before max slope (post-baseline)",
        min_value=3,
        max_value=25,
        value=10,
        step=1,
        help="Start fitting this many cycles before the inflection point "
             "(cycle of maximum slope). Mirrors 'Cycles after max slope'. "
             "Smaller = more focus on exponential rise, less baseline. "
             "Default: 8."
    )

    cycles_after_max = st.sidebar.slider(
        "Cycles after max slope",
        min_value=0,
        max_value=10,
        value=4,
        step=1,
        help="Cutoff = cycle at maximum slope + this many cycles. "
             "Default: 4 cycles captures plateau onset."
    )
    auto_truncate = True
    truncate_cycle = None
    custom_bounds_dict = None

    # Copy Number Conversion (sidebar)
    st.sidebar.markdown("---")
    st.sidebar.subheader("Copy Number Conversion")
    calibration_method = st.sidebar.radio(
        "Calibration method",
        ["Auto-detect standards", "Limited dilution", "Manual CF", "Single-copy D0", "None"],
        index=0,
        help="Choose how to convert D0 to copy numbers."
    )
    st.session_state['calibration_method'] = {
        "Auto-detect standards": "auto",
        "Limited dilution": "limited_dilution",
        "Manual CF": "manual_cf",
        "Single-copy D0": "d0_single",
        "None": "none",
    }[calibration_method]

    if calibration_method == "Limited dilution":
        ld_source = st.sidebar.radio(
            "Identify limited dilution wells by:",
            ["Sample Name (from metadata)", "Manual well selection"],
            index=0,
        )
        sample_metadata_sidebar = st.session_state.get('sample_metadata', {})

        if ld_source == "Sample Name (from metadata)":
            if sample_metadata_sidebar:
                unique_names = sorted(set(
                    str(m.get('Sample Name', '')).strip()
                    for m in sample_metadata_sidebar.values()
                    if m.get('Sample Name') and str(m.get('Sample Name')).strip()
                    and str(m.get('Sample Name')).strip() != 'nan'
                ))
                if unique_names:
                    ld_sample_name = st.sidebar.selectbox(
                        "Select limited dilution sample name",
                        unique_names,
                        help="All wells with this Sample Name are limited dilution wells."
                    )
                    ld_wells = [
                        well for well, meta in sample_metadata_sidebar.items()
                        if str(meta.get('Sample Name', '')).strip() == ld_sample_name
                    ]
                    st.sidebar.info(f"Found {len(ld_wells)} wells for '{ld_sample_name}'")
                    st.session_state['ld_wells'] = ld_wells
                else:
                    st.sidebar.warning("No sample names found in metadata.")
                    st.session_state['ld_wells'] = []
            else:
                st.sidebar.warning(
                    "No sample metadata available. "
                    "Upload a QuantStudio file or use manual well selection."
                )
                st.session_state['ld_wells'] = []
        else:  # Manual well selection
            ld_wells_input = st.sidebar.text_area(
                "Enter limited dilution well positions",
                placeholder="A1, A2, A3, B1, B2, B3, ...",
                help="Comma-separated well positions. Include ALL wells "
                     "(both positive and negative)."
            )
            ld_wells = [w.strip() for w in ld_wells_input.split(',') if w.strip()]
            st.session_state['ld_wells'] = ld_wells
            if ld_wells:
                st.sidebar.info(f"{len(ld_wells)} wells entered")

    elif calibration_method == "Manual CF":
        manual_cf = st.sidebar.number_input(
            "Manual conversion factor (copies/D0)",
            min_value=0.0,
            value=0.0,
            format="%.2e",
            help="Enter a known copies-per-D0-unit conversion factor. "
                 "Leave at 0 to skip."
        )
        if manual_cf > 0:
            st.session_state['manual_conversion_factor'] = manual_cf
        else:
            st.session_state.pop('manual_conversion_factor', None)

    elif calibration_method == "Auto-detect standards":
        manual_cf = st.sidebar.number_input(
            "Fallback manual CF (if no standards found)",
            min_value=0.0,
            value=0.0,
            format="%.2e",
            help="Used only if no standard curve is available."
        )
        if manual_cf > 0:
            st.session_state['manual_conversion_factor'] = manual_cf
        else:
            st.session_state.pop('manual_conversion_factor', None)

    elif calibration_method == "Single-copy D0":
        st.session_state.pop('manual_conversion_factor', None)
        _all_targets_sidebar = st.session_state.get('all_targets', [])
        if _all_targets_sidebar:
            st.sidebar.caption(
                "Enter D0_single for each target. Leave at 0 to skip that target."
            )
            _d0s_per_target = {}
            for _tgt in _all_targets_sidebar:
                _prev = st.session_state.get('d0_single_per_target', {}).get(_tgt, 0.0)
                _val = st.sidebar.number_input(
                    f"D0_single — {_tgt}",
                    min_value=0.0,
                    value=float(_prev),
                    format="%.3e",
                    help=f"D0 for a single copy of {_tgt}. Copies = D0 / D0_single.",
                    key=f"d0_single_{_tgt}",
                )
                _d0s_per_target[_tgt] = _val
            st.session_state['d0_single_per_target'] = _d0s_per_target
            st.session_state.pop('d0_single_value', None)
        else:
            _d0s_val = st.sidebar.number_input(
                "D0 for single copy",
                min_value=0.0,
                value=float(st.session_state.get('d0_single_value', 0.0)),
                format="%.3e",
                help="D0 value corresponding to exactly 1 copy. Copies = D0 / D0_single.",
            )
            if _d0s_val > 0:
                st.session_state['d0_single_value'] = _d0s_val
            else:
                st.session_state.pop('d0_single_value', None)
            st.session_state.pop('d0_single_per_target', None)

    else:  # None
        st.session_state.pop('manual_conversion_factor', None)

    # Replicate Analysis Options (only show for batch mode)
    if batch_mode and all_samples:
        with st.sidebar.expander("📊 Replicate Analysis", expanded=False):
            st.markdown("**Group samples into replicates**")

            enable_replicate_analysis = st.checkbox(
                "Enable replicate analysis",
                value=False,
                help="Automatically group samples and calculate statistics"
            )

            if enable_replicate_analysis:
                # Build grouping options — include metadata option if available
                grouping_options = []
                sample_metadata_check = st.session_state.get('sample_metadata')
                has_file_names = (
                    sample_metadata_check is not None and
                    any(
                        m.get('Sample Name') and str(m.get('Sample Name')).strip()
                        and str(m.get('Sample Name')).strip() != 'nan'
                        for m in sample_metadata_check.values()
                    )
                )
                if has_file_names:
                    grouping_options.append("Sample name (from file)")
                grouping_options.extend([
                    "Dot - last (F1.1, F1.2 → F1)",
                    "Dot - first (X1.R1.1, X1.R2.1 → X1)",
                    "Underscore (Sample_A_1, Sample_A_2 → Sample_A)",
                    "Manual grouping (define your own)",
                ])

                grouping_pattern = st.radio(
                    "Group samples by:",
                    grouping_options,
                    index=0,
                    help="How to parse sample names into replicate groups"
                )

                # Preview groups with current samples
                if grouping_pattern == "Sample name (from file)":
                    pattern_key = 'sample_name'
                elif "Dot - last" in grouping_pattern:
                    pattern_key = 'dot'
                elif "Dot - first" in grouping_pattern:
                    pattern_key = 'first_part'
                elif grouping_pattern == "Underscore (Sample_A_1, Sample_A_2 → Sample_A)":
                    pattern_key = 'underscore'
                elif grouping_pattern == "Manual grouping (define your own)":
                    pattern_key = 'manual'
                else:
                    pattern_key = 'dot'

                # Show preview of groups that will be created
                if pattern_key == 'manual':
                    # Manual grouping interface
                    st.markdown("**Define replicate groups manually:**")
                    st.caption("Enter one group per line. Format: GroupName: Sample1, Sample2, Sample3")

                    # Show available samples
                    with st.expander("📋 Available samples", expanded=False):
                        sample_list = list(all_samples.keys())
                        for i, sample in enumerate(sample_list, 1):
                            st.write(f"{i}. {sample}")

                    # Text area for manual grouping
                    manual_groups_text = st.text_area(
                        "Define groups (one per line):",
                        value=st.session_state.get('manual_groups_text', ''),
                        height=200,
                        placeholder="Example:\nGroup1: Sample_1, Sample_2, Sample_3\nGroup2: Sample_4, Sample_5, Sample_6",
                        help="Each line defines one group. Use format 'GroupName: sample1, sample2, sample3'"
                    )

                    st.session_state['manual_groups_text'] = manual_groups_text

                    # Parse manual groups
                    preview_groups = {}
                    if manual_groups_text.strip():
                        for line in manual_groups_text.strip().split('\n'):
                            if ':' in line:
                                group_name, samples_str = line.split(':', 1)
                                group_name = group_name.strip()
                                samples = [s.strip() for s in samples_str.split(',')]
                                # Validate that samples exist
                                valid_samples = [s for s in samples if s in all_samples]
                                if valid_samples:
                                    preview_groups[group_name] = valid_samples
                                else:
                                    st.warning(f"⚠️ Group '{group_name}' has no valid samples")
                elif pattern_key == 'sample_name' and sample_metadata_check:
                    # Group by Sample Name from file metadata
                    preview_groups = {}
                    has_mt = any('::' in k for k in all_samples.keys())
                    for key, meta in sample_metadata_check.items():
                        sname = meta.get('Sample Name')
                        if sname and str(sname) != 'nan' and str(sname).strip():
                            sname = str(sname).strip()
                            if has_mt and meta.get('_target'):
                                group_label = f"{meta['_target']} — {sname}"
                            else:
                                group_label = sname
                            if key in all_samples:
                                if group_label not in preview_groups:
                                    preview_groups[group_label] = []
                                preview_groups[group_label].append(key)
                    # Keep only groups with > 1 replicate
                    preview_groups = {k: v for k, v in preview_groups.items() if len(v) > 1}
                else:
                    preview_groups = parse_sample_groups(list(all_samples.keys()), pattern=pattern_key)

                if len(preview_groups) > 0:
                    st.success(f"✅ Found {len(preview_groups)} replicate groups")
                    with st.expander("Preview groups", expanded=False):
                        for group, samples in list(preview_groups.items())[:5]:  # Show first 5
                            st.write(f"**{group}:** {', '.join(samples)}")
                        if len(preview_groups) > 5:
                            st.write(f"... and {len(preview_groups) - 5} more groups")
                else:
                    st.warning("⚠️ No replicate groups found with this pattern")

                # Dilution series options
                analyze_as_dilution = False
                dilution_info = None

                if len(preview_groups) >= 3:
                    analyze_as_dilution = st.checkbox(
                        "Analyze as dilution series",
                        value=False,
                        help="Check linearity and efficiency for dilution series"
                    )

                    if analyze_as_dilution:
                        dilution_info = st.selectbox(
                            "Dilution type:",
                            ["2-fold serial dilutions",
                             "10-fold serial dilutions",
                             "5-fold serial dilutions",
                             "Custom dilution factors"],
                            help="How are your samples diluted?"
                        )

                        # Determine dilution factor
                        if "2-fold" in dilution_info:
                            dilution_factor = 2
                            custom_dilution_factors = None
                        elif "10-fold" in dilution_info:
                            dilution_factor = 10
                            custom_dilution_factors = None
                        elif "5-fold" in dilution_info:
                            dilution_factor = 5
                            custom_dilution_factors = None
                        else:
                            # Custom dilution factors
                            dilution_factor = None
                            st.markdown("**Define dilution factors for each group:**")
                            st.caption("Enter one line per group. Format: GroupName: dilution_factor")
                            st.caption("Example: if Group1 is undiluted (1×), Group2 is 10× diluted, Group3 is 100× diluted")

                            custom_dilution_text = st.text_area(
                                "Define dilution factors:",
                                value=st.session_state.get('custom_dilution_text', ''),
                                height=150,
                                placeholder="Example:\nGroup1: 1\nGroup2: 10\nGroup3: 100\nGroup4: 1000",
                                help="Dilution factor = how many times diluted (1 = undiluted, 10 = 10× diluted, etc.)"
                            )

                            st.session_state['custom_dilution_text'] = custom_dilution_text

                            # Parse custom dilution factors
                            custom_dilution_factors = {}
                            if custom_dilution_text.strip():
                                for line in custom_dilution_text.strip().split('\n'):
                                    if ':' in line:
                                        group_name, factor_str = line.split(':', 1)
                                        group_name = group_name.strip()
                                        try:
                                            factor = float(factor_str.strip())
                                            if group_name in preview_groups:
                                                custom_dilution_factors[group_name] = factor
                                            else:
                                                st.warning(f"⚠️ Group '{group_name}' not found in replicate groups")
                                        except ValueError:
                                            st.warning(f"⚠️ Invalid dilution factor for '{group_name}': {factor_str}")

                        st.session_state['dilution_factor'] = dilution_factor
                        st.session_state['custom_dilution_factors'] = custom_dilution_factors if dilution_factor is None else None

                        # Option to exclude groups from dilution series
                        st.markdown("**Exclude groups from dilution series (optional):**")
                        st.caption("Select groups to exclude (e.g., most diluted samples with poor amplification)")

                        groups_to_exclude = st.multiselect(
                            "Exclude these groups:",
                            options=list(preview_groups.keys()),
                            default=[],
                            help="These groups will be excluded from dilution series analysis"
                        )

                        st.session_state['exclude_from_dilution'] = groups_to_exclude

                # Store settings in session state
                st.session_state['replicate_analysis_enabled'] = True
                st.session_state['grouping_pattern'] = pattern_key
                st.session_state['preview_groups'] = preview_groups
                st.session_state['analyze_as_dilution'] = analyze_as_dilution
                st.session_state['dilution_info'] = dilution_info
            else:
                st.session_state['replicate_analysis_enabled'] = False

    # Fit button(s)
    if batch_mode and all_samples:
        # Batch mode - fit all samples
        _fit_btn = st.sidebar.button("🔬 Batch Fit All Samples", type="primary")
        _resume_pending = '_resume_checkpoint' in st.session_state
        if _fit_btn or _resume_pending:
          _is_resuming = _resume_pending
          _resume_cp = st.session_state.pop('_resume_checkpoint', None) if _is_resuming else None
          try:
            # Clear any previous manual fit results
            if 'fitted_params' in st.session_state:
                del st.session_state['fitted_params']
            if 'optimizer' in st.session_state:
                del st.session_state['optimizer']

            st.subheader("🔄 Batch Fitting Results")

            # Capture stdout from optimizer for diagnostic display
            import io as _io_batch
            _batch_captured = _io_batch.StringIO()
            _batch_old_stdout = sys.stdout
            sys.stdout = _batch_captured

            # ── Resume from checkpoint: restore all state ────────────────────────
            if _is_resuming and _resume_cp is not None:
                from collections import OrderedDict as _OD
                _cp_sample_names = _resume_cp['sample_names']
                all_samples_to_fit = _OD(
                    (k, np.array(_resume_cp['all_samples'][k]))
                    for k in _cp_sample_names
                )
                no_signal_samples = _resume_cp.get('no_signal_samples', {})
                cycles = np.array(_resume_cp['cycles'])
                _cp_settings = _resume_cp['settings']
                first_fit_cycle    = _cp_settings.get('first_fit_cycle', 1)
                cycles_before_max  = _cp_settings.get('cycles_before_max', 3)
                cycles_after_max   = _cp_settings.get('cycles_after_max', 3)
                auto_truncate      = _cp_settings.get('auto_truncate', True)
                truncate_cycle     = _cp_settings.get('truncate_cycle', None)
                custom_bounds_dict = _cp_settings.get('custom_bounds_dict', {})
                global_threshold      = _cp_settings.get('global_threshold')
                global_baseline_mean  = _cp_settings.get('global_baseline_mean', 0.0)
                channel_thresholds    = _cp_settings.get('channel_thresholds', {})
                channel_baseline_means = _cp_settings.get('channel_baseline_means', {})
                sample_metadata = _resume_cp.get('sample_metadata')
                results_list = _resume_cp['results_list']
                _pass1_start_idx = _resume_cp['pass1_completed_count']
                _pass2_start_set = set(_resume_cp.get('pass2_completed_indices', []))
                _cp_current_pass = _resume_cp.get('current_pass', 'pass1')
                progress_bar = st.progress(_pass1_start_idx / len(all_samples_to_fit))
                status_text = st.empty()
                if _cp_current_pass == 'pass1':
                    st.info(f"▶️ Resuming Pass 1 from well {_pass1_start_idx + 1}/{len(all_samples_to_fit)}")
                else:
                    st.info(f"▶️ Resuming Pass 2 retries")

                # Restore ROX data if available
                if _resume_cp.get('rox_by_well'):
                    st.session_state['rox_by_well'] = {
                        k: np.array(v) for k, v in _resume_cp['rox_by_well'].items()
                    }
                if _resume_cp.get('rox_normalized') is not None:
                    st.session_state['rox_normalized'] = _resume_cp['rox_normalized']
                if sample_metadata:
                    st.session_state['sample_metadata'] = sample_metadata

            else:
                # Fresh run — normal signal detection + threshold computation
                _pass1_start_idx = 0
                _pass2_start_set = set()
                _cp_current_pass = 'pass1'

            # === SIGNAL DETECTION + THRESHOLD COMPUTATION ===
            # (skipped on resume — restored from checkpoint above)
            if not (_is_resuming and _resume_cp is not None):
                from data_processing import detect_no_signal_samples

                with st.status("🔍 Detecting samples with no signal...", expanded=True) as status:
                    st.write(f"Analyzing {len(all_samples)} samples...")

                    # Run signal detection PER CHANNEL so each channel's
                    # range threshold is relative to its own max, not the
                    # plate-wide max.  This prevents e.g. CY3 wells (lower
                    # absolute fluorescence) from being rejected because FAM
                    # has a much larger absolute range.
                    _sel_channels = st.session_state.get('selected_channels', [])
                    if _sel_channels and len(_sel_channels) > 1:
                        valid_samples = {}
                        no_signal_samples = {}
                        plate_stats = {'max_range': 0}
                        for _det_ch in _sel_channels:
                            _det_ch_samples = {
                                name: fluor for name, fluor in all_samples.items()
                                if _ch(name) == _det_ch
                            }
                            if not _det_ch_samples:
                                continue
                            _v, _ns, _ps = detect_no_signal_samples(
                                cycles, _det_ch_samples,
                                min_range_pct=2.0, min_r2=0.80, verbose=False
                            )
                            valid_samples.update(_v)
                            no_signal_samples.update(_ns)
                            plate_stats['max_range'] = max(
                                plate_stats['max_range'], _ps.get('max_range', 0)
                            )
                        st.write(
                            f"Per-channel signal detection: "
                            f"{', '.join(f'{ch}: {sum(1 for n in valid_samples if _ch(n)==ch)}' for ch in _sel_channels)}"
                        )
                    else:
                        valid_samples, no_signal_samples, plate_stats = detect_no_signal_samples(
                            cycles, all_samples,
                            min_range_pct=2.0, min_r2=0.80, verbose=False
                        )

                    # ── Reconcile with instrument metadata (highest authority) ─────────────
                    _smeta = st.session_state.get('sample_metadata', {})
                    if _smeta:
                        rescued, suppressed = [], []
                        for sname, info in list(no_signal_samples.items()):
                            wm = _smeta.get(sname, {})
                            if 'Ct_instrument' not in wm:
                                continue
                            inst_ct = wm.get('Ct_instrument')
                            inst_noamp = wm.get('NOAMP', False)
                            inst_undetermined = (
                                inst_ct is None
                                or (isinstance(inst_ct, float) and np.isnan(inst_ct))
                            )
                            if not inst_undetermined and not inst_noamp:
                                valid_samples[sname] = all_samples[sname]
                                del no_signal_samples[sname]
                                rescued.append(sname)
                        if rescued:
                            st.write(f"🔁 Rescued {len(rescued)} wells overridden by instrument metadata: "
                                     f"{', '.join(rescued[:5])}{'…' if len(rescued) > 5 else ''}")

                    st.write(f"✅ Found {len(valid_samples)} samples with valid signal")
                    if no_signal_samples:
                        st.write(f"⚠️ Flagged {len(no_signal_samples)} samples with no detectable signal")

                    status.update(label="✅ Signal detection complete", state="complete")

                # Show flagged samples if any
                if no_signal_samples:
                    with st.expander(f"⚠️ {len(no_signal_samples)} samples flagged (will be skipped)", expanded=True):
                        st.info("These samples have insufficient signal for MAK2+ fitting (likely NTC, failed reactions, or no template).")

                        no_signal_df = pd.DataFrame([
                            {
                                'Sample': name,
                                'Reason': info['reason'],
                                'Fluorescence Range': f"{info['F_range']:.4f}",
                                '% of Max on Plate': f"{info['F_range_pct']:.1f}%"
                            }
                            for name, info in no_signal_samples.items()
                        ])
                        st.dataframe(no_signal_df, use_container_width=True)
                        st.session_state['_no_signal_df'] = no_signal_df

                # Update all_samples to only include valid samples
                all_samples_to_fit = valid_samples

                # Save no-signal fluor data NOW while all_samples is guaranteed in scope
                st.session_state['batch_no_signal_fluor'] = {
                    name: all_samples[name] for name in no_signal_samples if name in all_samples
                }

                results_list = []
                progress_bar = st.progress(0)
                status_text = st.empty()

                # ── Per-channel threshold computation ────────────────────────────────
                status_text.text("Calculating per-channel thresholds for Ct analysis...")
                global_threshold     = None
                global_baseline_mean = 0.0
                channel_thresholds      = {}
                channel_baseline_means  = {}

                all_fluorescence = list(all_samples_to_fit.values())

                if all_fluorescence:
                    baseline_end = max(3, int(len(cycles) * 0.15))

                    ch_arrays = {}
                    for sname, fd in all_samples_to_fit.items():
                        ch_arrays.setdefault(_ch(sname), []).append(fd)

                    for ch, arrays in ch_arrays.items():
                        ch_sds    = []
                        ch_means  = []
                        ch_early_all = []
                        for fd in arrays:
                            early = fd[:baseline_end]
                            ch_early_all.extend(early)
                            ch_sds.append(np.std(early))
                            ch_means.append(np.mean(early))

                        ch_sub = (np.mean(ch_early_all) < 0.1) or \
                                 np.any(np.array(ch_early_all) < 0)
                        ch_baseline_mean = 0.0 if ch_sub else np.median(ch_means)
                        ch_median_sd     = np.median(ch_sds)
                        ch_max           = np.max([np.max(fd) for fd in arrays])
                        ch_dyn_range     = ch_max - ch_baseline_mean

                        ch_thresh = max(
                            10 * ch_median_sd if ch_median_sd > 0 else 0.01,
                            0.05 * ch_dyn_range
                        )
                        channel_thresholds[ch]     = ch_thresh
                        channel_baseline_means[ch] = ch_baseline_mean

                    if len(channel_thresholds) == 1:
                        global_threshold     = next(iter(channel_thresholds.values()))
                        global_baseline_mean = next(iter(channel_baseline_means.values()))
                    else:
                        all_sds = [s for arrays in ch_arrays.values()
                                   for fd in arrays
                                   for s in [np.std(fd[:baseline_end])]]
                        all_means = [np.mean(fd[:baseline_end])
                                     for arrays in ch_arrays.values()
                                     for fd in arrays]
                        already_sub = (np.mean([v for fd in all_fluorescence
                                                for v in fd[:baseline_end]]) < 0.1)
                        global_baseline_mean = 0.0 if already_sub else np.median(all_means)
                        plate_max     = np.max([np.max(f) for f in all_fluorescence])
                        global_threshold = max(
                            10 * np.median(all_sds) if all_sds else 0.01,
                            0.05 * (plate_max - global_baseline_mean)
                        )

                    if len(channel_thresholds) > 1:
                        lines = ', '.join(
                            f'{c}: {v:,.4g}' for c, v in sorted(channel_thresholds.items())
                        )
                        st.info(f"📊 Per-channel thresholds (above baseline): {lines}")
                    else:
                        st.info(
                            f"📊 Using threshold: {global_threshold:,.4g} "
                            f"(from {len(all_samples_to_fit)} samples)"
                        )

                # Override thresholds with instrument values when metadata is loaded.
                abi_results_meta = st.session_state.get('abi_results_meta')
                rox_norm_active  = st.session_state.get('rox_normalized', False)
                if abi_results_meta and rox_norm_active:
                    inst_thresholds = abi_results_meta.get('channel_thresholds', {})
                    if inst_thresholds:
                        for ch, val in inst_thresholds.items():
                            if ch in channel_thresholds:
                                channel_thresholds[ch] = val
                        if len(inst_thresholds) == 1:
                            global_threshold = next(iter(inst_thresholds.values()))
                        lines = ', '.join(
                            f'{c}={v}' for c, v in sorted(inst_thresholds.items())
                        )
                        st.success(f"✅ Using instrument Ct thresholds (ΔRn): {lines}")

                # Get sample metadata for instrument CT values
                sample_metadata = st.session_state.get('sample_metadata')

            # ── Build checkpoint object ──────────────────────────────────────────
            from datetime import datetime as _dt
            _checkpoint = {
                'version': 1,
                'created_at': _resume_cp['created_at'] if (_is_resuming and _resume_cp) else _dt.utcnow().isoformat(),
                'updated_at': _dt.utcnow().isoformat(),
                'current_pass': _cp_current_pass,
                'pass1_completed_count': _pass1_start_idx,
                'pass1_total': len(all_samples_to_fit),
                'pass2_completed_indices': list(_pass2_start_set),
                'retry_indices': _resume_cp.get('retry_indices', []) if (_is_resuming and _resume_cp) else [],
                'results_list': results_list,
                'sample_names': list(all_samples_to_fit.keys()),
                'all_samples': {k: v.tolist() if hasattr(v, 'tolist') else v
                                for k, v in all_samples_to_fit.items()},
                'no_signal_samples': no_signal_samples,
                'cycles': cycles.tolist() if hasattr(cycles, 'tolist') else cycles,
                'settings': {
                    'first_fit_cycle': first_fit_cycle,
                    'cycles_before_max': cycles_before_max,
                    'cycles_after_max': cycles_after_max,
                    'auto_truncate': auto_truncate,
                    'truncate_cycle': truncate_cycle,
                    'custom_bounds_dict': custom_bounds_dict,
                    'global_threshold': global_threshold,
                    'global_baseline_mean': global_baseline_mean,
                    'channel_thresholds': channel_thresholds,
                    'channel_baseline_means': channel_baseline_means,
                },
                'sample_metadata': st.session_state.get('sample_metadata'),
                'rox_by_well': {k: v.tolist() if hasattr(v, 'tolist') else v
                                for k, v in st.session_state.get('rox_by_well', {}).items()},
                'rox_normalized': st.session_state.get('rox_normalized', False),
            }

            # Pass 1: Fit all samples normally
            # (if resuming into Pass 2, skip Pass 1 entirely)
            _skip_pass1 = (_is_resuming and _resume_cp is not None
                           and _cp_current_pass == 'pass2')
            for i, (sample_name, fluor_data) in enumerate(all_samples_to_fit.items()):
                if _skip_pass1 or i < _pass1_start_idx:
                    continue  # already completed in previous session
                status_text.text(f"Pass 1: Fitting {sample_name}... ({i+1}/{len(all_samples_to_fit)})")

                # Phase-1 unified per-well pipeline. Inputs assembled from the
                # batch-mode globals; everything else (no-amp pre-check, take-off
                # detection, bg pre-estimation, optimizer, toe stages, Ct,
                # tier, instrument status) lives inside fit_well so all callers
                # (this app, run_batch, the PCRedux scoring driver) share one
                # implementation.
                from fit_well import fit_well as _fit_well_v1
                wm = sample_metadata.get(sample_name, {}) if sample_metadata else None
                _ch_thresh = channel_thresholds.get(_ch(sample_name), global_threshold)
                _well_pos  = _get_well_pos(sample_name)
                _rox       = st.session_state.get('rox_by_well', {}).get(_well_pos)
                _result = _fit_well_v1(
                    cycles, fluor_data,
                    first_fit_cycle=first_fit_cycle,
                    cycles_before_max=cycles_before_max,
                    cycles_after_max=cycles_after_max,
                    auto_truncate=auto_truncate,
                    truncate_cycle=truncate_cycle,
                    fit_bounds=dict(custom_bounds_dict) if custom_bounds_dict else None,
                    sample_name=sample_name,
                    metadata=wm,
                    rox=_rox,
                    channel_threshold=_ch_thresh,
                )
                results_list.append(_result)
                
                progress_bar.progress((i + 1) / len(all_samples_to_fit))

                # ── Checkpoint: save progress after each well ────────────────────
                _checkpoint['pass1_completed_count'] = i + 1
                _checkpoint['results_list'] = results_list
                _checkpoint['updated_at'] = _dt.utcnow().isoformat()
                _save_checkpoint(_checkpoint)

            # ── Mark Pass 1 complete in checkpoint ────────────────────────────────
            _checkpoint['current_pass'] = 'pass2'
            _checkpoint['results_list'] = results_list
            _save_checkpoint(_checkpoint)

            # ── Pass 2: Channel-aware retry ─────────────────────────────────────────
            # Strategy: learn per-channel priors from reliable pass-1 fits, then
            # retry every sample that is: (a) high SSR, (b) errored, or
            # (c) has a degenerate k > 0.5 (unrealistic for real qPCR).
            #
            # Key fix vs. the old pass-2:
            #  • F_bg_slope bounds are now data-scale-aware (raw ABI slopes ~2000-5000,
            #    not ±0.1 which was 50,000× too narrow and silently broke every retry).
            #  • Statistics are computed per channel (FAM / JOE) so priors are not
            #    contaminated by mixing channels with very different backgrounds.
            #  • Error cases are included in the retry, not silently skipped.

            # _ch() is defined above the threshold block and reused here.

            # Channel-prior stats + retry-candidate identification — shared
            # canonical implementation in pass2_helpers, used by both this
            # app and run_batch.run_pass2.
            from pass2_helpers import compute_channel_priors, identify_retry_candidates
            channel_medians, plate_medians = compute_channel_priors(results_list)
            retry_indices, _skip_count = identify_retry_candidates(
                results_list, cycles, cycles_after_max,
            )
            if _skip_count > 0:
                status_text.text(f"Skipped {_skip_count} wells below R² 0.85 threshold")
            # ``all_k_vals`` is still referenced by the gating condition just
            # below ("if retry_indices and (channel_medians or all_k_vals):"),
            # so synthesise an equivalent from the channel medians.
            all_k_vals = [m['k'] for m in channel_medians.values()]

            # Save retry indices to checkpoint
            _checkpoint['retry_indices'] = retry_indices
            _save_checkpoint(_checkpoint)

            if retry_indices and (channel_medians or all_k_vals):
                status_text.text(
                    f"Pass 2: Retrying {len(retry_indices)} samples with "
                    f"channel-informed priors ({len(channel_medians)} channel(s) learned)…"
                )

                import time as _time_mod
                for idx in retry_indices:
                    # Skip retries already completed in a previous session
                    if idx in _pass2_start_set:
                        continue
                    _retry_t0   = _time_mod.perf_counter()
                    result      = results_list[idx]
                    # Detect late amplifiers early: if pass-1 fit_end is at/near
                    # the last cycle, use a shorter timeout since 0.995 is acceptable.
                    _pass1_late = (
                        result.get('fit_end_cycle') is not None
                        and result['fit_end_cycle'] >= float(cycles[-1]) - min(max(1, cycles_after_max), 5)
                    )
                    _RETRY_TIMEOUT = 10.0 if _pass1_late else 30.0
                    sample_name = result['Sample']
                    fluor_data  = result['fluor_data']
                    if fluor_data is None:
                        continue

                    ch     = _ch(sample_name)
                    priors = channel_medians.get(ch, plate_medians)

                    pk     = priors['k']
                    pP0    = priors['P0']
                    pFbg   = priors['F_bg_intercept']
                    pSlope = priors['F_bg_slope']

                    # Use the per-well pre-estimated background when available.
                    # This is more reliable than channel medians for background
                    # bounds because it comes from the actual baseline region of
                    # THIS well, not an average of other wells.
                    bg_slope_est_r     = result.get('bg_slope_est')
                    bg_intercept_est_r = result.get('bg_intercept_est')

                    status_text.text(
                        f"Pass 2: {sample_name}  ch={ch}  "
                        f"k≈{pk:.3f}  P0≈{pP0:.2g}"
                        + (f"  bg_slope≈{bg_slope_est_r:.1f}" if bg_slope_est_r is not None else "")
                    )

                    try:
                        model_retry     = MAK2Model()
                        optimizer_retry = MAK2Optimizer(model_retry)

                        F_max     = float(np.max(fluor_data))
                        F_range_r = float(np.max(fluor_data) - np.min(fluor_data))

                        # Background bounds: prefer per-well estimate (tight ±40%),
                        # fall back to channel-median prior (loose ±300%) so that
                        # wells without metadata still benefit from the retry.
                        if bg_slope_est_r is not None and bg_intercept_est_r is not None:
                            slope_margin = max(abs(bg_slope_est_r) * 0.40,
                                               F_range_r * 0.01)
                            int_margin   = max(abs(bg_intercept_est_r) * 0.15,
                                               F_max * 0.02)
                            bg_slope_bounds = (bg_slope_est_r - slope_margin,
                                               bg_slope_est_r + slope_margin)
                            bg_int_bounds   = (max(0.0, bg_intercept_est_r - int_margin),
                                               bg_intercept_est_r + int_margin)
                        else:
                            slope_delta   = max(abs(pSlope) * 3.0, F_range_r * 0.05)
                            bg_slope_bounds = (pSlope - slope_delta, pSlope + slope_delta)
                            fbg_lo = max(0.0, pFbg * 0.30)
                            fbg_hi = pFbg * 3.0 if pFbg > 0 else F_max
                            bg_int_bounds = (fbg_lo, fbg_hi)

                        informed_bounds = {
                            'k':              (max(0.01, pk * 0.20), min(1.0, max(0.5, pk * 5.0))),
                            'P0':             (max(pP0 * 0.05, F_range_r * 0.01), max(pP0 * 7.0, F_range_r * 2.0)),
                            'D0':             (1e-15, F_range_r * 10),
                            'F_bg_intercept': bg_int_bounds,
                            'F_bg_slope':     bg_slope_bounds,
                        }

                        # Smart start (no guard — background pre-estimated from window)
                        _r_floor        = int(np.searchsorted(cycles, first_fit_cycle))

                        # Try to get baseline end from metadata (same source as pass-1)
                        _r_meta_bl_end = None
                        _r_meta_bl_end_cycle = None
                        if sample_metadata:
                            _r_wm = sample_metadata.get(sample_name, {})
                            _r_meta_bl_end_val = _r_wm.get('Baseline End', None)
                            if _r_meta_bl_end_val is not None:
                                try:
                                    _r_meta_bl_end = int(np.searchsorted(cycles, float(_r_meta_bl_end_val)))
                                    _r_meta_bl_end_cycle = float(_r_meta_bl_end_val)
                                except (ValueError, TypeError):
                                    pass

                        # Always run estimate_baseline_end() — on Rn if ROX available, else raw
                        from data_processing import estimate_baseline_end
                        _r_rox_arr_bl = st.session_state.get('rox_by_well', {}).get(_get_well_pos(sample_name), None)
                        if _r_rox_arr_bl is not None and len(_r_rox_arr_bl) == len(fluor_data):
                            _r_fluor_for_bl = fluor_data / np.maximum(_r_rox_arr_bl, 1e-10)
                        else:
                            _r_fluor_for_bl = fluor_data
                        _r_est_bl_end_idx   = estimate_baseline_end(cycles, _r_fluor_for_bl, first_cycle_idx=_r_floor)
                        _r_est_bl_end_cycle = float(cycles[min(_r_est_bl_end_idx, len(cycles) - 1)])

                        # Same ceiling-check as pass-1: only use estimated end when
                        # it is reliably before the end of run.
                        _r_LATE_CEIL = int(len(cycles) * 0.85)
                        if _r_meta_bl_end is not None:
                            if _r_est_bl_end_idx < _r_LATE_CEIL:
                                _r_baseline_end_idx = max(_r_meta_bl_end, _r_est_bl_end_idx)
                            else:
                                _r_baseline_end_idx = _r_meta_bl_end
                        else:
                            _r_baseline_end_idx = _r_est_bl_end_idx

                        # Same right-to-left inflection search as pass-1.
                        _r_full_seg = fluor_data[_r_floor:]
                        if len(_r_full_seg) >= 5:
                            _r_raw_g  = np.gradient(_r_full_seg)
                            _r_kern   = np.ones(5) / 5.0
                            _r_smooth_g = np.convolve(_r_raw_g, _r_kern, mode='same')
                            _r_grad_range = float(np.max(_r_smooth_g) - np.min(_r_smooth_g))
                            _r_grad_floor = _r_grad_range * 0.05
                            _r_best_val = _r_smooth_g[-1]
                            _r_best_idx = len(_r_smooth_g) - 1
                            _r_found_peak = False
                            for _rj in range(len(_r_smooth_g) - 2, -1, -1):
                                if _r_smooth_g[_rj] > _r_best_val:
                                    _r_best_val = _r_smooth_g[_rj]
                                    _r_best_idx = _rj
                                elif (_r_best_val > _r_grad_floor
                                      and _r_smooth_g[_rj] < _r_best_val * 0.5):
                                    _r_found_peak = True
                                    break
                            _r_ms_offset = _r_best_idx if _r_found_peak else int(np.argmax(_r_smooth_g))
                        elif len(_r_full_seg) >= 2:
                            _r_ms_offset = int(np.argmax(np.gradient(_r_full_seg)))
                        else:
                            _r_ms_offset = 0
                        _r_max_slope_idx = _r_floor + _r_ms_offset
                        retry_start_idx  = max(_r_floor, _r_max_slope_idx - cycles_before_max)

                        # Background: purely pre-window regression
                        _r_bg_pre_start = max(_r_floor, retry_start_idx - 8)
                        _r_bg_c = cycles[_r_bg_pre_start:retry_start_idx]
                        _r_bg_f = fluor_data[_r_bg_pre_start:retry_start_idx]

                        if len(_r_bg_c) >= 2:
                            _r_bg_coeffs    = np.polyfit(_r_bg_c, _r_bg_f, 1)
                            _r_bg_slope_win = float(_r_bg_coeffs[0])
                            _r_bg_int_win   = float(_r_bg_coeffs[1])
                        else:
                            _r_bg_slope_win = 0.0
                            _r_bg_int_win   = float(fluor_data[retry_start_idx]) if retry_start_idx < len(fluor_data) else 0.0

                        # Adaptive extension: extend backwards until we include
                        # at least 3 cycles of baseline.
                        _r_bl_noise = float(np.std(_r_bg_f)) if len(_r_bg_f) >= 2 else 0.0
                        # Skip adaptive extension for background-subtracted data
                        # (bl_noise ≈ 0) — extra zero cycles hurt more than help.
                        _r_bg_mean = float(np.mean(np.abs(_r_bg_f))) if len(_r_bg_f) >= 2 else 0.0
                        _r_skip_ext = (_r_bl_noise < 1e-6 and _r_bg_mean < 1e-6)
                        _r_n_bl = 0
                        if not _r_skip_ext:
                            for _rbi in range(retry_start_idx, min(retry_start_idx + 6, len(cycles))):
                                _r_bg_lev = _r_bg_slope_win * cycles[_rbi] + _r_bg_int_win
                                if fluor_data[_rbi] <= _r_bg_lev + 3.0 * _r_bl_noise:
                                    _r_n_bl += 1
                        else:
                            _r_n_bl = 3  # pretend we have enough baseline
                        if _r_n_bl < 3:
                            _r_try = cycles_before_max
                            while _r_n_bl < 3 and _r_try < _r_max_slope_idx - _r_floor:
                                _r_try += 2
                                _r_try_start = max(_r_floor, _r_max_slope_idx - _r_try)
                                if _r_try_start == retry_start_idx:
                                    break
                                _r_n_bl = 0
                                for _rbi in range(_r_try_start, min(_r_try_start + 6, len(cycles))):
                                    _r_bg_lev = _r_bg_slope_win * cycles[_rbi] + _r_bg_int_win
                                    if fluor_data[_rbi] <= _r_bg_lev + 3.0 * _r_bl_noise:
                                        _r_n_bl += 1
                                retry_start_idx = _r_try_start
                            _r_bg_pre_start = max(_r_floor, retry_start_idx - 8)
                            _r_bg_c = cycles[_r_bg_pre_start:retry_start_idx]
                            _r_bg_f = fluor_data[_r_bg_pre_start:retry_start_idx]
                            if len(_r_bg_c) >= 2:
                                _r_bg_coeffs    = np.polyfit(_r_bg_c, _r_bg_f, 1)
                                _r_bg_slope_win = float(_r_bg_coeffs[0])
                                _r_bg_int_win   = float(_r_bg_coeffs[1])

                        cycles_retry    = cycles[retry_start_idx:]
                        fluor_retry     = fluor_data[retry_start_idx:]

                        _r_sm          = max(abs(_r_bg_slope_win) * 0.40, F_range_r * 0.002)
                        _r_s_min       = _r_bg_slope_win - _r_sm
                        _r_s_max       = _r_bg_slope_win + _r_sm
                        # Symmetric intercept bounds centered on regression intercept
                        _r_int_delta   = max(abs(_r_bg_int_win) * 0.005, F_range_r * 0.03)
                        _r_int_lo      = _r_bg_int_win - _r_int_delta
                        _r_int_hi      = _r_bg_int_win + _r_int_delta
                        informed_bounds['D0']             = (1e-8, max(F_range_r, 1.0))
                        informed_bounds['F_bg_slope']     = (_r_s_min, _r_s_max)
                        informed_bounds['F_bg_intercept'] = (_r_int_lo, _r_int_hi)

                        # Late-amp enhancement: when pass-1 failed badly and the
                        # well is a late amplifier, run analytical exponential
                        # phase detection to get tight D0/k priors.
                        _pass1_r2 = result.get('R2')
                        _pass1_failed = (_pass1_r2 is None
                                         or (isinstance(_pass1_r2, float) and
                                             (np.isnan(_pass1_r2) or _pass1_r2 < 0.90)))
                        if _pass1_late and _pass1_failed:
                            try:
                                from mak2_model import estimate_MAK2_params_from_exponential
                                _la_est, _la_bounds = estimate_MAK2_params_from_exponential(
                                    cycles_retry, fluor_retry,
                                    P0_assumed=pP0 if pP0 > 0 else 1.0,
                                    verbose=False,
                                )
                                if 'D0' in _la_bounds:
                                    informed_bounds['D0'] = _la_bounds['D0']
                                if 'k' in _la_bounds:
                                    _la_k_lo = max(0.01, _la_bounds['k'][0] * 0.5)
                                    _la_k_hi = min(1.2, _la_bounds['k'][1] * 2.0)
                                    informed_bounds['k'] = (_la_k_lo, _la_k_hi)
                            except Exception:
                                pass  # fall through to existing retry logic

                        # Retry uses extended truncation (+7 cycles after max slope)
                        # so the optimizer sees the post-truncation data that constrains k.
                        # Without this, wells that overshoot at the end of the pass-1
                        # window (model above data → negative residuals → high SSR → retry
                        # triggered) still converge to the same local minimum because
                        # lowering k makes the overshoot worse within the same window.
                        # The extra 7 cycles show the continuing rise / plateau onset,
                        # allowing the optimizer to find the correct k.
                        _retry_cam = cycles_after_max + 3
                        params_retry  = optimizer_retry.fit(
                            cycles_retry, fluor_retry,
                            cycles_after_max=_retry_cam,
                            auto_truncate=auto_truncate,
                            truncate_cycle=truncate_cycle,
                            bounds=informed_bounds,
                            fixed_background_values={
                                'F_bg_slope': _r_bg_slope_win,
                                'F_bg_intercept': _r_bg_int_win,
                            },
                            verbose=False,
                        )
                        _fit_end_cycle_r = float(optimizer_retry.cycles_fit[-1]) \
                            if optimizer_retry.cycles_fit is not None and len(optimizer_retry.cycles_fit) > 0 \
                            else float(cycles[-1])
                        metrics_retry = optimizer_retry.calculate_fit_metrics()

                        # Always target 0.999 during retries — only relax to
                        # 0.995 for late amplifiers at the final acceptance step
                        # (after all retry attempts including extended baseline).
                        _retry_is_late = (_fit_end_cycle_r >= _last_cyc - min(max(1, cycles_after_max), 5))
                        _r2_target = 0.999

                        # Recalculate Ct — reuse the metadata baseline window
                        ct_retry              = np.nan
                        ct_bl_slope_retry     = 0.0
                        ct_bl_intercept_retry = 0.0
                        ct_baseline_retry     = channel_baseline_means.get(ch, global_baseline_mean)
                        try:
                            ch_thresh = channel_thresholds.get(ch, global_threshold)
                            # Re-derive baseline window from metadata for this sample
                            bl_param_retry = None
                            if sample_metadata:
                                wm_r     = sample_metadata.get(sample_name, {})
                                bl_start = wm_r.get('Baseline Start')
                                bl_end   = wm_r.get('Baseline End')
                                try:
                                    bl_si = int(np.searchsorted(cycles, float(bl_start)))
                                    bl_ei = int(np.searchsorted(cycles, float(bl_end)))
                                    if bl_ei > bl_si + 1:
                                        bl_param_retry = (bl_si, bl_ei)
                                except (TypeError, ValueError):
                                    pass
                            # Use ROX-normalized signal for Ct if available.
                            # Use full cycle array so baseline_cycles indices are correct.
                            _rox_by_well_r = st.session_state.get('rox_by_well', {})
                            _well_pos_for_rox_r = _get_well_pos(sample_name)
                            _rox_arr_r = _rox_by_well_r.get(_well_pos_for_rox_r, None)
                            if _rox_arr_r is not None and len(_rox_arr_r) == len(fluor_data):
                                _fluor_for_ct_r = fluor_data / np.maximum(_rox_arr_r, 1e-10)
                                _ct_rox_mean_r  = float(np.mean(_rox_arr_r))
                                _ct_thresh_r    = ch_thresh
                            else:
                                _fluor_for_ct_r = fluor_data
                                _ct_rox_mean_r  = None
                                _ct_thresh_r    = None  # auto from raw
                            _orig_cycles_fit_r = optimizer_retry.cycles_fit
                            _orig_fluor_fit_r  = optimizer_retry.fluorescence_fit
                            optimizer_retry.cycles_fit       = cycles
                            optimizer_retry.fluorescence_fit = _fluor_for_ct_r
                            ct_res                = optimizer_retry.calculate_ct(
                                method='threshold',
                                threshold=_ct_thresh_r,
                                baseline_cycles=bl_param_retry,
                            )
                            optimizer_retry.cycles_fit       = _orig_cycles_fit_r
                            optimizer_retry.fluorescence_fit = _orig_fluor_fit_r
                            ct_retry              = ct_res['ct']
                            ct_baseline_retry     = ct_res.get('baseline_mean', ct_baseline_retry)
                            ct_bl_slope_retry     = ct_res.get('baseline_slope', 0.0)
                            ct_bl_intercept_retry = ct_res.get('baseline_intercept', 0.0)
                        except Exception:
                            pass

                        # Accept the retry if: original had no fit, OR retry R² is better.
                        original_r2  = result.get('R2')
                        retry_r2     = metrics_retry['r_squared']
                        retry_better = (
                            original_r2 is None
                            or (retry_r2 is not None and original_r2 is not None
                                and retry_r2 > original_r2)
                        )

                        # Track best result across all retry attempts.
                        # Start from the original pass-1 result so we only
                        # replace it if a retry genuinely improves R².
                        _best_r2     = original_r2 if original_r2 is not None else -999.0
                        _best_result = None
                        if retry_better:
                            _best_r2     = retry_r2 if retry_r2 is not None else _best_r2
                            _best_result = {
                                'params':       params_retry,
                                'metrics':      metrics_retry,
                                'optimizer':    optimizer_retry,
                                'start_idx':    retry_start_idx,
                                'fit_end':      _fit_end_cycle_r,
                                'ct':           ct_retry,
                                'ct_bl_mean':   ct_baseline_retry,
                                'ct_bl_slope':  ct_bl_slope_retry,
                                'ct_bl_int':    ct_bl_intercept_retry,
                                'ct_rox_mean':  _ct_rox_mean_r,
                            }

                        # Track retry progress for diagnostics
                        _retry_stage = 'initial-retry'
                        _retry_attempts = 1
                        _retry_timed_out = False

                        # If still below target, try window variations.
                        # First try extending baseline (more pre-max cycles),
                        # then try shrinking/extending after.
                        if (_best_r2 < _r2_target
                                and _time_mod.perf_counter() - _retry_t0 < _RETRY_TIMEOUT):
                            _retry_stage = 'window-variations'
                            _win_variations = [
                                # Phase 1: try original short truncation (no extension)
                                (cycles_before_max,     cycles_after_max),
                                (cycles_before_max,     max(3, cycles_after_max - 1)),
                                # Phase 2: extend baseline for better background anchor
                                (cycles_before_max + 4, cycles_after_max),
                                (cycles_before_max + 8, cycles_after_max),
                                # Phase 3: extend after for more plateau data
                                (cycles_before_max,     cycles_after_max + 3),
                                (cycles_before_max - 2, cycles_after_max + 3),
                                (cycles_before_max + 4, cycles_after_max + 3),
                                (cycles_before_max - 4, cycles_after_max),
                            ]
                            for _wv_before, _wv_cam in _win_variations:
                                _wv_before = max(3, _wv_before)
                                _wv_start  = max(_r_floor, _r_max_slope_idx - _wv_before)
                                _wv_c      = cycles[_wv_start:]
                                _wv_f      = fluor_data[_wv_start:]

                                # Re-estimate background for this window
                                _wv_bg_pre  = max(_r_floor, _wv_start - 6)
                                _wv_bg_post = min(len(cycles), _wv_start + 2)
                                _wv_bg_c    = cycles[_wv_bg_pre:_wv_bg_post]
                                _wv_bg_f    = fluor_data[_wv_bg_pre:_wv_bg_post]
                                if len(_wv_bg_c) >= 2:
                                    _wv_coeffs = np.polyfit(_wv_bg_c, _wv_bg_f, 1)
                                    _wv_slope  = float(_wv_coeffs[0])
                                    _wv_int    = float(_wv_coeffs[1])
                                else:
                                    _wv_slope = 0.0
                                    _wv_int   = float(_wv_f[0]) if len(_wv_f) else 0.0

                                _wv_sm     = max(abs(_wv_slope) * 0.40, F_range_r * 0.002)
                                _wv_id     = max(abs(_wv_int) * 0.005, F_range_r * 0.03)
                                _wv_bounds = dict(informed_bounds)
                                _wv_bounds['D0']             = (1e-8, max(F_range_r, 1.0))
                                _wv_bounds['F_bg_slope']     = (_wv_slope - _wv_sm, _wv_slope + _wv_sm)
                                _wv_bounds['F_bg_intercept'] = (_wv_int - _wv_id, _wv_int + _wv_id)

                                try:
                                    _wv_model = MAK2Model()
                                    _wv_opt   = MAK2Optimizer(_wv_model)
                                    _wv_params = _wv_opt.fit(
                                        _wv_c, _wv_f,
                                        cycles_after_max=_wv_cam,
                                        auto_truncate=auto_truncate,
                                        truncate_cycle=truncate_cycle,
                                        bounds=_wv_bounds,
                                        fixed_background_values={
                                            'F_bg_slope': _wv_slope,
                                            'F_bg_intercept': _wv_int,
                                        },
                                        verbose=False,
                                    )
                                    _wv_metrics = _wv_opt.calculate_fit_metrics()
                                    _wv_r2 = _wv_metrics['r_squared']
                                    if _wv_r2 is not None and _wv_r2 > _best_r2:
                                        _best_r2 = _wv_r2
                                        _wv_fe = float(_wv_opt.cycles_fit[-1]) \
                                            if _wv_opt.cycles_fit is not None and len(_wv_opt.cycles_fit) > 0 \
                                            else float(cycles[-1])
                                        _best_result = {
                                            'params':       _wv_params,
                                            'metrics':      _wv_metrics,
                                            'optimizer':    _wv_opt,
                                            'start_idx':    _wv_start,
                                            'fit_end':      _wv_fe,
                                            'ct':           ct_retry,
                                            'ct_bl_mean':   ct_baseline_retry,
                                            'ct_bl_slope':  ct_bl_slope_retry,
                                            'ct_bl_int':    ct_bl_intercept_retry,
                                            'ct_rox_mean':  _ct_rox_mean_r,
                                        }
                                        if _best_r2 >= _r2_target:
                                            break  # good enough
                                except Exception:
                                    pass
                                _retry_attempts += 1
                                if _time_mod.perf_counter() - _retry_t0 >= _RETRY_TIMEOUT:
                                    _retry_timed_out = True
                                    break  # timeout

                        # If still below target and k is at or near the lower
                        # bound, retry with relaxed k bounds.  Late amplifiers
                        # with incomplete S-curves sometimes need a lower k
                        # floor to converge.
                        if (_best_r2 < _r2_target
                                and not _retry_timed_out
                                and _time_mod.perf_counter() - _retry_t0 < _RETRY_TIMEOUT):
                            _retry_stage = 'k-relaxation'
                            _retry_attempts += 1
                            _cur_k = (_best_result['params']['k']
                                      if _best_result else params_retry.get('k'))
                            _k_lo  = informed_bounds['k'][0]
                            if _cur_k is not None and _cur_k < _k_lo * 1.5:
                                _relax_bounds = dict(informed_bounds)
                                _relax_bounds['k'] = (0.001, _relax_bounds['k'][1])
                                _relax_bounds['D0'] = (1e-15, max(F_range_r * 10, 1.0))
                                _relax_bounds['F_bg_slope']     = (_r_s_min, _r_s_max)
                                _relax_bounds['F_bg_intercept'] = (_r_int_lo, _r_int_hi)
                                try:
                                    _rk_model = MAK2Model()
                                    _rk_opt   = MAK2Optimizer(_rk_model)
                                    _rk_params = _rk_opt.fit(
                                        cycles_retry, fluor_retry,
                                        cycles_after_max=_retry_cam,
                                        auto_truncate=auto_truncate,
                                        truncate_cycle=truncate_cycle,
                                        bounds=_relax_bounds,
                                        fixed_background_values={
                                            'F_bg_slope': _r_bg_slope_win,
                                            'F_bg_intercept': _r_bg_int_win,
                                        },
                                        verbose=False,
                                    )
                                    _rk_metrics = _rk_opt.calculate_fit_metrics()
                                    _rk_r2 = _rk_metrics['r_squared']
                                    if _rk_r2 is not None and _rk_r2 > _best_r2:
                                        _best_r2 = _rk_r2
                                        _rk_fe = float(_rk_opt.cycles_fit[-1]) \
                                            if _rk_opt.cycles_fit is not None and len(_rk_opt.cycles_fit) > 0 \
                                            else float(cycles[-1])
                                        _best_result = {
                                            'params':       _rk_params,
                                            'metrics':      _rk_metrics,
                                            'optimizer':    _rk_opt,
                                            'start_idx':    retry_start_idx,
                                            'fit_end':      _rk_fe,
                                            'ct':           ct_retry,
                                            'ct_bl_mean':   ct_baseline_retry,
                                            'ct_bl_slope':  ct_bl_slope_retry,
                                            'ct_bl_int':    ct_bl_intercept_retry,
                                            'ct_rox_mean':  _ct_rox_mean_r,
                                        }
                                except Exception:
                                    pass

                        # Accept best result if it improved over original
                        if _best_result is not None and (
                            original_r2 is None or _best_r2 > original_r2
                        ):
                            _br = _best_result
                            results_list[idx] = {
                                'Sample':                sample_name,
                                'Ct':                    _br['ct'],
                                'Ct_baseline_mean':      _br['ct_bl_mean'],
                                'Ct_baseline_slope':     _br['ct_bl_slope'],
                                'Ct_baseline_intercept': _br['ct_bl_int'],
                                'D0':                    _br['params']['D0'],
                                'k':                     _br['params']['k'],
                                'P0':                    _br['params']['P0'],
                                'F_bg_intercept':        _br['params']['F_bg_intercept'],
                                'F_bg_slope':            _br['params']['F_bg_slope'],
                                'R2':                    _br['metrics']['r_squared'],
                                'RMSE':                  _br['metrics']['rmse'],
                                'NRMSE':                 _br['metrics']['nrmse'] * 100,
                                'SSR':                   _br['metrics']['ssr'],
                                'Tier':                  result.get('Tier'),
                                'Instrument':            result.get('Instrument', ''),
                                'Success':               ('✓ (window-retry)'
                                                          if _best_r2 >= 0.999
                                                          else ('✓ (late-amp)'
                                                                if _retry_is_late and _best_r2 >= 0.995
                                                                else ('⚠️ timeout@' + _retry_stage
                                                                      if _retry_timed_out
                                                                      else ('✓ (noisy data)'
                                                                            if _best_r2 >= 0.995
                                                                            else '⚠️ R² below target')))),
                                'retry_stage':           _retry_stage,
                                'retry_attempts':        _retry_attempts,
                                'retry_elapsed_s':       round(_time_mod.perf_counter() - _retry_t0, 1),
                                'retry_timed_out':       _retry_timed_out,
                                'FixedBG':               '',
                                'Fallback':              '',
                                'FallbackOK':            '',
                                'bg_slope_est':          bg_slope_est_r,
                                'bg_intercept_est':      bg_intercept_est_r,
                                'bl_end_meta':           _r_meta_bl_end_cycle,
                                'bl_end_est':            _r_est_bl_end_cycle,
                                'fit_start_cycle':       float(cycles[_br['start_idx']]),
                                'fit_end_cycle':         _br['fit_end'],
                                'ct_rox_mean':           _br['ct_rox_mean'],
                                'fluor_data':            fluor_data,
                            }
                        else:
                            # Keep original — accept if R² is good enough
                            orig_r2 = result.get('R2')
                            _orig_fe = result.get('fit_end_cycle')
                            _orig_late = (_orig_fe is not None and
                                          not (isinstance(_orig_fe, float) and np.isnan(_orig_fe)) and
                                          float(_orig_fe) >= _last_cyc - min(max(1, cycles_after_max), 5))
                            _orig_r2_thr = 0.995 if _orig_late else 0.999
                            if orig_r2 is not None and orig_r2 >= _orig_r2_thr:
                                results_list[idx]['Success'] = '✓'
                            else:
                                orig_k = result.get('k')
                                if orig_k is not None and orig_k > 0.5:
                                    results_list[idx]['Success'] = '⚠️ Degenerate k'
                                else:
                                    if orig_r2 is not None and orig_r2 >= 0.995:
                                        results_list[idx]['Success'] = '✓ (noisy data)'
                                    else:
                                        results_list[idx]['Success'] = '⚠️ R² below target'

                    except Exception as e:
                        if result['k'] is None:
                            results_list[idx]['Success'] = f'✗ Error: {str(e)[:30]}'
                        else:
                            results_list[idx]['Success'] = '⚠️ Retry failed'

                    # ── Checkpoint: save progress after each retry ──────────
                    _checkpoint['pass2_completed_indices'].append(idx)
                    _checkpoint['results_list'] = results_list
                    _checkpoint['updated_at'] = _dt.utcnow().isoformat()
                    _save_checkpoint(_checkpoint)

            # ── Post-fit non-amplification check ───────────────────────
            # Even when the optimizer converges, the well may not show
            # real amplification.  Three quality gates catch false
            # positives (fitting noise/drift rather than a real sigmoid):
            #
            #  1. Minimum fold change: F_max / F_baseline within the fit
            #     window must exceed 2×.  Real amplification shows ≥ 3×.
            #  2. Fit window width: fit_end − fit_start must be ≥ 5
            #     cycles.  A narrower window means no clear transition.
            #  3. Sigmoid shape: the fitted curve must show a clear
            #     inflection (second derivative sign change) within the
            #     fit window.  A monotone curve is drift, not sigmoid.
            #
            # Wells failing ANY gate are reclassified as non-amplifying.
            for _pf_idx, _pf_r in enumerate(results_list):
                if _pf_r.get('error') is not None:
                    continue  # already flagged
                _pf_r2 = _pf_r.get('R2')
                _pf_reject = False
                _pf_reason = ''

                # Detect late amplifiers early — needed by Gate 0 and Gate 3.
                _pf_fe_g0 = _pf_r.get('fit_end_cycle')
                _pf_is_late = (
                    _pf_fe_g0 is not None
                    and not (isinstance(_pf_fe_g0, float) and np.isnan(_pf_fe_g0))
                    and _pf_fe_g0 >= float(cycles[-1]) - min(max(1, cycles_after_max), 5)
                )

                # Gate 0: poor R² — a MAK2 sigmoid+linear-background model
                # can fit monotonic drift/noise to R²≈0.98, so the threshold
                # must be well above that.  All legitimate amplifications on
                # real qPCR data give R² ≥ 0.996; 0.99 provides margin.
                # Late amplifiers (fit extends to last cycle) get a relaxed
                # threshold of 0.85 because incomplete S-curves inherently
                # have lower R² — the linear-vs-MAK2 gate catches drift.
                # 0.85 accommodates very late amplifiers where only the
                # exponential rise is captured without any plateau.
                _pf_r2_thresh = 0.85 if _pf_is_late else 0.99
                if _pf_r2 is not None and _pf_r2 < _pf_r2_thresh:
                    _pf_reject = True
                    _pf_reason = f'R\u00b2 = {_pf_r2:.4f} < {_pf_r2_thresh}'

                # Gate 1: removed — signal departure check is redundant
                # with upstream detect_no_signal_samples() and too
                # aggressive for raw RFU data with high baseline noise.

                # Gate 2: fit window too narrow
                if not _pf_reject:
                    _pf_fs2 = _pf_r.get('fit_start_cycle')
                    _pf_fe2 = _pf_r.get('fit_end_cycle')
                    if (_pf_fs2 is not None and _pf_fe2 is not None
                            and _pf_fe2 - _pf_fs2 < 10):
                        _pf_reject = True
                        _pf_reason = f'Fit window {_pf_fe2 - _pf_fs2:.0f} cycles < 10'

                # Gate 2b: linear-vs-MAK2 comparison on pre-inflection data.
                # In real amplification the pre-inflection region (fit_start
                # to max-slope cycle) is exponential growth where a sigmoid
                # massively outperforms a straight line.  In drift, both
                # fit equally well.  Reject if MAK2 R² improvement < 0.05.
                # Skip for late amplifiers ONLY if R² >= 0.995 — a genuine
                # late amplifier with exponential growth will have high R².
                # Linear drift that happens to extend to the last cycle
                # (triggering late-amp classification) must still be tested.
                _pf_late_bypass_2b = _pf_is_late and _pf_r2 is not None and _pf_r2 >= 0.995
                if (not _pf_reject and not _pf_late_bypass_2b
                        and _pf_r.get('D0') is not None
                        and not (isinstance(_pf_r['D0'], float) and np.isnan(_pf_r['D0']))
                        and _pf_r.get('fluor_data') is not None):
                    try:
                        _pf_fs2b = _pf_r.get('fit_start_cycle')
                        _pf_fe2b = _pf_r.get('fit_end_cycle')
                        if _pf_fs2b is not None and _pf_fe2b is not None:
                            _pf_m2b = MAK2Model()
                            _pf_c_full2b = cycles[:len(_pf_r['fluor_data'])]
                            _pf_pred2b = _pf_m2b.simulate_to_cycle(
                                D0=_pf_r['D0'], k=_pf_r['k'], P0=_pf_r['P0'],
                                cycles=_pf_c_full2b,
                                F_bg_intercept=_pf_r['F_bg_intercept'],
                                F_bg_slope=_pf_r['F_bg_slope'],
                            )
                            # Restrict to fit window
                            _pf_win2b = (_pf_c_full2b >= _pf_fs2b) & (_pf_c_full2b <= _pf_fe2b)
                            _pf_cycles_win = _pf_c_full2b[_pf_win2b]
                            _pf_pred_win2b = _pf_pred2b[_pf_win2b]
                            _pf_fluor_full2b = np.asarray(_pf_r['fluor_data'])
                            _pf_fluor_win2b = _pf_fluor_full2b[_pf_win2b]

                            # Find max-slope cycle from MAK2 prediction
                            _pf_d1_2b = np.gradient(_pf_pred_win2b, _pf_cycles_win)
                            _pf_max_slope_idx = int(np.argmax(_pf_d1_2b))
                            _pf_max_slope_cycle = _pf_cycles_win[_pf_max_slope_idx]

                            # Pre-inflection region: fit_start → max_slope_cycle
                            _pf_pre_mask = _pf_cycles_win <= _pf_max_slope_cycle
                            _pf_fluor_pre = _pf_fluor_win2b[_pf_pre_mask]
                            _pf_cycles_pre = _pf_cycles_win[_pf_pre_mask]

                            if len(_pf_fluor_pre) >= 4:
                                # R² of linear fit on pre-inflection data
                                _pf_coeffs = np.polyfit(_pf_cycles_pre, _pf_fluor_pre, 1)
                                _pf_lin_pred = np.polyval(_pf_coeffs, _pf_cycles_pre)
                                _pf_ss_tot = float(np.sum((_pf_fluor_pre - np.mean(_pf_fluor_pre))**2))
                                if _pf_ss_tot > 0:
                                    _pf_ss_res_lin = float(np.sum((_pf_fluor_pre - _pf_lin_pred)**2))
                                    _pf_r2_lin = 1.0 - _pf_ss_res_lin / _pf_ss_tot

                                    # R² of MAK2 on same pre-inflection region
                                    _pf_mak2_pre = _pf_pred_win2b[_pf_pre_mask]
                                    _pf_ss_res_mak = float(np.sum((_pf_fluor_pre - _pf_mak2_pre)**2))
                                    _pf_r2_mak = 1.0 - _pf_ss_res_mak / _pf_ss_tot

                                    # Reject if MAK2 doesn't outperform linear
                                    if _pf_r2_mak - _pf_r2_lin < 0.05:
                                        _pf_reject = True
                                        _pf_reason = (
                                            f'MAK2 not better than linear in growth region '
                                            f'(R\u00b2_MAK2={_pf_r2_mak:.4f}, '
                                            f'R\u00b2_lin={_pf_r2_lin:.4f})'
                                        )
                    except Exception:
                        pass

                # Gate 3: sigmoid shape — the fitted curve within the fit
                # window must show a clear inflection (second derivative
                # sign change).  Only check within fit window, not the
                # full curve, and require the curvature to be significant
                # relative to the signal range (not just numerical noise).
                # Skip for late amplifiers — they may only capture the
                # exponential rise without reaching the inflection point.
                # Skip for high-R² fits (≥ 0.999) — the fit quality
                # already validates the curve shape.
                # _pf_is_late already computed above (before Gate 0)
                _pf_fit_width = (
                    (_pf_r.get('fit_end_cycle') or 0)
                    - (_pf_r.get('fit_start_cycle') or 0)
                )
                _pf_high_r2 = (
                    _pf_r2 is not None
                    and _pf_r2 >= 0.999
                    and _pf_fit_width >= 10
                )
                _pf_late_bypass_3 = _pf_is_late and _pf_r2 is not None and _pf_r2 >= 0.995
                if (not _pf_reject and not _pf_late_bypass_3 and not _pf_high_r2
                        and _pf_r.get('D0') is not None
                        and not (isinstance(_pf_r['D0'], float) and np.isnan(_pf_r['D0']))
                        and _pf_r.get('fluor_data') is not None):
                    try:
                        _pf_fs3 = _pf_r.get('fit_start_cycle')
                        _pf_fe3 = _pf_r.get('fit_end_cycle')
                        if _pf_fs3 is not None and _pf_fe3 is not None:
                            _pf_m = MAK2Model()
                            _pf_c_full = cycles[:len(_pf_r['fluor_data'])]
                            _pf_pred = _pf_m.simulate_to_cycle(
                                D0=_pf_r['D0'], k=_pf_r['k'], P0=_pf_r['P0'],
                                cycles=_pf_c_full,
                                F_bg_intercept=_pf_r['F_bg_intercept'],
                                F_bg_slope=_pf_r['F_bg_slope'],
                            )
                            # Restrict to fit window
                            _pf_win_mask = (_pf_c_full >= _pf_fs3) & (_pf_c_full <= _pf_fe3)
                            _pf_pred_win = _pf_pred[_pf_win_mask]
                            if len(_pf_pred_win) >= 5:
                                _pf_d1 = np.gradient(_pf_pred_win)
                                _pf_d2 = np.gradient(_pf_d1)
                                _pf_pred_range = float(np.max(_pf_pred_win) - np.min(_pf_pred_win))
                                # Curvature must be significant (> 1% of signal range)
                                _pf_d2_thresh = _pf_pred_range * 0.01
                                _pf_has_inflection = (
                                    np.any(_pf_d2 > _pf_d2_thresh)
                                    and np.any(_pf_d2 < -_pf_d2_thresh)
                                )
                                if not _pf_has_inflection:
                                    _pf_reject = True
                                    _pf_reason = 'No inflection (monotone curve)'
                    except Exception:
                        pass

                if _pf_reject:
                    results_list[_pf_idx]['Success'] = ''
                    results_list[_pf_idx]['error'] = f'No amplification detected ({_pf_reason})'
                    results_list[_pf_idx]['D0'] = None
                    results_list[_pf_idx]['k'] = None
                    results_list[_pf_idx]['P0'] = None
                    results_list[_pf_idx]['Ct'] = None

            # Restore stdout and save captured optimizer output
            sys.stdout = _batch_old_stdout
            _batch_debug = _batch_captured.getvalue()
            st.session_state['batch_debug_output'] = _batch_debug

            status_text.text("✅ Batch fitting complete!")
            st.toast("Fitting complete — results saved!", icon="✅")

            # Create results dataframe (remove fluor_data before display)
            _hidden = {'fluor_data', 'bg_slope_est', 'bg_intercept_est'}
            display_results = [{k: v for k, v in r.items() if k not in _hidden} for r in results_list]
            results_df = pd.DataFrame(display_results)

            # Store core session state FIRST so results are never lost
            st.session_state['batch_results'] = results_df
            st.session_state['batch_results_list'] = results_list
            st.session_state['batch_all_samples'] = all_samples_to_fit if _is_resuming else all_samples
            st.session_state['batch_no_signal_samples'] = no_signal_samples
            st.session_state['batch_cycles'] = cycles
            st.session_state['batch_settings'] = {
                'first_fit_cycle': first_fit_cycle,
                'cycles_before_max': cycles_before_max,
                'cycles_after_max': cycles_after_max,
                'auto_truncate': auto_truncate,
                'truncate_cycle': truncate_cycle,
                'custom_bounds_dict': custom_bounds_dict,
                'global_threshold': global_threshold,
                'global_baseline_mean': global_baseline_mean,
                'channel_thresholds': channel_thresholds,
                'channel_baseline_means': channel_baseline_means,
            }

            # Persist to disk so results survive session resets
            # Collect fluorescence data for no-signal wells
            # Start with the early save (line ~1922) so we never clobber it
            _ns_fluor_save = dict(st.session_state.get('batch_no_signal_fluor', {}))
            _all_src = all_samples if not _is_resuming else {}
            for _ns_key in no_signal_samples:
                if _ns_key not in _ns_fluor_save:
                    if _ns_key in _all_src:
                        _ns_fluor_save[_ns_key] = _all_src[_ns_key]
                    elif _ns_key in st.session_state.get('upload_batch_samples', {}):
                        _ns_fluor_save[_ns_key] = st.session_state['upload_batch_samples'][_ns_key]
            st.session_state['batch_no_signal_fluor'] = _ns_fluor_save
            _save_results_to_disk(
                results_df, results_list,
                all_samples_to_fit if _is_resuming else all_samples,
                no_signal_samples, cycles,
                st.session_state['batch_settings'],
                no_signal_fluor=_ns_fluor_save,
            )

            # Batch complete — clear the checkpoint (no longer needed)
            _clear_checkpoint()

            # Now add display columns (if this fails, raw results still persist)
            try:
                # Add Target and Well columns for multiplexed data (target::well format)
                all_targets = st.session_state.get('all_targets', [])
                if all_targets and results_df['Sample'].str.contains('::').any():
                    results_df.insert(0, 'Target', results_df['Sample'].str.split('::').str[0])
                    results_df.insert(1, 'Well', results_df['Sample'].str.split('::').str[1])

                # Add Channel column for multi-channel data (FAM_A1 format)
                all_channels = st.session_state.get('selected_channels', [])
                if all_channels and len(all_channels) > 1:
                    results_df.insert(0, 'Channel', results_df['Sample'].map(_ch))

                # Annotate with sample metadata (Sample Name, Task, Known_Copies) if available
                sample_metadata = st.session_state.get('sample_metadata')
                if sample_metadata:
                    results_df.insert(
                        results_df.columns.get_loc('Sample') + 1,
                        'Sample_Name',
                        results_df['Sample'].map(
                            lambda w: sample_metadata.get(w, {}).get('Sample Name', '')
                        )
                    )
                    results_df.insert(
                        results_df.columns.get_loc('Sample_Name') + 1,
                        'Task',
                        results_df['Sample'].map(
                            lambda w: sample_metadata.get(w, {}).get('Task', '')
                        )
                    )
                    results_df['Known_Copies'] = results_df['Sample'].map(
                        lambda w: sample_metadata.get(w, {}).get('Quantity', np.nan)
                    )

                    # Add instrument Ct column for comparison (from ABI results CSV)
                    if any('Ct_instrument' in sample_metadata.get(w, {})
                           for w in results_df['Sample']):
                        ct_inst_col = results_df['Sample'].map(
                            lambda w: sample_metadata.get(w, {}).get('Ct_instrument', np.nan)
                        )
                        ct_loc = results_df.columns.get_loc('Ct') + 1
                        results_df.insert(ct_loc, 'Ct_instrument', ct_inst_col)

                # Update session state with enriched DataFrame
                st.session_state['batch_results'] = results_df
            except Exception as _build_err:
                import traceback as _tb
                st.error(f"Error enriching results: {_build_err}")

          except Exception as _batch_fatal:
            # Restore stdout if it was redirected
            if sys.stdout is not _batch_old_stdout:
                sys.stdout = _batch_old_stdout
                _batch_debug = _batch_captured.getvalue()
                st.session_state['batch_debug_output'] = _batch_debug
            import traceback as _tb_fatal
            st.error(f"❌ Batch fitting crashed: {_batch_fatal}")
            st.code(_tb_fatal.format_exc())


    # Display batch results (outside button block, always visible if results exist)
    if not (batch_mode and all_samples) and 'batch_results' not in st.session_state:
        # Single sample mode
        if st.sidebar.button("🔬 Fit Model", type="primary"):
            with st.spinner("Fitting MAK2 model..."):
                model = MAK2Model()
                optimizer = MAK2Optimizer(model)

                # Capture stdout to display debug output
                import io
                import sys
                captured_output = io.StringIO()
                old_stdout = sys.stdout
                sys.stdout = captured_output

                try:
                    # Phase-1 unified pipeline. Single-sample mode now runs the
                    # full per-well fit (no-amp pre-check, take-off detection,
                    # bg pre-estimation, optimizer, toe stages, Ct, tier). The
                    # gates below operate on the reconstructed optimizer state.
                    from fit_well import fit_well as _fit_well_single
                    _result = _fit_well_single(
                        cycles, fluorescence,
                        first_fit_cycle=first_fit_cycle,
                        cycles_before_max=cycles_before_max,
                        cycles_after_max=cycles_after_max,
                        auto_truncate=auto_truncate,
                        truncate_cycle=truncate_cycle,
                        fit_bounds=dict(custom_bounds_dict) if custom_bounds_dict else None,
                        sample_name='single_sample',
                    )
                    if _result.get('error') == 'No amplification detected':
                        sys.stdout = old_stdout
                        st.warning("No amplification detected in this sample — "
                                   "signal range is within baseline noise.")
                        st.stop()
                    if _result.get('error'):
                        raise RuntimeError(_result['error'])

                    # Pack fit_well's result into the {fitted_params, optimizer}
                    # shape the display code at the bottom of this file expects.
                    fitted_params = {
                        'D0': _result['D0'], 'k': _result['k'], 'P0': _result['P0'],
                        'F_bg_intercept': _result['F_bg_intercept'],
                        'F_bg_slope':     _result['F_bg_slope'],
                        'tier':           _result['Tier'],
                    }
                    _fs_i = int(np.searchsorted(cycles, _result['fit_start_cycle']))
                    _fe_i = int(np.searchsorted(cycles, _result['fit_end_cycle'])) + 1
                    optimizer.cycles_fit = cycles[_fs_i:_fe_i]
                    optimizer.fluorescence_fit = fluorescence[_fs_i:_fe_i]
                    optimizer.optimal_params = fitted_params
                    optimizer.metrics = {
                        'r_squared': _result['R2'], 'rmse': _result['RMSE'],
                        'ssr': _result['SSR'],
                        'nrmse': (_result['NRMSE'] / 100.0) if _result['NRMSE'] is not None else None,
                    }

                    # ── Post-fit quality gates (same as batch mode) ──
                    _sq_warnings = []
                    _sq_metrics = optimizer.calculate_fit_metrics()
                    _sq_r2 = _sq_metrics.get('r_squared')
                    _sq_fs = float(optimizer.cycles_fit[0]) if optimizer.cycles_fit is not None and len(optimizer.cycles_fit) > 0 else None
                    _sq_fe = float(optimizer.cycles_fit[-1]) if optimizer.cycles_fit is not None and len(optimizer.cycles_fit) > 0 else None

                    # Gate 0: poor R² (drift/noise can fit to ~0.98)
                    if _sq_r2 is not None and _sq_r2 < 0.99:
                        _sq_warnings.append(f"R\u00b2 {_sq_r2:.4f} < 0.99")

                    if _sq_fs is not None and _sq_fe is not None:
                        # Gate 2: fit window width (≥ 8 cycles)
                        if _sq_fe - _sq_fs < 8:
                            _sq_warnings.append(f"Fit window {_sq_fe - _sq_fs:.0f} cycles < 8")

                    # Gate 3: sigmoid shape (inflection in fit window)
                    # Skip for late amplifiers — they may only capture the
                    # exponential rise without reaching the inflection point.
                    # Skip for high-R² fits (≥ 0.999) — fit quality validates shape.
                    _sq_is_late = (
                        _sq_fe is not None
                        and _sq_fe >= float(cycles[-1]) - min(max(1, cycles_after_max), 5)
                    )
                    _sq_r2_val = optimizer.metrics.get('r_squared', 0) if hasattr(optimizer, 'metrics') and optimizer.metrics else 0
                    _sq_fit_width = (_sq_fe - _sq_fs) if (_sq_fs is not None and _sq_fe is not None) else 0
                    _sq_high_r2 = (_sq_r2_val >= 0.999 and _sq_fit_width >= 10)
                    if not _sq_is_late and not _sq_high_r2:
                        try:
                            _sq_pred = optimizer.predict(cycles)
                            if _sq_fs is not None and _sq_fe is not None and len(_sq_pred) >= 5:
                                _sq_win_mask = (cycles >= _sq_fs) & (cycles <= _sq_fe)
                                _sq_pred_win = _sq_pred[_sq_win_mask]
                                if len(_sq_pred_win) >= 5:
                                    _sq_d1 = np.gradient(_sq_pred_win)
                                    _sq_d2 = np.gradient(_sq_d1)
                                    _sq_pred_range = float(np.max(_sq_pred_win) - np.min(_sq_pred_win))
                                    _sq_d2_thresh = _sq_pred_range * 0.01
                                    if not (np.any(_sq_d2 > _sq_d2_thresh) and np.any(_sq_d2 < -_sq_d2_thresh)):
                                        _sq_warnings.append("No inflection (monotone curve)")
                        except Exception:
                            pass

                    st.session_state['fitted_params'] = fitted_params
                    st.session_state['optimizer'] = optimizer

                    # Restore stdout and save captured output
                    sys.stdout = old_stdout
                    debug_output = captured_output.getvalue()
                    st.session_state['fit_debug_output'] = debug_output
                    # Store hash of data to validate fit matches current data
                    import hashlib
                    data_hash = hashlib.md5(f"{cycles.tobytes()}{fluorescence.tobytes()}".encode()).hexdigest()
                    st.session_state['fitted_data_hash'] = data_hash

                    if _sq_warnings:
                        st.warning(f"⚠️ Possible non-amplification: {'; '.join(_sq_warnings)}")
                    st.success("✅ Fitting complete!")

                except Exception as e:
                    # Restore stdout even if there was an error
                    sys.stdout = old_stdout
                    debug_output = captured_output.getvalue()
                    if debug_output:
                        st.session_state['fit_debug_output'] = debug_output
                    st.error(f"Fitting failed: {str(e)}")
                    st.stop()

# ============================================================================
# RESUME FROM CHECKPOINT — independent of file upload / data source state
# ============================================================================
if (_has_incomplete_checkpoint
    and 'batch_results' not in st.session_state
    and '_resume_checkpoint' not in st.session_state):
    _cp = _checkpoint_data
    _cp_total = _cp.get('pass1_total', 0)
    _cp_done = _cp.get('pass1_completed_count', 0)
    _cp_pass = _cp.get('current_pass', 'pass1')
    if _cp_pass == 'pass1':
        _cp_pct = (_cp_done / _cp_total * 100) if _cp_total > 0 else 0
        _cp_msg = f"Pass 1: {_cp_done}/{_cp_total} wells completed ({_cp_pct:.0f}%)"
    else:
        _cp_p2_done = len(_cp.get('pass2_completed_indices', []))
        _cp_p2_total = len(_cp.get('retry_indices', []))
        _cp_msg = f"Pass 1 complete. Pass 2: {_cp_p2_done}/{_cp_p2_total} retries completed"

    st.warning(f"⏸️ Incomplete batch fitting detected — {_cp_msg}")
    _cp_col1, _cp_col2 = st.columns(2)
    with _cp_col1:
        if st.button("▶️ Resume fitting", type="primary", key="resume_checkpoint"):
            # Store checkpoint in session state and rerun — the main batch
            # fitting path will pick it up via _is_resuming flag.
            st.session_state['_resume_checkpoint'] = _cp
            # Also restore sample data so batch_mode + all_samples are set
            _cp_samples = {}
            for k in _cp['sample_names']:
                _cp_samples[k] = np.array(_cp['all_samples'][k])
            st.session_state['upload_batch_samples'] = _cp_samples
            st.session_state['upload_cycles'] = np.array(_cp['cycles'])
            st.session_state['upload_ready_batch'] = True
            # Set a preview sample so the UI doesn't crash
            st.session_state['upload_preview_sample'] = _cp['sample_names'][0]
            if _cp.get('rox_by_well'):
                st.session_state['rox_by_well'] = {
                    k: np.array(v) for k, v in _cp['rox_by_well'].items()
                }
            if _cp.get('rox_normalized') is not None:
                st.session_state['rox_normalized'] = _cp['rox_normalized']
            if _cp.get('sample_metadata'):
                st.session_state['sample_metadata'] = _cp['sample_metadata']
            st.rerun()
    with _cp_col2:
        if st.button("🗑️ Discard and start fresh", key="discard_checkpoint"):
            _clear_checkpoint()
            st.rerun()

# ============================================================================
# RESULTS DISPLAY — independent of file upload / cycles state
# These blocks only need st.session_state, so they render even if
# the file uploader loses its reference during a long computation.
# ============================================================================

# ── Standalone Copy-Number Conversion sidebar ──────────────────────────────
# Rendered when batch_results exist but the main sidebar (inside the
# if-cycles block) was NOT shown — e.g., after "Load Previous Results".
if 'batch_results' in st.session_state and cycles is None:
    st.sidebar.markdown("---")
    st.sidebar.subheader("Copy Number Conversion")
    _standalone_cal = st.sidebar.radio(
        "Calibration method",
        ["Single-copy D0", "Manual CF", "None"],
        index=0,
        help="Choose how to convert D0 to copy numbers.",
        key="standalone_cal_method",
    )
    st.session_state['calibration_method'] = {
        "Single-copy D0": "d0_single",
        "Manual CF": "manual_cf",
        "None": "none",
    }[_standalone_cal]

    if _standalone_cal == "Single-copy D0":
        st.session_state.pop('manual_conversion_factor', None)
        # Check for targets in loaded results
        _loaded_results = st.session_state.get('batch_results')
        _loaded_targets = []
        if _loaded_results is not None and 'Target' in _loaded_results.columns:
            _loaded_targets = sorted(_loaded_results['Target'].dropna().unique().tolist())

        if _loaded_targets:
            st.sidebar.caption(
                "Enter D0_single for each target. Leave at 0 to skip."
            )
            _d0s_per_target = {}
            for _tgt in _loaded_targets:
                _prev = st.session_state.get('d0_single_per_target', {}).get(_tgt, 0.0)
                _val = st.sidebar.number_input(
                    f"D0_single — {_tgt}",
                    min_value=0.0,
                    value=float(_prev),
                    format="%.3e",
                    help=f"D0 for a single copy of {_tgt}. Copies = D0 / D0_single.",
                    key=f"d0_single_standalone_{_tgt}",
                )
                _d0s_per_target[_tgt] = _val
            st.session_state['d0_single_per_target'] = _d0s_per_target
            st.session_state.pop('d0_single_value', None)
        else:
            _d0s_val = st.sidebar.number_input(
                "D0 for single copy",
                min_value=0.0,
                value=float(st.session_state.get('d0_single_value', 0.0)),
                format="%.3e",
                help="D0 value corresponding to exactly 1 copy. Copies = D0 / D0_single.",
                key="d0_single_standalone_global",
            )
            if _d0s_val > 0:
                st.session_state['d0_single_value'] = _d0s_val
            else:
                st.session_state.pop('d0_single_value', None)
            st.session_state.pop('d0_single_per_target', None)

    elif _standalone_cal == "Manual CF":
        _manual_cf = st.sidebar.number_input(
            "Manual conversion factor (copies/D0)",
            min_value=0.0,
            value=0.0,
            format="%.2e",
            help="Enter a known copies-per-D0-unit conversion factor.",
            key="manual_cf_standalone",
        )
        if _manual_cf > 0:
            st.session_state['manual_conversion_factor'] = _manual_cf
        else:
            st.session_state.pop('manual_conversion_factor', None)

    else:  # None
        st.session_state.pop('manual_conversion_factor', None)
        st.session_state.pop('d0_single_value', None)
        st.session_state.pop('d0_single_per_target', None)

if 'batch_results' in st.session_state:
    st.subheader("🔄 Batch Fitting Results")

    # Show batch optimizer diagnostics (captured stdout)
    if 'batch_debug_output' in st.session_state and st.session_state['batch_debug_output']:
        # Filter to only show catastrophic-fit diagnostics + summary
        _raw_debug = st.session_state['batch_debug_output']
        # Extract CATASTROPHIC FIT blocks for prominent display
        _catast_lines = []
        _in_block = False
        for _dl in _raw_debug.splitlines():
            if 'CATASTROPHIC FIT' in _dl:
                _in_block = True
            if _in_block:
                _catast_lines.append(_dl)
                if _dl.startswith('====') and len(_catast_lines) > 2:
                    _in_block = False
        if _catast_lines:
            with st.expander("🔴 Catastrophic Fit Diagnostics", expanded=True):
                st.code('\n'.join(_catast_lines), language='text')
        with st.expander("🐛 Full Optimizer Log", expanded=False):
            st.code(_raw_debug[-50000:] if len(_raw_debug) > 50000 else _raw_debug, language='text')

    # Prominent download buttons at top of results
    try:
        _xl_results_top = st.session_state['batch_results']
        # Gather all available extra sheets (plain data tables)
        _xl_extra = {}
        _xl_keys = [
            ('_no_signal_df',            'No Signal Samples'),
            ('_replicate_stats_df',      'Replicate Statistics'),
            ('_precision_comparison_df', 'Precision Comparison'),
            ('_limited_dilution_df',     'Limited Dilution'),
        ]
        for _ss_key, _sheet_name in _xl_keys:
            _df = st.session_state.get(_ss_key)
            if _df is not None and len(_df) > 0:
                _xl_extra[_sheet_name] = _df

        # Copies sheet(s) — only rows where Copies_D0 was calculated
        if 'Copies_D0' in _xl_results_top.columns:
            _copies_base_cols = ['Sample', 'D0', 'Copies_D0', 'R2']
            if 'Target' in _xl_results_top.columns:
                _copies_base_cols.insert(1, 'Target')
            if 'Channel' in _xl_results_top.columns:
                _copies_base_cols.insert(1, 'Channel')
            _copies_all = _xl_results_top[
                _xl_results_top['Copies_D0'].notna()
            ][_copies_base_cols].copy()
            _copies_all.insert(
                _copies_all.columns.get_loc('Copies_D0') + 1,
                'D0_single (used)',
                _copies_all['D0'] / _copies_all['Copies_D0'],
            )
            if 'Target' in _copies_all.columns and _copies_all['Target'].nunique() > 1:
                # One sheet per target that had copies calculated
                for _cp_tgt in _copies_all['Target'].unique():
                    _cp_df = _copies_all[_copies_all['Target'] == _cp_tgt].copy()
                    _sheet_nm = f"Copies {_cp_tgt}"[:31]  # Excel 31-char sheet name limit
                    _xl_extra[_sheet_nm] = _cp_df
            elif len(_copies_all) > 0:
                _xl_extra['Copies'] = _copies_all

        # Per-channel/target standard curve variance sheets
        _xl_grp_keys = st.session_state.get('_std_curve_group_keys', [None])
        for _xl_gk in _xl_grp_keys:
            _gk_sfx = f"_{_xl_gk}" if _xl_gk else ""
            _gk_label = f" ({_xl_gk})" if _xl_gk else ""
            _var_df_gk = st.session_state.get(f'_std_curve_variance_df{_gk_sfx}')
            if _var_df_gk is not None and len(_var_df_gk) > 0:
                _sheet_nm = f'Std Curve Variance{_gk_label}'[:31]
                _xl_extra[_sheet_nm] = _var_df_gk

        # Gather chart sheets (data + summary + native Excel chart)
        # Per-channel/target D0 and Ct standard curve sheets
        _xl_charts = {}
        for _xl_gk in _xl_grp_keys:
            _gk_sfx = f"_{_xl_gk}" if _xl_gk else ""
            _gk_label = f" ({_xl_gk})" if _xl_gk else ""
            _d0_df = st.session_state.get(f'_std_curve_d0_df{_gk_sfx}')
            _d0_sum = st.session_state.get(f'_std_curve_d0_summary{_gk_sfx}')
            if _d0_df is not None and len(_d0_df) > 0:
                _xl_charts[f'Std Curve D0{_gk_label}'[:31]] = {
                    'data': _d0_df, 'summary': _d0_sum or {},
                    'chart_type': 'std_curve_d0',
                }
            _ct_df = st.session_state.get(f'_std_curve_ct_df{_gk_sfx}')
            _ct_sum = st.session_state.get(f'_std_curve_ct_summary{_gk_sfx}')
            if _ct_df is not None and len(_ct_df) > 0:
                _xl_charts[f'Std Curve Ct{_gk_label}'[:31]] = {
                    'data': _ct_df, 'summary': _ct_sum or {},
                    'chart_type': 'std_curve_ct',
                }
        _dil_df = st.session_state.get('_dilution_series_df')
        _dil_sum = st.session_state.get('_dilution_series_summary')
        if _dil_df is not None and len(_dil_df) > 0:
            _xl_charts['Dilution Series'] = {
                'data': _dil_df, 'summary': _dil_sum or {},
                'chart_type': 'dilution_series',
            }

        # ── Input Data sheet: raw fluorescence (cycles as rows, wells as cols) ──
        _xl_all_samples = st.session_state.get('batch_all_samples', {})
        _xl_cycles = st.session_state.get('batch_cycles')
        if _xl_cycles is not None and _xl_all_samples:
            _input_data = {'Cycle': _xl_cycles}
            # Include fitted wells from batch_all_samples
            for _wn, _wd in _xl_all_samples.items():
                _arr = np.asarray(_wd)
                if _arr.ndim >= 1:
                    _input_data[_wn] = _arr[:len(_xl_cycles)]
            # Include no-signal wells from saved fluor or upload source
            try:
                _xl_ns_fluor = st.session_state.get('batch_no_signal_fluor', {})
                _xl_upload_samples = st.session_state.get('upload_batch_samples', {})
                _xl_no_signal_keys = st.session_state.get('batch_no_signal_samples', {})
                for _ns_name in _xl_no_signal_keys:
                    if _ns_name not in _input_data:
                        _ns_src = _xl_ns_fluor.get(_ns_name)
                        if _ns_src is None:
                            _ns_src = _xl_upload_samples.get(_ns_name)
                        if _ns_src is not None:
                            _arr = np.asarray(_ns_src)
                            if _arr.ndim >= 1:
                                _input_data[_ns_name] = _arr[:len(_xl_cycles)]
            except Exception:
                pass  # export proceeds without no-signal wells in Input Data
            # Also check results_list for fluor_data of wells not yet included
            _xl_results_list = st.session_state.get('batch_results_list', [])
            for _rl in _xl_results_list:
                _rn = _rl.get('Sample', '')
                if _rn and _rn not in _input_data and _rl.get('fluor_data') is not None:
                    _arr = np.asarray(_rl['fluor_data'])
                    if _arr.ndim >= 1:
                        _input_data[_rn] = _arr[:len(_xl_cycles)]
            _xl_extra['Input Data'] = pd.DataFrame(_input_data)

        # ── Metadata sheet: sample_metadata table ──
        _xl_meta = st.session_state.get('sample_metadata')
        if _xl_meta and isinstance(_xl_meta, dict):
            _meta_rows = []
            for _mk, _mv in _xl_meta.items():
                _row = {'Well_Key': _mk}
                if isinstance(_mv, dict):
                    _row.update(_mv)
                _meta_rows.append(_row)
            if _meta_rows:
                _xl_extra['Metadata'] = pd.DataFrame(_meta_rows)

        # ── Settings sheet: batch_settings key-value pairs ──
        _xl_batch_settings = st.session_state.get('batch_settings')
        if _xl_batch_settings:
            _settings_rows = []
            for _sk, _sv in _xl_batch_settings.items():
                if isinstance(_sv, dict):
                    # Flatten nested dicts (e.g. channel_thresholds)
                    for _sk2, _sv2 in _sv.items():
                        _settings_rows.append({
                            'Setting': f'{_sk}.{_sk2}',
                            'Value': str(_sv2) if _sv2 is not None else '',
                        })
                else:
                    _settings_rows.append({
                        'Setting': _sk,
                        'Value': str(_sv) if _sv is not None else '',
                    })
            if _settings_rows:
                _xl_extra['Settings'] = pd.DataFrame(_settings_rows)

        _xl_bytes_top = _build_excel_download(
            _xl_results_top, _xl_extra, chart_sheets=_xl_charts
        )
        _xl_sheet_names = (
            ["Batch Results"] + list(_xl_extra.keys()) + list(_xl_charts.keys())
        )
        _n_sheets = len(_xl_sheet_names)
        st.download_button(
            f"📥 Download Complete Results (.xlsx, {_n_sheets} sheet{'s' if _n_sheets > 1 else ''})",
            _xl_bytes_top,
            "batch_fit_results.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="batch_download_xlsx_top",
            type="primary",
        )
    except Exception as _xl_err:
        # Fallback to CSV if Excel generation fails
        st.warning(f"Excel export unavailable ({_xl_err}). Use CSV instead.")
        csv = st.session_state['batch_results'].to_csv(index=False)
        st.download_button(
            "📥 Download Results (.csv)",
            csv,
            "batch_fit_results.csv",
            "text/csv",
            key="batch_download_csv_fallback",
            type="primary",
        )

    results_df = st.session_state['batch_results']
    results_list = st.session_state.get('batch_results_list', [])

    # Target filter for multiplexed data
    display_df = results_df
    if 'Target' in results_df.columns:
        unique_targets = results_df['Target'].unique().tolist()
        if len(unique_targets) > 1:
            target_filter = st.selectbox(
                "Filter by target:",
                ["All targets"] + unique_targets,
                key="target_display_filter"
            )
            if target_filter != "All targets":
                display_df = results_df[results_df['Target'] == target_filter]

    # Format numeric columns
    format_dict = {
        'Ct': '{:.2f}',
        'Ct_instrument': '{:.2f}',
        'D0': '{:.2e}',
        'k': '{:.6f}',
        'P0': '{:.2e}',
        'F_bg_intercept': '{:.6f}',
        'F_bg_slope': '{:.6f}',
        'R2': '{:.6f}',
        'RMSE': '{:.4f}',
        'NRMSE': '{:.2f}',
        'SSR': '{:.6f}',
        'Known_Copies': '{:.2e}',
        'Copies_D0': '{:.2e}',
        'Copies_Ct': '{:.2e}',
    }

    try:
        st.dataframe(display_df.style.format(format_dict, na_rep='-'), use_container_width=True)
    except Exception as _disp_err:
        st.warning(f"Display formatting error: {_disp_err}")
        st.dataframe(display_df, use_container_width=True)
        
    # Summary statistics
    col1, col2, col3 = st.columns(3)
    successful = results_df['Success'].fillna('').str.contains('✓').sum()
    col1.metric("Successful Fits", f"{successful}/{len(results_list)}")
    if successful > 0:
        _ok_r2 = results_df.loc[results_df['Success'].fillna('').str.contains('✓'), 'R2'].dropna()
        if len(_ok_r2) > 0:
            col2.metric("Mean R²", f"{_ok_r2.mean():.4f}")
            col3.metric("Median R²", f"{_ok_r2.median():.4f}")
        
    # Show no-signal samples if any
    if 'batch_no_signal_samples' in st.session_state and st.session_state['batch_no_signal_samples']:
        no_signal_samples = st.session_state['batch_no_signal_samples']
        with st.expander(f"⚠️ {len(no_signal_samples)} samples skipped (no signal detected)"):
            no_signal_df = pd.DataFrame([
                {
                    'Sample': name,
                    'Reason': info['reason'],
                    'Fluorescence Range': f"{info['F_range']:.4f}",
                    '% of Max on Plate': f"{info['F_range_pct']:.1f}%"
                }
                for name, info in no_signal_samples.items()
            ])
            st.dataframe(no_signal_df, use_container_width=True)

    # ============================================================================
    # CALIBRATION SECTION
    # ============================================================================
    sample_metadata = st.session_state.get('sample_metadata')
    manual_cf_val = st.session_state.get('manual_conversion_factor')
    cal_method = st.session_state.get('calibration_method', 'auto')

    has_standards = (
        sample_metadata is not None and
        any(m.get('Task') == 'STANDARD' for m in sample_metadata.values())
    )

    ld_wells = st.session_state.get('ld_wells', [])
    no_signal_wells = list(st.session_state.get('batch_no_signal_samples', {}).keys())

    if has_standards:
        st.markdown("---")
        st.subheader("📐 Standard Curve Calibration")

        # Determine if we need per-channel standard curves.
        # If Channel column doesn't exist, try to derive from sample names
        # (e.g. FAM_A1 → FAM) — handles cases where selected_channels
        # wasn't set during batch fit.
        if 'Channel' not in results_df.columns:
            _derived_ch = results_df['Sample'].map(_ch)
            _uniq_ch = _derived_ch[_derived_ch != 'default'].unique()
            if len(_uniq_ch) > 1:
                results_df.insert(0, 'Channel', _derived_ch)
        _has_multi_ch = 'Channel' in results_df.columns and results_df['Channel'].nunique() > 1
        _cal_channels = list(results_df['Channel'].unique()) if _has_multi_ch else [None]

        # Also detect per-target standard curves for multiplexed data
        _has_multi_tgt = (
            'Target' in results_df.columns
            and results_df['Target'].nunique() > 1
            and not _has_multi_ch  # targets and channels are mutually exclusive groupings
        )

        # Build iteration list: (channel_or_None, target_or_None) pairs
        if _has_multi_tgt:
            # Per-target: find which targets have standards
            _tgt_with_stds = set()
            for _mk, _mv in (sample_metadata or {}).items():
                if _mv.get('Task') == 'STANDARD':
                    _tgt_with_stds.add(_mv.get('_target', ''))
            # Also check results_df directly
            if 'Task' in results_df.columns and 'Target' in results_df.columns:
                _std_rows = results_df[results_df['Task'] == 'STANDARD']
                _tgt_with_stds.update(_std_rows['Target'].dropna().unique())
            _cal_groups = [(None, t) for t in sorted(_tgt_with_stds) if t]
        else:
            _cal_groups = [(_ch, None) for _ch in _cal_channels]

        st.session_state['_std_curve_channels'] = _cal_channels
        # Store all group keys (channels or targets) for Excel export lookup
        _all_grp_keys = [(_cal_ch or _cal_tgt) for _cal_ch, _cal_tgt in _cal_groups]
        st.session_state['_std_curve_group_keys'] = _all_grp_keys

        _any_cal_succeeded = False
        _per_ch_cals = {}  # channel/target → (calibration, ct_calibration)

        for _cal_ch, _cal_tgt in _cal_groups:
            # Filter results and metadata to this channel or target
            if _cal_ch is not None:
                _ch_mask = results_df['Channel'] == _cal_ch
                _ch_df = results_df[_ch_mask].copy()
                _ch_meta = {k: v for k, v in (sample_metadata or {}).items()
                            if k.startswith(f"{_cal_ch}_")}
            elif _cal_tgt is not None:
                _ch_mask = results_df['Target'] == _cal_tgt
                _ch_df = results_df[_ch_mask].copy()
                # Filter metadata to keys belonging to this target
                _ch_meta = {k: v for k, v in (sample_metadata or {}).items()
                            if v.get('_target') == _cal_tgt
                            or k.startswith(f"{_cal_tgt}::")}
            else:
                _ch_df = results_df
                _ch_meta = sample_metadata

            calibration = build_standard_curve(_ch_df, _ch_meta)
            ct_calibration = build_ct_standard_curve(_ch_df, _ch_meta)
            _grp_key = _cal_ch or _cal_tgt
            _per_ch_cals[_grp_key] = (calibration, ct_calibration)

            if calibration is None and ct_calibration is None:
                continue
            _any_cal_succeeded = True

            _ch_label = f" ({_cal_ch})" if _cal_ch else (f" ({_cal_tgt})" if _cal_tgt else "")

            # ── Enrich standard curve data with D0_single predicted copies ──
            _d0s_for_std = None
            _d0s_per_tgt = st.session_state.get('d0_single_per_target', {})
            _d0s_global_v = st.session_state.get('d0_single_value')
            if cal_method == 'd0_single':
                if _cal_tgt and _d0s_per_tgt.get(_cal_tgt, 0) > 0:
                    _d0s_for_std = _d0s_per_tgt[_cal_tgt]
                elif _d0s_per_tgt:
                    # Fallback: find any matching target
                    for _st_tgt, _st_val in _d0s_per_tgt.items():
                        if _st_val and _st_val > 0:
                            _d0s_for_std = _st_val
                            break
                elif _d0s_global_v and _d0s_global_v > 0:
                    _d0s_for_std = _d0s_global_v

            if _d0s_for_std and _d0s_for_std > 0:
                if calibration is not None:
                    _ppd = calibration['per_point_data']
                    _ppd['Predicted_Copies'] = _ppd['D0'] / _d0s_for_std
                    _ppd['log10_Predicted'] = np.log10(_ppd['Predicted_Copies'])
                if ct_calibration is not None:
                    _ct_ppd = ct_calibration['per_point_data']
                    if 'D0' not in _ct_ppd.columns:
                        # Ct per_point_data may not have D0; look it up from results
                        _ct_d0_map = dict(zip(results_df['Sample'], results_df['D0']))
                        _ct_ppd['D0'] = _ct_ppd['Well'].map(_ct_d0_map)
                    _valid_d0 = _ct_ppd['D0'].notna() & (_ct_ppd['D0'] > 0)
                    _ct_ppd.loc[_valid_d0, 'Predicted_Copies'] = (
                        _ct_ppd.loc[_valid_d0, 'D0'] / _d0s_for_std
                    )

            # Determine side-by-side layout
            both_succeeded = calibration is not None and ct_calibration is not None
            if both_succeeded:
                col_d0, col_ct = st.columns(2)
            else:
                col_d0 = st.container() if calibration is not None else None
                col_ct = st.container() if ct_calibration is not None else None

            # ── D0 Calibration (left column) ──────────────────────
            if calibration is not None and col_d0 is not None:
                with col_d0:
                    st.markdown(f"#### D0-Based Standard Curve{_ch_label}")
                    m1, m2 = st.columns(2)
                    m1.metric("Slope", f"{calibration['slope']:.4f}")
                    m2.metric("R\u00b2", f"{calibration['r_squared']:.6f}")

                    st.markdown(
                        f"**log\u2081\u2080(copies) = {calibration['slope']:.4f} "
                        f"\u00d7 log\u2081\u2080(D0) + {calibration['intercept']:.4f}**"
                    )
                    st.caption(
                        f"Standards: {calibration['n_standards']} wells, "
                        f"{calibration['n_concentrations']} levels"
                    )
                    cf_spread = calibration.get('cf_spread', np.nan)
                    if not np.isnan(cf_spread):
                        st.caption(
                            f"Median CF = {calibration['median_cf']:.2e} "
                            f"(spread: {cf_spread:.2f}\u00d7 across levels)"
                        )
                    if not np.isnan(calibration['pooled_replicate_cv']):
                        st.caption(
                            f"Replicate pooled CV = {calibration['pooled_replicate_cv']:.1f}%"
                        )

                    for w in calibration.get('warnings', []):
                        st.warning(w)

                    fig_cal = plot_calibration(calibration, channel_label=_ch_label)

                    # Add predicted copies from D0_single if available
                    _ppd = calibration['per_point_data']
                    if 'Predicted_Copies' in _ppd.columns and _d0s_for_std:
                        import plotly.graph_objects as go
                        fig_cal.add_trace(go.Scatter(
                            x=_ppd['log10_D0'],
                            y=_ppd['log10_Predicted'],
                            mode='markers',
                            name='D0_single predicted',
                            marker=dict(size=10, color='red', symbol='x',
                                        line=dict(width=2)),
                            hovertemplate=(
                                'Well: %{customdata[0]}<br>'
                                'D0: %{customdata[1]:.4e}<br>'
                                'Known: %{customdata[2]:.2e}<br>'
                                'Predicted: %{customdata[3]:.2e}'
                                '<extra></extra>'
                            ),
                            customdata=_ppd[['Well', 'D0', 'Known_Copies', 'Predicted_Copies']].values,
                        ))
                        fig_cal.update_layout(
                            title=dict(text=(
                                fig_cal.layout.title.text or ''
                            ) + f'<br><sub>D0_single = {_d0s_for_std:.3e}</sub>'),
                        )

                    st.plotly_chart(fig_cal, use_container_width=True)

                    # Store D0 standard curve data for Excel export
                    _d0_key_suffix = f"_{_grp_key}" if _grp_key else ""
                    st.session_state[f'_std_curve_d0_df{_d0_key_suffix}'] = calibration['per_point_data'].copy()
                    st.session_state[f'_std_curve_d0_summary{_d0_key_suffix}'] = {
                        'slope': calibration['slope'],
                        'intercept': calibration['intercept'],
                        'r_squared': calibration['r_squared'],
                        'n_standards': calibration['n_standards'],
                        'n_concentrations': calibration['n_concentrations'],
                        'median_cf': calibration.get('median_cf', np.nan),
                    }

            # ── Ct Calibration (right column) ─────────────────────
            if ct_calibration is not None and col_ct is not None:
                with col_ct:
                    st.markdown(f"#### Ct-Based Standard Curve{_ch_label}")
                    m1, m2 = st.columns(2)
                    m1.metric("Efficiency", f"{ct_calibration['efficiency']*100:.1f}%")
                    m2.metric("R\u00b2", f"{ct_calibration['r_squared']:.4f}")

                    st.markdown(
                        f"**log\u2081\u2080(copies) = {ct_calibration['slope']:.4f} \u00d7 Ct "
                        f"+ {ct_calibration['intercept']:.4f}**"
                    )
                    _ct_src = ct_calibration.get('ct_source', 'MAK2+ (calculated)')
                    st.caption(
                        f"Standards: {ct_calibration['n_standards']} wells, "
                        f"{ct_calibration['n_concentrations']} levels  \n"
                        f"Ct source: {_ct_src}"
                    )
                    if not np.isnan(ct_calibration['pooled_replicate_cv']):
                        st.caption(
                            f"Replicate pooled Ct CV = {ct_calibration['pooled_replicate_cv']:.1f}%"
                        )

                    for w in ct_calibration.get('warnings', []):
                        st.warning(w)

                    fig_ct = plot_ct_calibration(ct_calibration, channel_label=_ch_label)
                    st.plotly_chart(fig_ct, use_container_width=True)

                    # Store Ct standard curve data for Excel export
                    _ct_key_suffix = f"_{_grp_key}" if _grp_key else ""
                    st.session_state[f'_std_curve_ct_df{_ct_key_suffix}'] = ct_calibration['per_point_data'].copy()
                    st.session_state[f'_std_curve_ct_summary{_ct_key_suffix}'] = {
                        'slope': ct_calibration['slope'],
                        'intercept': ct_calibration['intercept'],
                        'r_squared': ct_calibration['r_squared'],
                        'efficiency': ct_calibration['efficiency'],
                        'n_standards': ct_calibration['n_standards'],
                        'n_concentrations': ct_calibration['n_concentrations'],
                    }

            # ── Per-level replicate variance ───────────────────────
            with st.expander(f"📊 Replicate variance by concentration level{_ch_label}"):
                var_data = []
                if calibration is not None:
                    for copies_val, var_info in sorted(
                        calibration['replicate_variance'].items(), reverse=True
                    ):
                        row = {
                            'Known Copies': f"{copies_val:.2e}",
                            'N Replicates': var_info['n_replicates'],
                            'Mean D0': f"{var_info['mean_D0']:.4e}",
                            'SD D0': f"{var_info['sd_D0']:.4e}" if not np.isnan(var_info['sd_D0']) else '-',
                            'D0 CV%': f"{var_info['cv_D0_pct']:.1f}" if not np.isnan(var_info['cv_D0_pct']) else '-',
                        }
                        if _d0s_for_std and _d0s_for_std > 0:
                            _pred_copies = var_info['mean_D0'] / _d0s_for_std
                            row['Predicted Copies'] = f"{_pred_copies:.2e}"
                            _ratio = _pred_copies / copies_val if copies_val > 0 else np.nan
                            row['Pred/Known'] = f"{_ratio:.2f}" if not np.isnan(_ratio) else '-'
                        if ct_calibration is not None:
                            ct_var = ct_calibration['replicate_variance'].get(copies_val)
                            if ct_var:
                                row['Mean Ct'] = f"{ct_var['mean_Ct']:.2f}"
                                row['SD Ct'] = f"{ct_var['sd_Ct']:.3f}" if not np.isnan(ct_var['sd_Ct']) else '-'
                                row['Ct CV%'] = f"{ct_var['cv_Ct_pct']:.2f}" if not np.isnan(ct_var['cv_Ct_pct']) else '-'
                        var_data.append(row)
                _var_df = pd.DataFrame(var_data)
                st.dataframe(_var_df, use_container_width=True)
                # Store for Excel export
                if len(_var_df) > 0:
                    _var_key_suffix = f"_{_grp_key}" if _grp_key else ""
                    st.session_state[f'_std_curve_variance_df{_var_key_suffix}'] = _var_df

            # ── Apply calibration (only when auto mode) ──
            if cal_method == 'auto':
                # Apply to the appropriate subset (channel, target, or all)
                if _cal_ch is not None:
                    _apply_mask = results_df['Channel'] == _cal_ch
                elif _cal_tgt is not None:
                    _apply_mask = results_df['Target'] == _cal_tgt
                else:
                    _apply_mask = None  # apply to all

                if _apply_mask is not None:
                    if calibration is not None:
                        cal_subset = apply_calibration(results_df[_apply_mask].copy(), calibration=calibration)
                        results_df.loc[_apply_mask, 'Copies_D0'] = cal_subset['Copies_D0']
                    if ct_calibration is not None:
                        ct_subset = apply_ct_calibration(results_df[_apply_mask].copy(), ct_calibration)
                        results_df.loc[_apply_mask, 'Copies_Ct'] = ct_subset['Copies_Ct']
                else:
                    if calibration is not None:
                        results_df = apply_calibration(results_df, calibration=calibration)
                    if ct_calibration is not None:
                        results_df = apply_ct_calibration(results_df, ct_calibration)

        if not _any_cal_succeeded:
            st.warning("Could not build standard curve (need ≥ 2 concentration levels with successful fits)")
            if cal_method == 'auto' and manual_cf_val and manual_cf_val > 0:
                results_df = apply_calibration(results_df, manual_cf=manual_cf_val)
                st.info(f"Using manual conversion factor: {manual_cf_val:.2e} copies/D0")

        if cal_method == 'auto':
            st.session_state['batch_results'] = results_df

    if cal_method == 'limited_dilution' and len(ld_wells) >= 3:
        st.markdown("---")
        st.subheader("📐 Limited Dilution Calibration")

        ld_calibration = build_limited_dilution_calibration(
            results_df=results_df,
            ld_wells=ld_wells,
            no_signal_wells=no_signal_wells,
        )

        if ld_calibration is None or 'conversion_factor' not in ld_calibration:
            st.error("Limited dilution calibration failed.")
            if ld_calibration and ld_calibration.get('warnings'):
                for w in ld_calibration['warnings']:
                    st.warning(w)
        else:
            # Display metrics
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("\u03bb (copies/well)", f"{ld_calibration['lambda_hat']:.3f}")
            col2.metric("Conversion Factor", f"{ld_calibration['conversion_factor']:.2e}")
            col3.metric("Positive Wells",
                        f"{ld_calibration['n_positive']}/{ld_calibration['n_total']}")
            col4.metric("D0 (1 copy)", f"{ld_calibration['d0_single_copy']:.3e}")

            # CF with CI
            if not np.isnan(ld_calibration['cf_ci_lower']):
                st.markdown(
                    f"**Conversion Factor:** {ld_calibration['conversion_factor']:.2e} copies/D0 "
                    f"(95% CI: {ld_calibration['cf_ci_lower']:.2e} \u2013 "
                    f"{ld_calibration['cf_ci_upper']:.2e})"
                )
            else:
                st.markdown(
                    f"**Conversion Factor:** {ld_calibration['conversion_factor']:.2e} copies/D0"
                )

            st.markdown(
                f"**Poisson rate (\u03bb):** {ld_calibration['lambda_hat']:.3f} copies/well "
                f"(SE: {ld_calibration['lambda_se']:.3f})"
            )
            st.markdown(
                f"**Expected copies per positive well:** "
                f"{ld_calibration['expected_copies_per_positive']:.3f}"
            )

            # Warnings
            for w in ld_calibration.get('warnings', []):
                st.warning(w)

            # Diagnostic plots
            fig_ld = plot_limited_dilution_diagnostics(ld_calibration)
            st.plotly_chart(fig_ld, use_container_width=True)

            # Positive well D0 details
            with st.expander("🔬 Positive well D0 details"):
                pos_wells_list = ld_calibration['positive_wells']
                pos_details = results_df[
                    results_df['Sample'].isin(pos_wells_list)
                ][['Sample', 'D0', 'R2']].copy()
                pos_details['Est. Copies'] = (
                    pos_details['D0'] * ld_calibration['conversion_factor']
                )
                st.dataframe(pos_details, use_container_width=True)
                st.session_state['_limited_dilution_df'] = pos_details

                st.markdown(
                    f"**D0 statistics:** mean = {ld_calibration['mean_d0_positive']:.3e}, "
                    f"median = {ld_calibration['median_d0_positive']:.3e}"
                )
                if not np.isnan(ld_calibration['cv_d0_positive']):
                    st.markdown(
                        f"**CV:** {ld_calibration['cv_d0_positive']:.1f}%"
                    )

            # Apply calibration — feed CF into apply_calibration via manual_cf
            results_df = apply_calibration(
                results_df, manual_cf=ld_calibration['conversion_factor']
            )
            st.session_state['batch_results'] = results_df
            st.success("✅ Copies_D0 column added from limited dilution calibration")

    elif cal_method == 'manual_cf' and manual_cf_val and manual_cf_val > 0:
        # Manual CF provided
        st.markdown("---")
        st.subheader("📐 Copy Number Conversion (Manual)")
        results_df = apply_calibration(results_df, manual_cf=manual_cf_val)
        st.session_state['batch_results'] = results_df
        st.info(f"Applied manual conversion factor: {manual_cf_val:.2e} copies/D0")

    elif cal_method == 'auto' and manual_cf_val and manual_cf_val > 0:
        # Auto mode but no standards found — use fallback manual CF
        st.markdown("---")
        st.subheader("📐 Copy Number Conversion (Manual Fallback)")
        results_df = apply_calibration(results_df, manual_cf=manual_cf_val)
        st.session_state['batch_results'] = results_df
        st.info(f"No standards detected. Applied fallback manual CF: {manual_cf_val:.2e} copies/D0")

    elif cal_method == 'd0_single':
        _d0s_per_target = st.session_state.get('d0_single_per_target', {})
        _d0s_global = st.session_state.get('d0_single_value')
        _has_targets_col = 'Target' in results_df.columns

        # Build list of (target_name_or_None, d0_single_value) to apply
        _d0s_to_apply = []
        if _has_targets_col and _d0s_per_target:
            for _tgt, _d0s in _d0s_per_target.items():
                if _d0s and _d0s > 0:
                    _d0s_to_apply.append((_tgt, _d0s))
        elif _d0s_global and _d0s_global > 0:
            _d0s_to_apply.append((None, _d0s_global))

        if _d0s_to_apply:
            # Apply calibration
            for _tgt, _d0s in _d0s_to_apply:
                _cf = 1.0 / _d0s
                if _tgt is not None:
                    _tgt_mask = results_df['Target'] == _tgt
                    if _tgt_mask.any():
                        _tgt_cal = apply_calibration(
                            results_df[_tgt_mask].copy(), manual_cf=_cf
                        )
                        results_df.loc[_tgt_mask, 'Copies_D0'] = _tgt_cal['Copies_D0']
                else:
                    results_df = apply_calibration(results_df, manual_cf=_cf)
            st.session_state['batch_results'] = results_df

    # ── Sync Copies_D0 back to batch_results_list (used by visualizations) ──
    if 'Copies_D0' in results_df.columns:
        _sync_list = st.session_state.get('batch_results_list', [])
        if _sync_list:
            _sample_to_copies = dict(
                zip(results_df['Sample'], results_df['Copies_D0'])
            )
            for _rl_item in _sync_list:
                _sn = _rl_item.get('Sample', '')
                if _sn in _sample_to_copies:
                    _rl_item['Copies_D0'] = _sample_to_copies[_sn]
            st.session_state['batch_results_list'] = _sync_list

    # CSV export (always available immediately)
    csv = results_df.to_csv(index=False)
    st.download_button(
        "📥 Download Results (.csv)",
        csv,
        "batch_fit_results.csv",
        "text/csv",
        key="batch_download_csv"
    )

    # ============================================================================
    # REPLICATE ANALYSIS SECTION
    # ============================================================================
    if st.session_state.get('replicate_analysis_enabled', False):
        st.markdown("---")
        st.subheader("📊 Replicate Analysis")

        # Get grouping pattern from session state
        pattern = st.session_state.get('grouping_pattern', 'dot')

        # Parse sample groups — use Sample Name from metadata when available
        sample_metadata_for_groups = st.session_state.get('sample_metadata')
        has_multi_targets = 'Target' in results_df.columns and results_df['Target'].nunique() > 1
        if sample_metadata_for_groups and pattern == 'sample_name':
            # Group by Sample Name from instrument metadata
            name_groups = {}
            for key, meta in sample_metadata_for_groups.items():
                sname = meta.get('Sample Name')
                if sname and str(sname) != 'nan' and str(sname).strip():
                    sname = str(sname).strip()
                    # For multi-target, prefix group name with target
                    if has_multi_targets and meta.get('_target'):
                        group_label = f"{meta['_target']} — {sname}"
                    else:
                        group_label = sname
                    if key in results_df['Sample'].values:
                        if group_label not in name_groups:
                            name_groups[group_label] = []
                        name_groups[group_label].append(key)
            # Keep only groups with > 1 replicate
            sample_groups = {k: v for k, v in name_groups.items() if len(v) > 1}
            if sample_groups:
                st.info("Grouping by Sample Name from instrument metadata")
            else:
                # Fall back to pattern-based grouping
                sample_groups = parse_sample_groups(
                    results_df['Sample'].tolist(), pattern='dot'
                )
        elif pattern == 'manual':
            # Use manually defined groups from session state
            sample_groups = {}
            manual_groups_text = st.session_state.get('manual_groups_text', '')
            if manual_groups_text.strip():
                for line in manual_groups_text.strip().split('\n'):
                    if ':' in line:
                        group_name, samples_str = line.split(':', 1)
                        group_name = group_name.strip()
                        samples = [s.strip() for s in samples_str.split(',')]
                        # Validate that samples exist in results
                        valid_samples = [s for s in samples if s in results_df['Sample'].tolist()]
                        if valid_samples:
                            sample_groups[group_name] = valid_samples
        else:
            sample_groups = parse_sample_groups(
                results_df['Sample'].tolist(),
                pattern=pattern
            )

        if len(sample_groups) == 0:
            st.warning("⚠️ No replicate groups found with the selected pattern")
        else:
            # Add Group column to results
            results_with_groups = results_df.copy()
            group_mapping = {}
            for group, samples in sample_groups.items():
                for sample in samples:
                    group_mapping[sample] = group

            results_with_groups['Group'] = results_with_groups['Sample'].map(
                lambda x: group_mapping.get(x, x)
            )

            # Filter to only samples that are in groups
            results_with_groups = results_with_groups[
                results_with_groups['Group'] != results_with_groups['Sample']
            ]

            if len(results_with_groups) == 0:
                st.warning("⚠️ No samples matched the grouping pattern")
            else:
                st.success(f"✅ Analyzed {len(sample_groups)} replicate groups ({len(results_with_groups)} samples total)")

                # Determine which metrics to track (only include columns that exist)
                rep_metrics = []
                for m in ['Ct', 'D0']:
                    if m in results_with_groups.columns:
                        rep_metrics.append(m)
                if 'Copies_D0' in results_with_groups.columns and results_with_groups['Copies_D0'].notna().any():
                    rep_metrics.append('Copies_D0')
                if 'Copies_Ct' in results_with_groups.columns and results_with_groups['Copies_Ct'].notna().any():
                    rep_metrics.append('Copies_Ct')

                # Calculate replicate statistics
                replicate_stats = calculate_replicate_stats(
                    results_with_groups,
                    group_column='Group',
                    metrics=rep_metrics
                )

                # Precision comparison (use default efficiency unless dilution series provides better estimate)
                # This will be updated if dilution series is analyzed
                precision_comparison = compare_precision(replicate_stats, efficiency=0.95)

                # Store for Excel export
                st.session_state['_replicate_stats_df'] = replicate_stats
                st.session_state['_precision_comparison_df'] = precision_comparison

                # Display tabs for different analyses
                tab1, tab2, tab3, tab4 = st.tabs([
                    "📈 Replicate Statistics",
                    "🔬 Replicate Visualization",
                    "🎯 Precision Comparison",
                    "📉 Dilution Series"
                ])

                with tab1:
                    st.markdown("**Replicate Statistics (Mean ± SD)**")
                    st.markdown("*Coefficient of Variation (CV%) = (SD / Mean) × 100*")

                    # Format for display (convert None to NaN for proper na_rep display)
                    display_stats = replicate_stats.fillna(value=np.nan).copy()

                    # Add Wells column so user can identify which wells belong to each group
                    def _get_well_ids(group_name):
                        keys = sample_groups.get(group_name, [])
                        # Extract well ID: "Target::Well" -> "Well", or just use key as-is
                        wells = [k.split('::')[-1] if '::' in k else k for k in keys]
                        return ', '.join(sorted(wells))

                    display_stats.insert(
                        1, 'Wells',
                        display_stats['Group'].apply(_get_well_ids)
                    )

                    format_stats_dict = {
                        'Ct_Mean': '{:.2f}',
                        'Ct_SD': '{:.3f}',
                        'Ct_CV': '{:.2f}%',
                        'Ct_Min': '{:.2f}',
                        'Ct_Max': '{:.2f}',
                        'D0_Mean': '{:.2e}',
                        'D0_SD': '{:.2e}',
                        'D0_CV': '{:.2f}%',
                        'D0_Min': '{:.2e}',
                        'D0_Max': '{:.2e}',
                        'Copies_D0_Mean': '{:.2e}',
                        'Copies_D0_SD': '{:.2e}',
                        'Copies_D0_CV': '{:.2f}%',
                        'Copies_D0_Min': '{:.2e}',
                        'Copies_D0_Max': '{:.2e}',
                        'Copies_Ct_Mean': '{:.2e}',
                        'Copies_Ct_SD': '{:.2e}',
                        'Copies_Ct_CV': '{:.2f}%',
                        'Copies_Ct_Min': '{:.2e}',
                        'Copies_Ct_Max': '{:.2e}',
                    }

                    st.dataframe(
                        display_stats.style.format(format_stats_dict, na_rep='-'),
                        use_container_width=True
                    )

                    # Download button for stats
                    stats_csv = replicate_stats.to_csv(index=False)
                    st.download_button(
                        "📥 Download Replicate Statistics (CSV)",
                        stats_csv,
                        "replicate_statistics.csv",
                        "text/csv",
                        key="replicate_stats_download"
                    )

                with tab2:
                    st.markdown("### Replicate Overlay")
                    st.markdown("Select a replicate group to overlay all replicates on one plot.")

                    group_names = sorted(sample_groups.keys())
                    selected_group = st.selectbox(
                        "Select replicate group:",
                        group_names,
                        key="replicate_group_selector"
                    )

                    if selected_group:
                        group_wells = sample_groups[selected_group]

                        # Plot overlay
                        batch_cycles = st.session_state.get('batch_cycles')
                        batch_all_samples = st.session_state.get('batch_all_samples', {})
                        batch_results_list = st.session_state.get('batch_results_list', [])

                        if batch_cycles is not None and batch_all_samples:
                            fig_overlay = plot_replicate_overlay(
                                cycles=batch_cycles,
                                all_samples=batch_all_samples,
                                results_list=batch_results_list,
                                group_wells=group_wells,
                                group_name=selected_group
                            )
                            st.plotly_chart(fig_overlay, use_container_width=True)

                            # Parameter summary table (mean +/- SE)
                            st.markdown("**Parameter Summary (Mean ± SE)**")
                            param_summary = calculate_replicate_param_summary(
                                results_df, group_wells
                            )
                            # Format for display
                            def format_mean_se(row):
                                mean = row['Mean']
                                se = row['SE']
                                n = int(row['N'])
                                if pd.isna(mean):
                                    return '-'
                                if pd.isna(se):
                                    return f'{mean:.4e} (n={n})'
                                return f'{mean:.4e} ± {se:.4e} (n={n})'

                            summary_display = param_summary.apply(format_mean_se, axis=1)
                            summary_display.name = 'Mean ± SE (n)'
                            st.dataframe(summary_display.to_frame(), use_container_width=True)
                        else:
                            st.warning("Batch fitting data not available for visualization.")

                with tab3:
                    st.markdown("**Precision Comparison: Ct vs D0**")
                    st.markdown("*Lower CV% = Better precision*")

                    # Show which efficiency is being used
                    if st.session_state.get('analyze_as_dilution', False) and len(sample_groups) >= 3:
                        st.info("ℹ️ **Note:** Ct values are logarithmic, so Ct CV% is converted to concentration CV% using efficiency from dilution series.")
                    else:
                        st.info("ℹ️ **Note:** Ct values are logarithmic, so Ct CV% is converted to concentration CV% assuming 95% efficiency. Enable dilution series analysis for actual efficiency.")

                    # Add color highlighting for better method
                    def highlight_better_precision(row):
                        if pd.isna(row['Ct_Conc_CV']) or pd.isna(row['D0_CV']):
                            return [''] * len(row)

                        colors = [''] * len(row)
                        ct_conc_cv_idx = precision_comparison.columns.get_loc('Ct_Conc_CV')
                        d0_cv_idx = precision_comparison.columns.get_loc('D0_CV')

                        if row['Better_Precision'] == 'Ct':
                            colors[ct_conc_cv_idx] = 'background-color: lightgreen'
                        else:
                            colors[d0_cv_idx] = 'background-color: lightgreen'

                        return colors

                    format_precision_dict = {
                        'Ct_Mean': '{:.2f}',
                        'Ct_SD': '{:.3f}',
                        'Ct_CV': '{:.2f}%',
                        'Ct_Conc_CV': '{:.2f}%',
                        'D0_Mean': '{:.2e}',
                        'D0_CV': '{:.2f}%',
                        'CV_Difference': '{:.2f}%'
                    }

                    st.dataframe(
                        precision_comparison.style.format(format_precision_dict, na_rep='-').apply(
                            highlight_better_precision, axis=1
                        ),
                        use_container_width=True
                    )

                    # Summary statistics
                    ct_wins = (precision_comparison['Better_Precision'] == 'Ct').sum()
                    d0_wins = (precision_comparison['Better_Precision'] == 'D0').sum()

                    col1, col2, col3 = st.columns(3)
                    col1.metric("Ct Better", f"{ct_wins}/{len(precision_comparison)}")
                    col2.metric("D0 Better", f"{d0_wins}/{len(precision_comparison)}")
                    avg_ct_conc_cv = precision_comparison['Ct_Conc_CV'].mean()
                    avg_d0_cv = precision_comparison['D0_CV'].mean()
                    col3.metric("Avg CV Diff", f"{abs(avg_ct_conc_cv - avg_d0_cv):.2f}%")

                    st.markdown("---")
                    st.markdown("**Interpretation:**")
                    if d0_wins > ct_wins:
                        st.success(f"✅ **D0 quantification** shows better precision in {d0_wins}/{len(precision_comparison)} groups")
                    elif ct_wins > d0_wins:
                        st.info(f"ℹ️ **Ct method** shows better precision in {ct_wins}/{len(precision_comparison)} groups")
                    else:
                        st.info("ℹ️ Both methods show similar precision across groups")

                    # Download button
                    precision_csv = precision_comparison.to_csv(index=False)
                    st.download_button(
                        "📥 Download Precision Comparison (CSV)",
                        precision_csv,
                        "precision_comparison.csv",
                        "text/csv",
                        key="precision_comparison_download"
                    )

                with tab4:
                    # Check if dilution series analysis is enabled
                    if st.session_state.get('analyze_as_dilution', False):
                        st.markdown("**Dilution Series Analysis**")

                        if len(sample_groups) < 3:
                            st.warning("⚠️ Need at least 3 dilution levels for analysis")
                        else:
                            # Get dilution factor from session state
                            dilution_factor = st.session_state.get('dilution_factor', 2)
                            custom_dilution_factors = st.session_state.get('custom_dilution_factors', None)
                            groups_to_exclude = st.session_state.get('exclude_from_dilution', [])

                            # Filter out excluded groups before analysis
                            if groups_to_exclude:
                                results_filtered = results_with_groups[
                                    ~results_with_groups['Group'].isin(groups_to_exclude)
                                ].copy()
                                st.info(f"📝 Excluding {len(groups_to_exclude)} group(s) from dilution analysis: {', '.join(groups_to_exclude)}")
                            else:
                                results_filtered = results_with_groups

                            # Run dilution series analysis
                            dilution_analysis = analyze_dilution_series(
                                results_filtered,
                                dilution_factors=custom_dilution_factors,  # Use custom if provided, else auto-generate
                                dilution_factor=dilution_factor if custom_dilution_factors is None else 2,
                                group_column='Group'
                            )

                            if 'error' in dilution_analysis:
                                st.error(dilution_analysis['error'])
                            else:
                                # Use efficiency from dilution series to recalculate precision comparison
                                ct_efficiency = dilution_analysis['ct_analysis']['efficiency'] / 100  # Convert % to fraction
                                precision_comparison = compare_precision(replicate_stats, efficiency=ct_efficiency)
                                # Update stored copies for Excel export
                                st.session_state['_precision_comparison_df'] = precision_comparison

                                # Display results
                                col1, col2 = st.columns(2)

                                with col1:
                                    st.markdown("**Ct Analysis**")
                                    ct_analysis = dilution_analysis['ct_analysis']
                                    st.metric("R²", f"{ct_analysis['r2']:.4f}")
                                    st.metric("Efficiency", f"{ct_analysis['efficiency']:.1f}%")
                                    st.metric("Slope", f"{ct_analysis['slope']:.4f}")

                                with col2:
                                    st.markdown("**D0 Analysis**")
                                    d0_analysis = dilution_analysis['d0_analysis']
                                    st.metric("R²", f"{d0_analysis['r2']:.4f}")
                                    st.metric("Slope", f"{d0_analysis['slope']:.4f}")
                                    st.caption(f"Expected: {d0_analysis['expected_slope']:.4f}")

                                # Plot comparison
                                st.markdown("---")
                                st.markdown("**Linearity Comparison**")

                                # Debug: Show error bar values
                                with st.expander("🔍 Debug: Error Bar Values"):
                                    data_debug = dilution_analysis['data']
                                    st.write("**Ct Standard Deviations:**")
                                    st.write(data_debug[['Group', 'Ct_Mean', 'Ct_SD']].to_string())
                                    st.write("\n**D0 Standard Deviations:**")
                                    st.write(data_debug[['Group', 'D0_Mean', 'D0_SD']].to_string())

                                # Store dilution data for Excel export
                                _dil_export = dilution_analysis['data'].copy()
                                _dil_export['log10_Dilution'] = np.log10(_dil_export['Dilution'].astype(float))
                                _dil_export['log10_D0_Mean'] = np.log10(_dil_export['D0_Mean'].astype(float))
                                st.session_state['_dilution_series_df'] = _dil_export
                                st.session_state['_dilution_series_summary'] = {
                                    'ct_slope': dilution_analysis['ct_analysis']['slope'],
                                    'ct_intercept': dilution_analysis['ct_analysis']['intercept'],
                                    'ct_r2': dilution_analysis['ct_analysis']['r2'],
                                    'ct_efficiency': dilution_analysis['ct_analysis']['efficiency'],
                                    'd0_slope': dilution_analysis['d0_analysis']['slope'],
                                    'd0_intercept': dilution_analysis['d0_analysis']['intercept'],
                                    'd0_r2': dilution_analysis['d0_analysis']['r2'],
                                    'd0_expected_slope': dilution_analysis['d0_analysis']['expected_slope'],
                                    'better_linearity': dilution_analysis['comparison']['better_linearity'],
                                }

                                dilution_plot = plot_dilution_series_comparison(dilution_analysis)
                                st.plotly_chart(dilution_plot, use_container_width=True)

                                # Summary
                                comparison = dilution_analysis['comparison']
                                if comparison['better_linearity'] == 'D0':
                                    st.success(f"✅ **D0 shows better linearity** (R² = {comparison['d0_r2']:.4f} vs {comparison['ct_r2']:.4f})")
                                else:
                                    st.info(f"ℹ️ **Ct shows better linearity** (R² = {comparison['ct_r2']:.4f} vs {comparison['d0_r2']:.4f})")
                    else:
                        st.info("💡 Enable 'Analyze as dilution series' in the sidebar to see linearity analysis")

# Batch visualization section (outside button block, always visible if results exist)
if 'batch_results_list' in st.session_state and 'batch_all_samples' in st.session_state:
    st.markdown("---")
    st.subheader("📊 Visualize Individual Fits")
        
    results_list = st.session_state['batch_results_list']
    all_samples = st.session_state['batch_all_samples']
    cycles = st.session_state['batch_cycles']
    batch_settings = st.session_state['batch_settings']
    sample_metadata = st.session_state.get('sample_metadata')

    # Sample selector
    sample_names = [r['Sample'] for r in results_list]
    selected_sample = st.selectbox(
        "Select sample to visualize:",
        sample_names,
        key="batch_viz_selector"
    )
        
    # Find the selected sample's index
    selected_idx = sample_names.index(selected_sample)
    selected_result = results_list[selected_idx]
        
    # Determine pass/fail status for plot title
    _success_val = selected_result.get('Success', '')
    _error_val = selected_result.get('error')
    _has_fit_params = (
        selected_result.get('D0') is not None
        and not (isinstance(selected_result.get('D0'), float) and np.isnan(selected_result['D0']))
    )
    if _error_val and not (isinstance(_error_val, float) and pd.isna(_error_val)):
        _status_label = f"FAIL — {_error_val}"
        _status_color = "red"
    elif _success_val and str(_success_val).startswith('✓'):
        _status_label = f"PASS ({_success_val})"
        _status_color = "green"
    elif _success_val:
        _status_label = f"PASS ({_success_val})"
        _status_color = "orange"
    else:
        _status_label = "FAIL"
        _status_color = "red"

    # Get fluorescence data — prefer all_samples, fall back to stored fluor_data
    if selected_sample in all_samples:
        sample_fluor = all_samples[selected_sample]
    elif selected_result.get('fluor_data') is not None:
        sample_fluor = np.asarray(selected_result['fluor_data'])
    else:
        sample_fluor = None

    try:
        # Generate fit prediction if parameters are available
        F_pred = None
        eff_batch = None
        if _has_fit_params:
            model_viz = MAK2Model()
            F_pred = model_viz.simulate_to_cycle(
                D0=selected_result['D0'],
                k=selected_result['k'],
                P0=selected_result['P0'],
                cycles=cycles,
                F_bg_intercept=selected_result['F_bg_intercept'],
                F_bg_slope=selected_result['F_bg_slope']
            )
            _, D_batch_eff, _ = model_viz.simulate_cycles(
                D0=selected_result['D0'],
                k=selected_result['k'],
                P0=selected_result['P0'],
                n_cycles=len(cycles),
                F_bg_intercept=selected_result['F_bg_intercept'],
                F_bg_slope=selected_result['F_bg_slope']
            )
            eff_batch = calculate_amplification_efficiency(D_batch_eff)

        # Build subplot layout — 3 rows if we have fit, 1 row if data only
        _has_full_viz = F_pred is not None and sample_fluor is not None
        if _has_full_viz:
            fig_batch = make_subplots(
                rows=3, cols=1,
                subplot_titles=(
                    f"<span style='color:{_status_color}'>{_status_label}</span> — {selected_sample}",
                    "Residuals", "Amplification Efficiency",
                ),
                vertical_spacing=0.08,
                row_heights=[0.50, 0.22, 0.28],
                shared_xaxes=True,
            )
        else:
            fig_batch = make_subplots(
                rows=1, cols=1,
                subplot_titles=(
                    f"<span style='color:{_status_color}'>{_status_label}</span> — {selected_sample}",
                ),
            )

        # Row 1: Data markers
        if sample_fluor is not None:
            fig_batch.add_trace(
                go.Scatter(
                    x=cycles[:len(sample_fluor)], y=sample_fluor,
                    mode='markers',
                    name='Data',
                    marker=dict(size=8, color='blue', opacity=0.6)
                ),
                row=1, col=1
            )

        if _has_full_viz:
            # Draw fit line only within the fit window (fit_start_cycle onwards).
            _vis_fit_start = selected_result.get('fit_start_cycle')
            if _vis_fit_start is not None:
                _fit_mask   = cycles >= _vis_fit_start
                _fit_cyc_v  = cycles[_fit_mask]
                _fit_pred_v = F_pred[_fit_mask]
            else:
                _fit_cyc_v  = cycles
                _fit_pred_v = F_pred
            fig_batch.add_trace(
                go.Scatter(
                    x=_fit_cyc_v, y=_fit_pred_v,
                    mode='lines',
                    name='MAK2 Fit',
                    line=dict(color='red', width=2)
                ),
                row=1, col=1
            )

            # Row 2: Residuals — only within the fit window
            residuals = sample_fluor - F_pred
            _res_in_window = residuals[cycles >= _vis_fit_start] if _vis_fit_start is not None else residuals
            fig_batch.add_trace(
                go.Scatter(
                    x=_fit_cyc_v, y=_res_in_window,
                    mode='markers',
                    name='Residuals',
                    marker=dict(size=6, color='green')
                ),
                row=2, col=1
            )
            fig_batch.add_hline(y=0, line_dash="dash", line_color="gray", row=2, col=1)

            # Row 3: Efficiency
            fig_batch.add_trace(
                go.Scatter(
                    x=np.arange(1, len(eff_batch)+1),
                    y=eff_batch,
                    mode='lines+markers',
                    name='Efficiency',
                    marker=dict(size=4),
                    line=dict(color='orange'),
                ),
                row=3, col=1
            )

        # ── Annotations only for full (fitted) visualizations ────────
        _n_viz_rows = 3 if _has_full_viz else 1

        if _has_full_viz:
            # Ct threshold line on fit plot (row 1).
            ch_thresholds_stored = batch_settings.get('channel_thresholds', {})
            sample_ch = _ch(selected_sample) if '_' in selected_sample or '::' in selected_sample else 'default'
            ct_threshold = ch_thresholds_stored.get(
                sample_ch, batch_settings.get('global_threshold'))

            ct_bl_slope_vis     = selected_result.get('Ct_baseline_slope', 0.0)
            ct_bl_intercept_vis = selected_result.get('Ct_baseline_intercept', 0.0)
            ct_baseline_mean    = selected_result.get(
                'Ct_baseline_mean',
                batch_settings.get('channel_baseline_means', {}).get(
                    sample_ch, batch_settings.get('global_baseline_mean', 0.0)
                )
            )

            # ── Fit window: start (orange) and end (green) vlines ────────────
            fit_start_vis = selected_result.get('fit_start_cycle')
            if fit_start_vis is not None:
                for row_idx in range(1, _n_viz_rows + 1):
                    fig_batch.add_vline(
                        x=fit_start_vis,
                        line_dash="dash",
                        line_color="orange",
                        row=row_idx, col=1,
                    )
                fig_batch.add_annotation(
                    x=fit_start_vis, y=1, yref="y domain",
                    text="Fit start", showarrow=False,
                    xanchor="left", yanchor="top",
                    font=dict(size=11, color="orange"),
                    row=1, col=1,
                )

            final_cycle = selected_result.get('fit_end_cycle')
            if final_cycle is None and sample_fluor is not None:
                from mak2_model import find_slope_threshold_cycle
                trunc_idx  = find_slope_threshold_cycle(
                    sample_fluor,
                    cycles_after_max=batch_settings.get('cycles_after_max', 3)
                )
                final_cycle = float(cycles[min(trunc_idx, len(cycles) - 1)])
            if final_cycle is not None and final_cycle < cycles[-1]:
                for row_idx in range(1, _n_viz_rows + 1):
                    fig_batch.add_vline(
                        x=final_cycle,
                        line_dash="dash",
                        line_color="green",
                        row=row_idx, col=1,
                    )
                fig_batch.add_annotation(
                    x=final_cycle, y=1, yref="y domain",
                    text="Fit end", showarrow=False,
                    xanchor="right", yanchor="top",
                    font=dict(size=11, color="green"),
                    row=1, col=1,
                )

            # ── Ct vertical line on all rows ─────────────────────────────────
            ct_val = selected_result.get('Ct', np.nan)
            ct_is_nan = ct_val is None or (isinstance(ct_val, float) and np.isnan(ct_val))
            if not ct_is_nan:
                for row_idx in range(1, _n_viz_rows + 1):
                    fig_batch.add_vline(
                        x=ct_val,
                        line_dash="dot",
                        line_color="gray",
                        row=row_idx, col=1,
                    )
                fig_batch.add_annotation(
                    x=ct_val, y=1, yref="y domain",
                    text="Ct", showarrow=False,
                    xanchor="left", yanchor="top",
                    font=dict(size=11, color="gray"),
                    row=1, col=1,
                )
            else:
                inst_status_vis = selected_result.get('Instrument', '')
                if inst_status_vis:
                    fig_batch.add_annotation(
                        x=0.5, xref="x domain",
                        y=0.95, yref="y domain",
                        text=f"Instrument: {inst_status_vis}",
                        showarrow=False,
                        xanchor="center", yanchor="top",
                        font=dict(size=12, color="crimson"),
                        bgcolor="rgba(255,255,255,0.8)",
                        row=1, col=1,
                    )

            # ── Threshold: short segment anchored at Ct crossing ──────────────
            if ct_threshold is not None and ct_threshold > 0:
                _abi_meta_vis   = st.session_state.get('abi_results_meta')
                _rox_active_vis = st.session_state.get('rox_normalized', False)
                inst_drn = None
                if _abi_meta_vis and _rox_active_vis:
                    inst_drn = _abi_meta_vis.get('channel_thresholds', {}).get(sample_ch)

                has_linear_baseline = (
                    abs(ct_bl_intercept_vis) > 1e-9
                    or abs(ct_bl_slope_vis) > 1e-12
                )

                _span = 5
                if not ct_is_nan:
                    _tx_lo = max(float(cycles[0]),  ct_val - _span)
                    _tx_hi = min(float(cycles[-1]), ct_val + _span)
                else:
                    _tx_lo, _tx_hi = float(cycles[0]), float(cycles[-1])
                _tx = np.linspace(_tx_lo, _tx_hi, 60)

                _ct_rox_mean_vis = selected_result.get('ct_rox_mean')

                if has_linear_baseline:
                    _ty_rn = ct_bl_slope_vis * _tx + ct_bl_intercept_vis + ct_threshold
                else:
                    _ty_rn = np.full_like(_tx, float(ct_baseline_mean + ct_threshold))

                if _ct_rox_mean_vis is not None and _ct_rox_mean_vis > 0:
                    _ty = _ty_rn * _ct_rox_mean_vis
                else:
                    _ty = _ty_rn

                if inst_drn is not None:
                    thresh_label = f"Threshold (ΔRn = {inst_drn})"
                else:
                    thresh_label = f"Threshold (ΔRn = {ct_threshold:.4f})"

                fig_batch.add_trace(
                    go.Scatter(
                        x=_tx, y=_ty,
                        mode='lines',
                        name='Threshold',
                        line=dict(color='purple', dash='dot', width=1.5),
                        showlegend=False,
                    ),
                    row=1, col=1
                )
                fig_batch.add_annotation(
                    x=float(_tx[-1]), y=float(_ty[-1]),
                    text=thresh_label,
                    showarrow=False,
                    xanchor="left", yanchor="middle",
                    font=dict(size=10, color="purple"),
                    row=1, col=1,
                )

        # ── Layout and render ──────────────────────────────────────────
        if _has_full_viz:
            fig_batch.update_xaxes(title_text="Cycle", row=3, col=1)
            fig_batch.update_yaxes(title_text="Fluorescence", row=1, col=1)
            fig_batch.update_yaxes(title_text="Residual", row=2, col=1)
            fig_batch.update_yaxes(title_text="Efficiency", row=3, col=1)
            fig_batch.update_layout(height=800, showlegend=True)
        else:
            fig_batch.update_xaxes(title_text="Cycle", row=1, col=1)
            fig_batch.update_yaxes(title_text="Fluorescence", row=1, col=1)
            fig_batch.update_layout(height=400, showlegend=True)

        st.plotly_chart(fig_batch, use_container_width=True)

        # Show parameters if available
        if _has_fit_params:
            _copies_val = selected_result.get('Copies_D0')
            _show_copies = (
                _copies_val is not None
                and not (isinstance(_copies_val, float) and np.isnan(_copies_val))
            )
            if _show_copies:
                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("D₀", f"{selected_result['D0']:.2e}")
                col2.metric("k", f"{selected_result['k']:.6f}")
                col3.metric("P₀", f"{selected_result['P0']:.2e}")
                col4.metric("R²", f"{selected_result['R2']:.4f}")
                col5.metric("Copies", f"{_copies_val:.2e}")
            else:
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("D₀", f"{selected_result['D0']:.2e}")
                col2.metric("k", f"{selected_result['k']:.6f}")
                col3.metric("P₀", f"{selected_result['P0']:.2e}")
                col4.metric("R²", f"{selected_result['R2']:.4f}")

            _bl_meta_vis = selected_result.get('bl_end_meta')
            _bl_est_vis  = selected_result.get('bl_end_est')
            if _bl_meta_vis is not None or _bl_est_vis is not None:
                _bl_parts = []
                if _bl_meta_vis is not None:
                    _bl_parts.append(f"Instrument baseline end: cycle **{_bl_meta_vis:.0f}**")
                if _bl_est_vis is not None:
                    _bl_parts.append(f"Estimated baseline end: cycle **{_bl_est_vis:.0f}**")
                if _bl_meta_vis is not None and _bl_est_vis is not None:
                    _bl_diff = abs(_bl_meta_vis - _bl_est_vis)
                    _bl_icon = "✅" if _bl_diff <= 2 else ("⚠️" if _bl_diff <= 5 else "❌")
                    _bl_parts.append(f"{_bl_icon} Δ = {_bl_diff:.0f} cycles")
                st.caption(" · ".join(_bl_parts))

    except Exception as e:
        st.error(f"Could not visualize fit: {str(e)}")

# ── Visualize skipped (no-signal) wells ──────────────────────────────────
if ('batch_no_signal_samples' in st.session_state
        and st.session_state['batch_no_signal_samples']
        and 'batch_cycles' in st.session_state):
    _ns_samples = st.session_state['batch_no_signal_samples']
    _ns_cycles = st.session_state['batch_cycles']
    # Collect fluorescence data for skipped wells
    _ns_fluor = {}
    _ns_saved = st.session_state.get('batch_no_signal_fluor', {})
    _ns_upload = st.session_state.get('upload_batch_samples', {})
    _ns_results_list = st.session_state.get('batch_results_list', [])
    for _ns_name in _ns_samples:
        if _ns_name in _ns_saved:
            _ns_fluor[_ns_name] = np.asarray(_ns_saved[_ns_name])
        elif _ns_name in _ns_upload:
            _ns_fluor[_ns_name] = np.asarray(_ns_upload[_ns_name])
        else:
            # Check if batch_no_signal_samples itself holds fluor arrays (import case)
            _ns_val = _ns_samples[_ns_name]
            if isinstance(_ns_val, (np.ndarray, list)) and not isinstance(_ns_val, dict):
                _ns_fluor[_ns_name] = np.asarray(_ns_val)
            else:
                # Try results_list fluor_data
                for _rl in _ns_results_list:
                    if _rl.get('Sample') == _ns_name and _rl.get('fluor_data') is not None:
                        _ns_fluor[_ns_name] = np.asarray(_rl['fluor_data'])
                        break
    if _ns_fluor:
        st.markdown("---")
        st.subheader("🚫 Visualize Skipped Wells (No Signal)")
        _ns_names = list(_ns_fluor.keys())
        _ns_selected = st.selectbox(
            "Select skipped well to visualize:",
            _ns_names,
            key="ns_viz_selector",
        )
        _ns_info = _ns_samples.get(_ns_selected, {})
        _ns_reason = _ns_info.get('reason', 'No signal detected') if isinstance(_ns_info, dict) else 'No signal detected'
        st.caption(f"Skip reason: {_ns_reason}")

        _ns_fd = _ns_fluor[_ns_selected]
        _ns_cyc = _ns_cycles[:len(_ns_fd)]

        _ns_fig = go.Figure()
        _ns_fig.add_trace(go.Scatter(
            x=_ns_cyc, y=_ns_fd,
            mode='markers+lines',
            name='Data',
            marker=dict(size=6, color='blue', opacity=0.6),
            line=dict(color='blue', width=1, dash='dot'),
        ))
        _ns_fig.update_layout(
            title=f"<span style='color:gray'>SKIPPED</span> — {_ns_selected}",
            xaxis_title="Cycle",
            yaxis_title="Fluorescence",
            height=400,
            showlegend=True,
        )
        st.plotly_chart(_ns_fig, use_container_width=True)

        # Show basic stats
        _ns_c1, _ns_c2, _ns_c3 = st.columns(3)
        _ns_c1.metric("Range", f"{float(np.max(_ns_fd) - np.min(_ns_fd)):.4f}")
        _ns_bl_sd = float(np.std(_ns_fd[:min(12, len(_ns_fd) // 4)])) if len(_ns_fd) >= 12 else float(np.std(_ns_fd))
        _ns_c2.metric("Baseline SD", f"{_ns_bl_sd:.4f}")
        _ns_range_ratio = float(np.max(_ns_fd) - np.min(_ns_fd)) / _ns_bl_sd if _ns_bl_sd > 0 else 0
        _ns_c3.metric("Range / Baseline SD", f"{_ns_range_ratio:.1f}×")


if 'fitted_params' in st.session_state:
    fitted_params = st.session_state['fitted_params']
    optimizer = st.session_state['optimizer']

    # Show debug output if available
    if 'fit_debug_output' in st.session_state and st.session_state['fit_debug_output']:
        with st.expander("🐛 Debug Output (Optimization Details)", expanded=True):
            st.code(st.session_state['fit_debug_output'], language='text')

    # Results tabs
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Fit Visualization", "📈 Parameters & Metrics", "🔬 Bootstrap CI", "💾 Export"])
        
    with tab1:
            # Predict fitted curve
            F_pred = optimizer.predict(cycles)

            # Cycle-by-cycle amplification efficiency
            _, D_eff, _ = optimizer.model.simulate_cycles(
                D0=fitted_params['D0'],
                k=fitted_params['k'],
                P0=fitted_params['P0'],
                n_cycles=len(cycles),
                F_bg_intercept=fitted_params['F_bg_intercept'],
                F_bg_slope=fitted_params['F_bg_slope']
            )
            eff_per_cycle = calculate_amplification_efficiency(D_eff)

            # Calculate truncation point for visualization
            from mak2_model import find_slope_threshold_cycle
            threshold_idx = find_slope_threshold_cycle(
                fluorescence,
                cycles_after_max=cycles_after_max
            )
            threshold_label = f"Max slope + {cycles_after_max} cycles"
            threshold_cycle_num = cycles[min(threshold_idx, len(cycles)-1)]

            # Calculate Ct for vertical line
            ct_threshold_single = None
            ct_baseline_mean_single = 0.0
            try:
                ct_res_single = optimizer.calculate_ct(method='threshold')
                ct_val_single = ct_res_single['ct']
                ct_threshold_single = ct_res_single.get('threshold')
                ct_baseline_mean_single = ct_res_single.get('baseline_mean', 0.0)
            except Exception:
                ct_val_single = np.nan

            # 3-row stacked plot: Fit, Residuals, Efficiency (shared x-axis)
            fig = make_subplots(
                rows=3, cols=1,
                subplot_titles=(
                    f"qPCR Curve Fit (Truncated at Max Slope + {cycles_after_max})",
                    "Residuals",
                    "Amplification Efficiency"
                ),
                vertical_spacing=0.08,
                row_heights=[0.50, 0.22, 0.28],
                shared_xaxes=True,
            )

            # Row 1: Data and fit
            fig.add_trace(
                go.Scatter(
                    x=cycles, y=fluorescence,
                    mode='markers',
                    name='Data',
                    marker=dict(size=8, color='blue', opacity=0.6)
                ),
                row=1, col=1
            )
            fig.add_trace(
                go.Scatter(
                    x=cycles, y=F_pred,
                    mode='lines',
                    name='MAK2 Fit',
                    line=dict(color='red', width=2)
                ),
                row=1, col=1
            )

            # Row 2: Residuals
            residuals = fluorescence - F_pred
            fig.add_trace(
                go.Scatter(
                    x=cycles, y=residuals,
                    mode='markers',
                    name='Residuals',
                    marker=dict(size=6, color='purple')
                ),
                row=2, col=1
            )
            fig.add_hline(y=0, line_dash="dash", line_color="gray", row=2, col=1)

            # Row 3: Efficiency
            fig.add_trace(
                go.Scatter(
                    x=np.arange(1, len(eff_per_cycle)+1),
                    y=eff_per_cycle,
                    mode='lines+markers',
                    name='Efficiency',
                    marker=dict(size=4),
                    line=dict(color='orange'),
                ),
                row=3, col=1
            )

            # Ct threshold horizontal line on fit plot (row 1)
            # The threshold is relative to baseline-subtracted data (delta_rn),
            # so plot it at baseline_mean + threshold on the raw fluorescence axis
            if ct_threshold_single is not None and ct_threshold_single > 0:
                threshold_plot_y = ct_threshold_single + ct_baseline_mean_single
                fig.add_hline(
                    y=threshold_plot_y,
                    line_dash="dot",
                    line_color="purple",
                    line_width=1,
                    row=1, col=1,
                )
                fig.add_annotation(
                    x=0, xref="x domain",
                    y=threshold_plot_y,
                    text=f"Threshold ({threshold_plot_y:.4f})",
                    showarrow=False,
                    xanchor="left", yanchor="bottom",
                    font=dict(size=10, color="purple"),
                    row=1, col=1,
                )

            # Ct vertical line on all rows
            if not np.isnan(ct_val_single):
                for row_idx in range(1, 4):
                    fig.add_vline(
                        x=ct_val_single,
                        line_dash="dot",
                        line_color="gray",
                        row=row_idx, col=1,
                    )
                    if row_idx == 1:
                        fig.add_annotation(
                            x=ct_val_single, y=1, yref="y domain",
                            text="Ct", showarrow=False,
                            xanchor="left", yanchor="top",
                            font=dict(size=11, color="gray"),
                            row=1, col=1,
                        )

            # Final fitted cycle vertical line on all rows
            if threshold_cycle_num < cycles[-1]:
                for row_idx in range(1, 4):
                    fig.add_vline(
                        x=threshold_cycle_num,
                        line_dash="dash",
                        line_color="green",
                        row=row_idx, col=1,
                    )
                    if row_idx == 1:
                        fig.add_annotation(
                            x=threshold_cycle_num, y=1, yref="y domain",
                            text="Final fitted cycle", showarrow=False,
                            xanchor="right", yanchor="top",
                            font=dict(size=11, color="green"),
                            row=1, col=1,
                        )

            fig.update_xaxes(title_text="Cycle", row=3, col=1)
            fig.update_yaxes(title_text="Fluorescence", row=1, col=1)
            fig.update_yaxes(title_text="Residual", row=2, col=1)
            fig.update_yaxes(title_text="Efficiency", row=3, col=1)
            fig.update_layout(height=800, showlegend=True)

            st.plotly_chart(fig, use_container_width=True)
                
            # Comprehensive goodness-of-fit metrics (calculated on fitted data only)
            st.subheader("Goodness of Fit Metrics")
            st.caption("⚠️ Metrics calculated only on the fitted region (after truncation), not the full dataset")
                
            metrics = optimizer.calculate_fit_metrics()
                
            # Display main metrics in columns
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("R²", f"{metrics['r_squared']:.6f}", 
                       help="Coefficient of determination. Can be negative for poor nonlinear fits!")
            col2.metric("RMSE", f"{metrics['rmse']:.4f}", 
                       help="Root Mean Squared Error (fluorescence units)")
            col3.metric("NRMSE", f"{metrics['nrmse']*100:.2f}%",
                       help="Normalized RMSE (% of signal range)")
            # Show quality indicator based on R²
            quality = "✅ Excellent" if metrics['r_squared'] >= 0.999 else "⚠️ Check fit"
            col4.metric("Fit Quality", quality,
                       help="Based on R² threshold (≥0.999 = excellent)")
                
            # Additional metrics in expander
            with st.expander("📊 Additional Fit Metrics"):
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("MAE", f"{metrics['mae']:.4f}", 
                             help="Mean Absolute Error")
                    st.metric("MAPE", f"{metrics['mape']:.2f}%",
                             help="Mean Absolute Percentage Error")
                    st.metric("SSR", f"{metrics['ssr']:.6f}",
                             help="Sum of Squared Residuals")
                with col2:
                    st.metric("AIC", f"{metrics['aic']:.2f}",
                             help="Akaike Information Criterion (lower is better)")
                    st.metric("BIC", f"{metrics['bic']:.2f}",
                             help="Bayesian Information Criterion (lower is better)")
                    st.metric("Reduced χ²", f"{metrics['reduced_chi_sq']:.4f}",
                             help="Chi-squared per degree of freedom")
                    
                st.caption(f"**Data points:** {metrics['n_points']} (after truncation) | "
                          f"**Parameters:** {metrics['n_params']} | "
                          f"**Degrees of freedom:** {metrics['dof']}")
            
    with tab2:
        st.subheader("Fitted Parameters")
                
        param_df = pd.DataFrame({
            'Parameter': ['D₀ (Initial DNA)', 'k (PCR constant)', 'P₀ (Initial primer)',
                         'F_bg (intercept)', 'F_bg_slope'],
            'Value': [
                f"{fitted_params['D0']:.2e}",  # D0 in fluorescence units - scientific notation
                f"{fitted_params['k']:.6f}",
                f"{fitted_params['P0']:.4e}",
                f"{fitted_params['F_bg_intercept']:.6f}",
                f"{fitted_params['F_bg_slope']:.6f}"
            ],
            'Description': [
                'Initial template fluorescence (fluorescence units)',
                'Ratio of primer binding to reannealing rate',
                'Initial primer concentration',
                'Background fluorescence (constant)',
                'Background fluorescence (linear slope)'
            ]
        })
                
        st.dataframe(param_df, use_container_width=True)
            
    with tab3:
        st.subheader("🔬 Bootstrap Confidence Intervals")
                
        st.markdown("""
        **Get professional-grade uncertainty estimates for your parameters!**
                
        Bootstrap analysis provides 95% confidence intervals by:
        1. Resampling residuals from your fit 1000 times
        2. Refitting the model to each bootstrap sample
        3. Calculating percentile-based confidence intervals
                
        ⏱️ **Analysis time:** 10-30 minutes  
        📊 **Output:** Confidence intervals + distribution plots
        """)
                
        # Initialize session state for bootstrap
        if 'bootstrap_running' not in st.session_state:
            st.session_state.bootstrap_running = False
        if 'bootstrap_results' not in st.session_state:
            st.session_state.bootstrap_results = None
        if 'bootstrap_sample_name' not in st.session_state:
            st.session_state.bootstrap_sample_name = None
                
        # Determine current sample name
        if batch_mode:
            current_sample_name = preview_sample
        else:
            current_sample_name = "single_sample"
                
        # Clear bootstrap results if sample changed
        if st.session_state.bootstrap_sample_name != current_sample_name:
            st.session_state.bootstrap_results = None
            st.session_state.bootstrap_sample_name = current_sample_name
                
        # Email input
        st.markdown("### Run Bootstrap Analysis")
        if batch_mode:
            st.info(f"📊 **Current sample:** {current_sample_name}\n\nBootstrap will run on this sample only. Change preview sample to bootstrap a different curve.")
        else:
            st.info("⚠️ Your browser will need to stay open during analysis (~10-30 minutes)")
                
        col1, col2 = st.columns([2, 1])
        with col1:
            n_bootstrap = st.number_input("Number of bootstrap iterations", 
                                         min_value=100, max_value=2000, value=1000, step=100,
                                         help="More iterations = better CI estimates but longer runtime",
                                         key="n_bootstrap_input")
        with col2:
            estimated_time = n_bootstrap / 1000 * 15  # ~15 min for 1000
            st.metric("Estimated Time", f"{estimated_time:.0f} min")
                
        # Show existing results first if they exist
        if st.session_state.bootstrap_results is not None:
            st.success("✅ Bootstrap analysis complete!")
            # Display results here (we'll move the display code up)
                
        # Run button (disabled if already running or have results)
        run_bootstrap = st.button(
            "▶️ Run Bootstrap Analysis", 
            type="primary", 
            use_container_width=True,
            disabled=st.session_state.bootstrap_running or st.session_state.bootstrap_results is not None,
            key="run_bootstrap_btn"
        )
                
        if run_bootstrap:
            # Create progress placeholder
            progress_bar = st.progress(0)
            status_text = st.empty()
                    
            try:
                status_text.text("Starting bootstrap analysis...")
                        
                # Run bootstrap with progress updates
                import time
                start_time = time.time()
                        
                # We'll run in smaller chunks to update progress
                status_text.text(f"Running {n_bootstrap} bootstrap iterations...")
                progress_bar.progress(0.1)
                        
                # Get the fitted region (what was actually used in the fit)
                cycles_fit = optimizer.cycles_fit
                fluorescence_fit = optimizer.fluorescence_fit
                        
                # Ensure arrays
                cycles_fit = np.asarray(cycles_fit)
                fluorescence_fit = np.asarray(fluorescence_fit)
                        
                # Get predicted values on fitted region (use all params including background)
                F_pred_fit = optimizer.predict(cycles_fit, fitted_params)
                        
                bootstrap_results = bootstrap_parameter_uncertainty(
                    cycles=cycles_fit,  # Use fitted cycles, not full data
                    fluorescence=fluorescence_fit,  # Use fitted fluorescence
                    original_params=fitted_params,  # Pass ALL params including background
                    original_fit=F_pred_fit,  # Prediction on fitted region
                    n_bootstrap=n_bootstrap,
                    confidence_level=0.95,
                    random_seed=None,
                    show_progress=False  # We'll use streamlit progress bar
                )
                        
                elapsed_time = time.time() - start_time
                        
                # Store results in session state
                st.session_state.bootstrap_results = bootstrap_results
                        
                progress_bar.progress(1.0)
                status_text.text(f"✅ Complete! ({elapsed_time/60:.1f} minutes)")
                        
                # Store success flag to prevent re-running bootstrap
                st.session_state.bootstrap_just_completed = True
                        
                st.success(f"🎉 Bootstrap complete! {bootstrap_results.n_successful}/{bootstrap_results.n_bootstrap} successful fits")
                st.rerun()  # Rerun to display results
                        
            except Exception as e:
                st.error(f"❌ Bootstrap failed: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
                
        # Display results if available
        if st.session_state.bootstrap_results is not None:
            results = st.session_state.bootstrap_results
                    
            st.markdown("---")
            st.markdown("## 📊 Bootstrap Results")
                    
            # Show diagnostics FIRST
            with st.expander("🔍 Bootstrap Diagnostics", expanded=True):
                D0_samples = results.D0_samples
                k_samples = results.k_samples
                P0_samples = results.P0_samples
                        
                diag_col1, diag_col2 = st.columns(2)
                with diag_col1:
                    st.write("**Sample Statistics:**")
                    st.write(f"- D₀: min={D0_samples.min():.2e}, max={D0_samples.max():.2e}, std={np.std(D0_samples):.2e}")
                    st.write(f"- k: min={k_samples.min():.4f}, max={k_samples.max():.4f}, std={np.std(k_samples):.4f}")
                    st.write(f"- P₀: min={P0_samples.min():.2e}, max={P0_samples.max():.2e}, std={np.std(P0_samples):.2e}")
                        
                with diag_col2:
                    st.write("**Variation Check:**")
                    st.write(f"- Unique D₀ values: {len(np.unique(D0_samples))}/{len(D0_samples)}")
                    st.write(f"- Unique k values: {len(np.unique(k_samples))}/{len(k_samples)}")
                    st.write(f"- Unique P₀ values: {len(np.unique(P0_samples))}/{len(P0_samples)}")
                            
                    if len(np.unique(D0_samples)) < 10:
                        st.warning("⚠️ Very few unique D₀ values - bootstrap may not be working correctly!")
                    if len(np.unique(k_samples)) < 10:
                        st.warning("⚠️ Very few unique k values - bootstrap may not be working correctly!")
                        
                st.write("**Original Fit (for comparison):**")
                st.write(f"- D₀={results.D0_point:.2e}, k={results.k_point:.4f}, P₀={results.P0_point:.2e}")
                    
            # Display confidence intervals in columns
            st.markdown("### Parameter Estimates with 95% Confidence Intervals")
            col1, col2, col3 = st.columns(3)
                    
            with col1:
                st.metric(
                    "D₀ (Initial DNA)",
                    f"{results.D0_point:.2e}",
                    help="Point estimate from original fit"
                )
                st.caption(f"**95% CI:** [{results.D0_ci[0]:.2e}, {results.D0_ci[1]:.2e}]")
                st.caption(f"**Std:** {np.std(results.D0_samples):.2e}")
                    
            with col2:
                st.metric(
                    "k (PCR constant)",
                    f"{results.k_point:.4f}",
                    help="Point estimate from original fit"
                )
                st.caption(f"**95% CI:** [{results.k_ci[0]:.4f}, {results.k_ci[1]:.4f}]")
                st.caption(f"**Std:** {np.std(results.k_samples):.4f}")
                    
            with col3:
                st.metric(
                    "P₀ (Initial primer)",
                    f"{results.P0_point:.2e}",
                    help="Point estimate from original fit"
                )
                st.caption(f"**95% CI:** [{results.P0_ci[0]:.2e}, {results.P0_ci[1]:.2e}]")
                st.caption(f"**Std:** {np.std(results.P0_samples):.2e}")
                    
            # Efficiency
            st.markdown("### Amplification Efficiency")
            col1, col2 = st.columns(2)
            with col1:
                st.metric(
                    "Efficiency (E)",
                    f"{results.efficiency_point:.4f}",
                    help="E = 1 + k*P₀"
                )
            with col2:
                st.caption(f"**95% CI:** [{results.efficiency_ci[0]:.4f}, {results.efficiency_ci[1]:.4f}]")
                    
            # Distribution plots
            st.markdown("### Parameter Distributions")
                    
            try:
                # Create figure using bootstrap analyzer
                analyzer = BootstrapAnalyzer(
                    model=MAK2Model(),
                    optimizer=MAK2Optimizer(MAK2Model())
                )
                fig = analyzer.plot_bootstrap_distributions(results, figsize=(15, 5))
                st.pyplot(fig)
            except Exception as e:
                st.warning(f"Could not create distribution plots: {e}")
                import traceback
                with st.expander("🔍 Error Details"):
                    st.code(traceback.format_exc())
                    
            # Summary table
            with st.expander("📊 Detailed Bootstrap Summary"):
                summary = results.summary_dict()
                        
                summary_df = pd.DataFrame({
                    'Parameter': ['D₀', 'k', 'P₀', 'Efficiency'],
                    'Estimate': [
                        f"{summary['D0']['estimate']:.2e}",
                        f"{summary['k']['estimate']:.4f}",
                        f"{summary['P0']['estimate']:.2e}",
                        f"{summary['efficiency']['estimate']:.4f}"
                    ],
                    'CI Lower': [
                        f"{summary['D0']['ci_lower']:.2e}",
                        f"{summary['k']['ci_lower']:.4f}",
                        f"{summary['P0']['ci_lower']:.2e}",
                        f"{summary['efficiency']['ci_lower']:.4f}"
                    ],
                    'CI Upper': [
                        f"{summary['D0']['ci_upper']:.2e}",
                        f"{summary['k']['ci_upper']:.4f}",
                        f"{summary['P0']['ci_upper']:.2e}",
                        f"{summary['efficiency']['ci_upper']:.4f}"
                    ],
                    'Std Dev': [
                        f"{summary['D0']['std']:.2e}",
                        f"{summary['k']['std']:.4f}",
                        f"{summary['P0']['std']:.2e}",
                        'N/A'
                    ]
                })
                        
                st.dataframe(summary_df, use_container_width=True)
                        
                st.caption(f"**Metadata:** {summary['metadata']['n_successful']}/{summary['metadata']['n_bootstrap']} "
                          f"successful fits ({summary['metadata']['success_rate']:.1%}) | "
                          f"Runtime: {summary['metadata']['runtime_seconds']/60:.1f} minutes | "
                          f"Confidence level: {summary['metadata']['confidence_level']:.0%}")
                    
            # Download bootstrap results
            st.markdown("### Download Bootstrap Results")
                    
            # Create CSV with bootstrap samples
            bootstrap_df = pd.DataFrame({
                'iteration': range(len(results.D0_samples)),
                'D0': results.D0_samples,
                'k': results.k_samples,
                'P0': results.P0_samples
            })
                    
            csv_bootstrap = bootstrap_df.to_csv(index=False)
            st.download_button(
                label="📥 Download Bootstrap Samples (CSV)",
                data=csv_bootstrap,
                file_name="bootstrap_samples.csv",
                mime="text/csv"
            )
                    
            # Clear button
            if st.button("🗑️ Clear Bootstrap Results", key="clear_bootstrap_btn"):
                st.session_state.bootstrap_results = None
                st.rerun()
            
    with tab4:
        st.subheader("Export Results")
                
        # Prepare export data
        export_df = pd.DataFrame({
            'Cycle': cycles,
            'Fluorescence_Data': fluorescence,
            'Fluorescence_Fit': F_pred,
            'Residual': residuals
        })
                
        # CSV download
        csv = export_df.to_csv(index=False)
        st.download_button(
            label="📥 Download Results (CSV)",
            data=csv,
            file_name="mak2_results.csv",
            mime="text/csv"
        )
                
        # Parameter summary
        st.subheader("Parameter Summary (for copying)")
        summary_text = f"""
MAK2 Fitting Results
====================
D₀ (Initial DNA): {fitted_params['D0']:.4e}
k (PCR constant): {fitted_params['k']:.6f}
P₀ (Initial primer): {fitted_params['P0']:.4e}
F_bg_intercept: {fitted_params['F_bg_intercept']:.6f}
F_bg_slope: {fitted_params['F_bg_slope']:.6f}
    
Goodness of Fit (on fitted region):
R² = {optimizer.calculate_fit_metrics()['r_squared']:.6f}
RMSE = {optimizer.calculate_fit_metrics()['rmse']:.4f}
NRMSE = {optimizer.calculate_fit_metrics()['nrmse']*100:.2f}%
"""
        st.text_area("Summary", summary_text, height=250)


if ('batch_results' not in st.session_state
        and 'fitted_params' not in st.session_state
        and (cycles is None or fluorescence is None)):
    st.info("👈 Please load data using the sidebar")
    
    # Show instructions
    st.markdown("""
    ### Getting Started
    
    1. **Load your data** using one of three methods:
       - Use provided example datasets
       - Upload a CSV or Excel file (first column: cycles, second: fluorescence)
       - Enter data manually
    
    2. **Configure fitting options** in the sidebar
    
    3. **Click "Fit Model"** to run the optimization
    
    4. **Explore results** in the tabs:
       - Visualize the fit quality
       - Examine fitted parameters
       - Export results
    
    ### About the MAK2 Model
    
    The MAK2 model is a mechanistic model of PCR that accounts for:
    - Competition between primer binding and DNA reannealing
    - Primer depletion over cycles
    - Background fluorescence (with linear drift)
    
    **Reference:** Boggy & Woolf (2010). A Mechanistic Model of PCR for Accurate Quantification 
    of Quantitative PCR Data. PLOS ONE 5(8): e12355.
    
    ### Example Datasets

    All three example datasets are from the [qpcR](https://cran.r-project.org/package=qpcR) R package:

    **Boggy** - 10-fold dilution series (Boggy & Woolf, 2010, *PLOS ONE*)
    **Rutledge** - High-throughput screen with 120 wells (Rutledge, 2004, *Nucleic Acids Research*)
    **Technical Replicates** - `reps` dataset: 7 dilutions with quad replicates (Spiess & Mueller, IHF Hamburg)

    [View dataset documentation](https://github.com/gboggy2/MAK2-plus/tree/main/example_data)
    """)

# Footer
st.sidebar.markdown("---")
st.sidebar.markdown("Built with Streamlit")

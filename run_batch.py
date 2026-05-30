#!/Users/boggy/anaconda3/envs/qpcr_mak2/bin/python
"""Offline batch-fitting driver: process every plate in a folder, write Excel.

Why this script exists alongside ``app.py``:

The Streamlit app does the same per-well fitting work, but it does
it interactively, one plate at a time, in a process that needs to
stay attached to the user's browser. This script is the headless
batch equivalent — point it at a folder of plate files and it
processes each one to completion overnight, producing
``Results/<PlateX>_MAK2Plus_Results.xlsx`` that can later be
re-loaded in the app for visualization via "Load Previous Results".

The per-well fitting logic was duplicated here from ``app.py`` (see
the Phase-1 unification task in CLAUDE.md). Keeping them in sync is
manual; the long-term plan is to extract a shared ``fit_well()``
function during the Next.js + FastAPI port.

Pipeline overview (orchestrated by ``process_plate`` per plate):

  1. ``load_plate_data`` — parse multicomponent + metadata CSVs.
  2. ``detect_no_signal_samples`` — triage NTCs and failed wells.
  3. ``run_pass1`` — fit every survivor with the MAK2 optimizer.
  4. ``run_pass2`` — channel-aware retry pass for borderline fits
     (using channel-median k/P0 as priors).
  5. ``run_quality_gates`` — mark final PASS / FAIL / INDETERMINATE
     status per well.
  6. ``build_replicate_groups`` — replicate stats from the metadata.
  7. ``build_standard_curve`` / ``build_ct_standard_curve`` — fit
     calibration curves from STANDARD wells.
  8. ``apply_calibration`` / ``apply_ct_calibration`` — D0→Copies.
  9. ``build_excel`` — write the multi-sheet result file.

Usage:
    caffeinate -s python run_batch.py

Output:
    Results/<PlateX>_MAK2Plus_Results.xlsx  (one per plate)
"""

import sys
import time
import io
import traceback
import numpy as np
import pandas as pd
from pathlib import Path

# ── Ensure the MAK2+ modules are importable ���─────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

from mak2_model import MAK2Model, estimate_MAK2_params_from_exponential
from optimizer import MAK2Optimizer
# ``prepare_fit_inputs`` is imported lazily inside ``run_pass1`` to avoid
# a circular dependency: ``fit_well`` imports ``smart_start`` /
# ``adaptive_window_extension`` from this module. A follow-up commit
# should move those helpers into ``fit_well`` (or a shared module) so
# the lazy import can become a top-level one.
from data_processing import detect_no_signal_samples, estimate_baseline_end
from qpcr_data_converter import QPCRDataConverter, load_abi_results_csv
from replicate_analysis import calculate_replicate_stats, compare_precision
from calibration import build_standard_curve, build_ct_standard_curve, apply_calibration, apply_ct_calibration
from config import DEFAULT_GATES

# ── Configuration ───────��────────────────────────────��────────────────────────
DATA_DIR = Path("/Users/boggy/Desktop/Desktop031424/Personal/InputDataFiles")
OUTPUT_DIR = DATA_DIR / "Results"
# NOTE: OUTPUT_DIR is created lazily in the ``__main__`` block below,
# not at module-import time. Importing this module (e.g. from tests
# or from the Phase 1 FastAPI layer) must not have filesystem side
# effects against a developer-specific path.

# Plates to process: (multicomponent_csv, metadata_csv, output_name)
PLATES = []
for letter in "ABCDEFGHI":
    mc_file = DATA_DIR / f"Plate_{letter}_Multicomponent.csv"
    meta_file = DATA_DIR / f"Plate_{letter}_data.csv"
    if mc_file.exists() and meta_file.exists():
        PLATES.append((mc_file, meta_file, f"Plate{letter}"))

# Default settings (matching app sidebar defaults)
FIRST_FIT_CYCLE = 3
# CYCLES_BEFORE_MAX: cycles of baseline retained before the inflection (1st
# derivative max) in the fit window. Increased from 10 to 15 after PCRedux
# false-FAIL inspection showed late amplifiers (inflection at cycle ~38)
# were getting only 2-3 baseline cycles in the fit window, which starved
# Gate 2b's MAK2-vs-linear discriminator: with too little baseline, a pure
# linear fit and a MAK2 sigmoid fit score equally well on the short
# pre-inflection window. Adding ~5 more baseline cycles widens the
# discrimination gap and recovers the borderline late amplifiers.
CYCLES_BEFORE_MAX = 15
CYCLES_AFTER_MAX = 4
AUTO_TRUNCATE = True
TRUNCATE_CYCLE = None
CUSTOM_BOUNDS_DICT = None

# Replicate grouping: "sample_name" uses Sample Name from metadata
REPLICATE_GROUPING = "sample_name"


# ── Helper functions ───────────���──────────────────────────────────────────────

# Sample-name parsing helpers live in pass2_helpers (canonical implementation
# shared with app.py). Re-exported under the legacy private names so that the
# many call sites in this module don't need to be renamed.
from pass2_helpers import channel_of as _ch, well_pos_of as _get_well_pos  # noqa: E402

def pre_estimate_background(cycles, fluor, bl_start_idx, bl_end_idx):
    """Linear regression over a baseline window — local copy of the canonical helper.

    This is functionally a duplicate of
    ``mak2_model.pre_estimate_background``. The local copy was added
    before the modules were aligned; keeping both paths in sync is
    a Phase-1 task. See CLAUDE.md.
    """
    bg_c = cycles[bl_start_idx:bl_end_idx + 1]
    bg_f = fluor[bl_start_idx:bl_end_idx + 1]
    if len(bg_c) >= 2:
        coeffs = np.polyfit(bg_c, bg_f, 1)
        return float(coeffs[0]), float(coeffs[1])
    return 0.0, float(fluor[bl_start_idx]) if bl_start_idx < len(fluor) else 0.0


def smart_start(fluor_data, cycles, floor_idx, cycles_before_max):
    """Locate the qPCR sigmoid inflection by scanning the smoothed gradient right-to-left.

    Why right-to-left: the inflection (max-slope cycle) of a real
    qPCR sigmoid is a *unique* peak in the smoothed gradient, but
    early-cycle baseline noise can produce spurious local maxima
    that an argmax over the full array would prefer. Scanning from
    the right finds the rightmost sustained gradient peak — the
    real inflection — and stops there.

    The peak is recognised when the running max has dropped to <50%
    of its peak value (and the peak itself exceeds 5% of the
    signal's gradient range, filtering out flat noise). If no peak
    is found, falls back to argmax over the whole gradient.

    The fit-window start is then set to ``cycles_before_max`` cycles
    upstream of the inflection (or the floor, whichever is larger).

    Also duplicated in ``app.py`` per the unification task in
    CLAUDE.md.

    Returns:
        ``(fit_start_idx, max_slope_idx)`` 0-based indices into
        ``cycles`` / ``fluor_data``.
    """
    full_seg = fluor_data[floor_idx:]
    if len(full_seg) >= 5:
        raw_grad = np.gradient(full_seg)
        kern = np.ones(5) / 5.0
        smooth_g = np.convolve(raw_grad, kern, mode='same')
        grad_range = float(np.max(smooth_g) - np.min(smooth_g))
        grad_floor = grad_range * 0.05
        best_val = smooth_g[-1]
        best_idx = len(smooth_g) - 1
        found_peak = False
        for j in range(len(smooth_g) - 2, -1, -1):
            if smooth_g[j] > best_val:
                best_val = smooth_g[j]
                best_idx = j
            elif best_val > grad_floor and smooth_g[j] < best_val * 0.5:
                found_peak = True
                break
        max_slope_offset = best_idx if found_peak else int(np.argmax(smooth_g))
    elif len(full_seg) >= 2:
        max_slope_offset = int(np.argmax(np.gradient(full_seg)))
    else:
        max_slope_offset = 0
    max_slope_idx = floor_idx + max_slope_offset
    fit_start_idx = max(floor_idx, max_slope_idx - cycles_before_max)
    return fit_start_idx, max_slope_idx


def adaptive_window_extension(fluor_data, cycles, fit_start_idx, max_slope_idx,
                              floor_idx, cycles_before_max, bg_slope, bg_int):
    """Walk the fit window's start cycle back until ≥3 baseline cycles are inside it.

    Why this matters: the optimizer needs *some* baseline cycles
    inside the fit window to constrain the background parameters
    against — if the window starts mid-amplification, the
    optimizer can't distinguish baseline drift from real signal and
    background ends up absorbing what should be D0 information.
    The MAK2 model fits cleanly when at least 3 of the early
    in-window cycles are below the projected background (within
    3σ noise).

    The procedure: extend ``fit_start_idx`` backward by 2 cycles at
    a time, recount baseline-region cycles, stop when ≥3 are found
    or the floor is reached. After the extension, refit the local
    background regression on the wider window so the bounds carried
    forward reflect the new starting point.

    A short-circuit: if the data immediately before
    ``fit_start_idx`` is essentially zero (background-subtracted
    plates), there's nothing to detect and the count is set to 3
    so no extension happens.
    """
    bg_pre_start = max(floor_idx, fit_start_idx - 8)
    bg_f = fluor_data[bg_pre_start:fit_start_idx]
    bl_noise = float(np.std(bg_f)) if len(bg_f) >= 2 else 0.0
    bg_mean_abs = float(np.mean(np.abs(bg_f))) if len(bg_f) >= 2 else 0.0
    skip_ext = (bl_noise < 1e-6 and bg_mean_abs < 1e-6)

    n_baseline = 0
    if not skip_ext:
        for bi in range(fit_start_idx, min(fit_start_idx + 6, len(cycles))):
            bg_level = bg_slope * cycles[bi] + bg_int
            if fluor_data[bi] <= bg_level + 3.0 * bl_noise:
                n_baseline += 1
    else:
        n_baseline = 3

    if n_baseline < 3:
        try_before = cycles_before_max
        while n_baseline < 3 and try_before < max_slope_idx - floor_idx:
            try_before += 2
            try_start = max(floor_idx, max_slope_idx - try_before)
            if try_start == fit_start_idx:
                break
            n_baseline = 0
            for bi in range(try_start, min(try_start + 6, len(cycles))):
                bg_level = bg_slope * cycles[bi] + bg_int
                if fluor_data[bi] <= bg_level + 3.0 * bl_noise:
                    n_baseline += 1
            fit_start_idx = try_start

        # Recompute background with wider window
        bg_pre_start = max(floor_idx, fit_start_idx - 8)
        bg_c = cycles[bg_pre_start:fit_start_idx]
        bg_f = fluor_data[bg_pre_start:fit_start_idx]
        if len(bg_c) >= 2:
            coeffs = np.polyfit(bg_c, bg_f, 1)
            bg_slope = float(coeffs[0])
            bg_int = float(coeffs[1])

    return fit_start_idx, bg_slope, bg_int


def compute_ct(optimizer_obj, cycles, fluor_data, sample_name, sample_metadata,
               rox_by_well, channel_thresholds, global_threshold,
               channel_baseline_means, global_baseline_mean):
    """Compute the Ct value for one well via the MAK2 optimizer's threshold method.

    Wraps ``MAK2Optimizer.calculate_ct`` with the right per-well
    inputs:

    - **Threshold**: per-channel from ``channel_thresholds`` (or the
      global fallback) — matches the instrument's Ct so the MAK2
      Ct is comparable.
    - **Baseline cycles**: extracted from the well's metadata
      (``Baseline Start`` / ``Baseline End``) so the baseline
      subtraction matches what the instrument did.
    - **ROX normalisation**: when the well has a ROX trace,
      fluorescence is divided by ROX before threshold-crossing
      detection; this corrects per-well volume variation.
    - **Instrument-Undetermined respect**: if the metadata says the
      instrument couldn't determine a Ct (NaN ``Ct_instrument``),
      we return NaN too rather than reporting a fitted Ct that the
      user might over-trust.

    The ``optimizer_obj.cycles_fit`` / ``fluorescence_fit`` fields
    are temporarily swapped to the full curve here (the optimizer
    normally holds the truncated fit window) so the Ct calculation
    sees the entire trace; they're restored before return.

    Returns:
        ``(ct_value, ct_baseline_val, ct_bl_slope, ct_bl_intercept,
        ct_rox_mean)``. Any failure path returns NaN ``ct_value``
        and zeroed baseline params rather than raising — failed Ct
        is a per-well outcome, not a pipeline error.
    """
    ct_value = np.nan
    ct_baseline_val = 0.0
    ct_bl_slope = 0.0
    ct_bl_intercept = 0.0
    ct_rox_mean = None

    try:
        ch = _ch(sample_name)
        sample_ch_thresh = channel_thresholds.get(ch, global_threshold)
        baseline_cycles_param = None

        if sample_metadata:
            wm_ct = sample_metadata.get(sample_name, {})
            bl_start = wm_ct.get('Baseline Start')
            bl_end = wm_ct.get('Baseline End')
            try:
                bl_start_i = int(np.searchsorted(cycles, float(bl_start)))
                bl_end_i = int(np.searchsorted(cycles, float(bl_end)))
                if bl_end_i > bl_start_i + 1:
                    baseline_cycles_param = (bl_start_i, bl_end_i)
            except (TypeError, ValueError):
                pass

        well_pos = _get_well_pos(sample_name)
        rox_arr = rox_by_well.get(well_pos, None)
        if rox_arr is not None and len(rox_arr) == len(fluor_data):
            fluor_for_ct = fluor_data / np.maximum(rox_arr, 1e-10)
            ct_rox_mean = float(np.mean(rox_arr))
            ct_threshold = sample_ch_thresh
        else:
            fluor_for_ct = fluor_data
            ct_rox_mean = None
            ct_threshold = None

        orig_cycles = optimizer_obj.cycles_fit
        orig_fluor = optimizer_obj.fluorescence_fit
        optimizer_obj.cycles_fit = cycles
        optimizer_obj.fluorescence_fit = fluor_for_ct
        ct_results = optimizer_obj.calculate_ct(
            method='threshold',
            threshold=ct_threshold,
            baseline_cycles=baseline_cycles_param,
        )
        optimizer_obj.cycles_fit = orig_cycles
        optimizer_obj.fluorescence_fit = orig_fluor
        ct_value = ct_results['ct']
        ct_baseline_val = ct_results.get('baseline_mean', 0.0)
        ct_bl_slope = ct_results.get('baseline_slope', 0.0)
        ct_bl_intercept = ct_results.get('baseline_intercept', 0.0)

        # Respect instrument Undetermined
        if sample_metadata:
            wm = sample_metadata.get(sample_name, {})
            if 'Ct_instrument' in wm:
                inst_ct = wm['Ct_instrument']
                inst_undetermined = (
                    inst_ct is None
                    or (isinstance(inst_ct, float) and np.isnan(inst_ct))
                )
                if inst_undetermined:
                    ct_value = np.nan
    except Exception:
        ct_value = np.nan

    return ct_value, ct_baseline_val, ct_bl_slope, ct_bl_intercept, ct_rox_mean


def load_plate_data(mc_file, meta_file):
    """Load and assemble both halves of an ABI plate's data files.

    A QuantStudio plate splits its export into two CSVs that have to
    be loaded together:

    - **Multicomponent CSV** (``mc_file``): per-channel fluorescence
      and the ROX passive reference for normalisation.
    - **Metadata CSV** (``meta_file``): per-well sample names, task
      types (UNKNOWN / STANDARD / NTC), known quantities for
      standards, baseline-window cycles, and instrument-reported
      Ct values.

    This function loads both, fans out the multicomponent data into
    a flat ``{f"{channel}_{well}": fluor_array}`` dict (the format
    the rest of the pipeline expects), and pulls the ROX trace into
    a separate ``{well: rox_array}`` dict.

    Args:
        mc_file: Path to the Multicomponent CSV.
        meta_file: Path to the Sample Setup / Results CSV.

    Returns:
        Six-tuple:
            ``(cycles, all_samples, channels, rox_by_well,
              sample_metadata, abi_results_meta)``.
    """
    converter = QPCRDataConverter()

    # Load multicomponent data
    cycles, samples, metadata = converter.load_from_file(str(mc_file))
    extra_info = metadata.get('extra_info', {})

    channels = extra_info.get('channels', [])
    passive_ref = extra_info.get('passive_reference', None)
    samples_by_channel = extra_info.get('samples_by_channel', {})

    # Extract fluorescence per channel (same as app.py channel loading)
    all_samples = {}
    rox_by_well = {}

    for channel in channels:
        c_cycles, c_samples, _ = converter.filter_by_channel(
            extra_info, channel, normalize_by_reference=False
        )
        for well_pos, fluor in c_samples.items():
            key = f"{channel}_{well_pos}"
            all_samples[key] = fluor

    # Extract ROX data
    if passive_ref and passive_ref in samples_by_channel:
        for well_pos, rox_fluor in samples_by_channel[passive_ref].items():
            rox_by_well[well_pos] = np.asarray(rox_fluor)

    # Load metadata
    abi_results_meta = load_abi_results_csv(str(meta_file))
    sample_metadata = abi_results_meta.get('sample_metadata', {})

    return cycles, all_samples, channels, rox_by_well, sample_metadata, abi_results_meta


def run_pass1(all_samples_to_fit, cycles, sample_metadata, rox_by_well,
              channel_thresholds, global_threshold, channel_baseline_means,
              global_baseline_mean):
    """First-pass MAK2 fitting over every survivor of the no-signal triage.

    Each well is fit by ``fit_well.fit_well`` — the canonical per-well
    entry point shared with the Streamlit app and the PCRedux scoring
    driver. Do not inline preprocessing or optimizer logic here;
    changes belong in ``fit_well``. See ARCHITECTURE.md § "The
    no-divergence rule".

    For each well, fit_well handles:

      1. **No-amplification pre-check.** A second, per-well-local
         test (the plate-wide one ran in
         ``detect_no_signal_samples``); compares signal range to
         baseline SD with a 5σ threshold and a tail-vs-baseline
         comparison. Failures are recorded with ``Tier=None`` and
         ``error='No amplification detected'`` and skipped without
         calling the optimizer.
      2. **Background pre-estimation** from the metadata baseline
         window when available (``Baseline Start`` / ``Baseline End``).
      3. **Smart-start inflection search** (``smart_start``) +
         **adaptive window extension** (``adaptive_window_extension``)
         to determine the fit window's start cycle.
      4. **Safety-net check**: if the truncated fit window's
         fluorescence range is < 70% of the full trace range, the
         smart-start probably missed the sigmoid (e.g. very late
         amplifier) — fall back to the floor cycle and recompute
         background.
      5. **Build per-well bounds** from the local background
         regression (slope and intercept ± data-driven margins).
      6. **Call the MAK2 optimizer** with fixed-background mode
         seeded by the per-well background estimate.
      7. **Compute Ct** via ``compute_ct``.
      8. **Classify the tier** (T1-Full / T1-Fixed / T2-LHS / T3-DE)
         based on which optimizer escalation tier produced the fit.

    Per-well failures are caught and recorded with
    ``Success='✗ Error: ...'`` rather than propagating, so one bad
    well doesn't kill the batch.

    The function preserves the per-well code path of ``app.py``'s
    batch mode line-for-line (see CLAUDE.md Phase-1 unification).

    Args:
        all_samples_to_fit: ``{sample_name: fluorescence_array}``
            for every survivor of the triage.
        cycles: Cycle-number array shared across all wells.
        sample_metadata: Per-well metadata dict from the Results
            CSV (keyed by composite ``"{channel}_{well}"`` name).
        rox_by_well: ``{well_pos: rox_array}`` for ROX-aware Ct.
        channel_thresholds: Per-channel Ct thresholds from the
            instrument metadata.
        global_threshold: Plate-wide fallback threshold.
        channel_baseline_means: Per-channel baseline-fluorescence
            mean (currently unused in fitting; passed for
            future per-channel features).
        global_baseline_mean: Plate-wide fallback baseline mean.

    Returns:
        List of per-well result dicts (every well in
        ``all_samples_to_fit`` produces exactly one entry, including
        failures).
    """
    from fit_well import fit_well  # local import — fit_well imports back here

    results_list = []
    total = len(all_samples_to_fit)
    fit_bounds_template = dict(CUSTOM_BOUNDS_DICT) if CUSTOM_BOUNDS_DICT else None

    for i, (sample_name, fluor_data) in enumerate(all_samples_to_fit.items()):
        print(f"  Pass 1: [{i+1}/{total}] {sample_name}", end="")
        t0 = time.perf_counter()

        # Pull per-well postprocessing inputs and let fit_well handle the
        # full pipeline (no-amp check, preprocessing, optimizer, toe stages,
        # Ct, tier, instrument status). This is the canonical per-well call.
        wm = sample_metadata.get(sample_name, {}) if sample_metadata else None
        ch = _ch(sample_name)
        ch_thresh = channel_thresholds.get(ch, global_threshold)
        well_pos = _get_well_pos(sample_name)
        rox = rox_by_well.get(well_pos) if rox_by_well else None

        result = fit_well(
            cycles, fluor_data,
            first_fit_cycle=FIRST_FIT_CYCLE,
            cycles_before_max=CYCLES_BEFORE_MAX,
            cycles_after_max=CYCLES_AFTER_MAX,
            auto_truncate=AUTO_TRUNCATE,
            truncate_cycle=TRUNCATE_CYCLE,
            fit_bounds=fit_bounds_template,
            sample_name=sample_name,
            metadata=wm,
            rox=rox,
            channel_threshold=ch_thresh,
        )
        results_list.append(result)

        elapsed = time.perf_counter() - t0
        if result.get('error') == 'No amplification detected':
            print(f"  → no amplification ({elapsed:.1f}s)")
        elif result.get('Success') == '✓':
            print(f"  → R²={result['R2']:.4f} k={result['k']:.4f} ({elapsed:.1f}s)")
        else:
            print(f"  → ERROR: {result.get('error', 'unknown')}")

    return results_list


def run_pass2(results_list, cycles, sample_metadata, rox_by_well,
              channel_thresholds, global_threshold, channel_baseline_means,
              global_baseline_mean):
    """Channel-aware retry pass for borderline-quality fits.

    Each per-well retry runs ``pass2_helpers.retry_one_well`` — the
    canonical retry implementation shared with the Streamlit app's
    batch loop. Do not inline retry variants here; new retry strategies
    belong in ``retry_one_well`` so both drivers pick them up. See
    ARCHITECTURE.md § "The no-divergence rule".

    Why a second pass exists: Pass 1 fits each well in isolation
    against generic data-driven bounds. After Pass 1 finishes, we
    know the channel-typical k and P0 distributions across the
    whole plate. Wells with poor R² (or otherwise suspect fits) are
    re-fit using the channel medians as priors — this rescues wells
    where Pass 1 landed in a bad local minimum without distorting
    fits that were already good.

    The retry replicates Pass 1's preprocessing (smart-start,
    adaptive window, background re-estimation) but uses tighter
    bounds anchored at the channel-typical values. Whichever fit
    has lower SSR (Pass 1 vs Pass 2) wins.

    Like ``run_pass1``, mirrors ``app.py`` line-for-line — see
    CLAUDE.md Phase-1 unification for the eventual extraction.

    Args:
        results_list: Output of ``run_pass1``. Mutated in place
            (the entries are replaced with retry results when the
            retry wins).
        cycles, sample_metadata, rox_by_well, channel_thresholds,
            global_threshold, channel_baseline_means, global_baseline_mean:
            Same meaning as in ``run_pass1``.

    Returns:
        The (mutated) ``results_list``.
    """
    from pass2_helpers import (
        compute_channel_priors, identify_retry_candidates, retry_one_well,
    )
    last_cyc = float(cycles[-1])  # noqa: F841 — kept for readability of summary line

    # Channel-prior stats + retry-candidate identification — shared
    # canonical implementation used by both this driver and app.py.
    channel_medians, plate_medians = compute_channel_priors(results_list)
    retry_indices, _skipped = identify_retry_candidates(
        results_list, cycles, CYCLES_AFTER_MAX,
    )
    if not retry_indices:
        print("  Pass 2: No samples need retry")
        return results_list

    print(f"  Pass 2: Retrying {len(retry_indices)} samples "
          f"({len(channel_medians)} channel(s) learned)")

    for idx in retry_indices:
        result = results_list[idx]
        sample_name = result.get('Sample', '?')
        retry_t0 = time.perf_counter()

        ch = _ch(sample_name)
        priors = channel_medians.get(ch, plate_medians)
        wm = sample_metadata.get(sample_name, {}) if sample_metadata else None
        well_pos = _get_well_pos(sample_name)
        rox = rox_by_well.get(well_pos) if rox_by_well else None
        ch_thresh = (channel_thresholds.get(ch, global_threshold)
                     if channel_thresholds else global_threshold)

        print(f"    [{idx}] {sample_name} (R²={result.get('R2', 'None')})", end="")
        new_result = retry_one_well(
            result, priors, cycles,
            metadata=wm, rox=rox, channel_threshold=ch_thresh,
            first_fit_cycle=FIRST_FIT_CYCLE,
            cycles_before_max=CYCLES_BEFORE_MAX,
            cycles_after_max=CYCLES_AFTER_MAX,
            auto_truncate=AUTO_TRUNCATE,
            truncate_cycle=TRUNCATE_CYCLE,
        )
        results_list[idx] = new_result
        elapsed = time.perf_counter() - retry_t0
        new_r2 = new_result.get('R2')
        if new_r2 is None:
            print(f"  → retry skipped ({elapsed:.1f}s)")
        elif new_r2 != result.get('R2'):
            print(f"  → R²={new_r2:.4f} ({elapsed:.1f}s)")
        else:
            print(f"  → kept original ({elapsed:.1f}s)")

    return results_list


def run_quality_gates(results_list, cycles, gates=None):
    """Apply the per-well quality gates and assign final ``Status``.

    After Pass 1+2 leave each well with its best available fit, this
    pass evaluates a series of quality gates and stamps each well
    with one of:

    - ``'PASS'``: cleared every applicable gate.
    - ``'FAIL'``: hit a hard rejection criterion (no amplification,
      a tier escalation that didn't recover, an obviously bad fit
      in some specific way).
    - ``'INDETERMINATE'``: the fit didn't pass cleanly but didn't
      fail outright — the result is reported but with reduced
      confidence and the failing gate reason in the Status.

    Gates evaluated (full list lives in the optimizer; this function
    just drives them and writes the verdict):

      - **Gate 0** (R² floor): R² ≥ 0.999 (relaxed for late
        amplifiers).
      - **Gate 2** (fit-window width): at least 12 cycles in the
        fitted window.
      - **Gate 2b** (sigmoid vs linear): MAK2 R² must be appreciably
        better than a linear fit on the same window — protects
        against the "any monotone curve fits both models" trap.
      - **Gate 3** (sigmoid shape): the second derivative of the
        fitted curve must change sign in the window — confirms the
        S-shape.
      - **Late-amplifier relaxation**: when the fit window's last
        cycle is the last data cycle, gate thresholds relax.

    Args:
        results_list: Output of ``run_pass2``. Mutated in place.
        cycles: Cycle-number array shared across wells.
        gates: Optional ``QualityGateConfig`` instance. Defaults to
            ``config.DEFAULT_GATES`` (the production thresholds).
            Tuning experiments pass alternative configs to score
            different gate parameterisations against PCRedux labels;
            tests can use a custom config to exercise specific gate
            edge cases.

    Returns:
        The mutated ``results_list``, with ``Status`` and
        ``status_detail`` populated.
    """
    # Thresholds + late-bypasses come from a single config so the tuning
    # driver can sweep over them. ``gates=None`` means "use production
    # defaults" — the active QualityGateConfig in config.DEFAULT_GATES.
    if gates is None:
        gates = DEFAULT_GATES

    for pf_idx, pf_r in enumerate(results_list):
        if pf_r.get('error') is not None:
            continue
        pf_r2 = pf_r.get('R2')
        pf_reject = False
        pf_reason = ''

        # Gate 5: Amplification amplitude — the model's growth component
        # (MAK2 prediction minus the fitted background line) must
        # account for a meaningful fraction of the observed fluorescence
        # range. Catches fits like media-1 maro2.69: a non-amplifying
        # curve with continuous gradual rise interpreted as a tiny
        # amplification (D0=4e-3, P0=5e-2) on top of bg_slope=+0.01.
        # The MAK2 fit explains the data via the bg line, not via
        # amplification — the growth component is only ~14% of F_range.
        # Real amplifiers have growth ≥ ~50% of F_range.
        pf_fluor_g5 = pf_r.get('fluor_data')
        pf_d0_g5 = pf_r.get('D0')
        if (not pf_reject and pf_fluor_g5 is not None
                and pf_d0_g5 is not None
                and not (isinstance(pf_d0_g5, float) and np.isnan(pf_d0_g5))
                and len(pf_fluor_g5) >= 10):
            pf_c_g5 = cycles[:len(pf_fluor_g5)]
            try:
                pf_model_g5 = MAK2Model().simulate_to_cycle(
                    D0=pf_r['D0'], k=pf_r['k'], P0=pf_r['P0'],
                    cycles=pf_c_g5,
                    F_bg_intercept=pf_r['F_bg_intercept'],
                    F_bg_slope=pf_r['F_bg_slope'],
                )
                pf_bg_line_g5 = (
                    pf_r['F_bg_intercept'] + pf_r['F_bg_slope'] * pf_c_g5
                )
                pf_growth_amp = float(np.max(pf_model_g5 - pf_bg_line_g5))
                pf_range_g5 = float(np.max(pf_fluor_g5) - np.min(pf_fluor_g5))
                if pf_range_g5 > 0:
                    pf_amp_pct = pf_growth_amp / pf_range_g5
                    if pf_amp_pct < 0.30:
                        pf_reject = True
                        pf_reason = (
                            f'Amplification amplitude too small '
                            f'({pf_amp_pct:.1%} of F_range, '
                            f'min 30%) — fit dominated by '
                            f'background line'
                        )
            except Exception:
                pass

        # Gate 4: Direction — raw fluorescence must end above its start.
        # Computed on the median of the first/last N cycles for noise
        # robustness; immune to model-parameter laundering (e.g., a real
        # downward trend absorbed into F_bg_slope to produce a high-R²
        # spurious fit on a non-amplifying well).
        pf_fluor_g4 = pf_r.get('fluor_data')
        if pf_fluor_g4 is not None and len(pf_fluor_g4) >= 2 * gates.direction_anchor_window:
            pf_n_g4 = gates.direction_anchor_window
            pf_first = float(np.median(pf_fluor_g4[:pf_n_g4]))
            pf_last  = float(np.median(pf_fluor_g4[-pf_n_g4:]))
            pf_range_g4 = float(np.max(pf_fluor_g4) - np.min(pf_fluor_g4))
            pf_growth = pf_last - pf_first
            if pf_range_g4 > 0:
                pf_growth_pct = pf_growth / pf_range_g4
                if pf_growth_pct < gates.min_growth_pct_of_range:
                    pf_reject = True
                    pf_reason = (
                        f'Curve does not rise '
                        f'(growth = {pf_growth_pct:+.1%} of range, '
                        f'min {gates.min_growth_pct_of_range:.1%})'
                    )

        # Late amplifier detection
        pf_fe_g0 = pf_r.get('fit_end_cycle')
        pf_is_late = (
            pf_fe_g0 is not None
            and not (isinstance(pf_fe_g0, float) and np.isnan(pf_fe_g0))
            and pf_fe_g0 >= float(cycles[-1]) - min(
                max(1, CYCLES_AFTER_MAX), gates.late_amplifier_tail_window
            )
        )

        # Gate 0: R² threshold
        pf_r2_thresh = (
            gates.r2_floor_late_amplifier if pf_is_late
            else gates.r2_floor_standard
        )
        if not pf_reject and pf_r2 is not None and pf_r2 < pf_r2_thresh:
            pf_reject = True
            pf_reason = f'R² = {pf_r2:.4f} < {pf_r2_thresh}'

        # Gate 2: fit window width
        if not pf_reject:
            pf_fs2 = pf_r.get('fit_start_cycle')
            pf_fe2 = pf_r.get('fit_end_cycle')
            if (pf_fs2 is not None and pf_fe2 is not None
                    and pf_fe2 - pf_fs2 < gates.min_fit_window_cycles):
                pf_reject = True
                pf_reason = (
                    f'Fit window {pf_fe2 - pf_fs2:.0f} cycles '
                    f'< {gates.min_fit_window_cycles}'
                )

        # Gate 2b: linear vs MAK2
        # The high-R² bypass applies universally: for clean sigmoids the
        # pre-inflection region inside the fit window is short (2–3
        # cycles for a sharp rise), so MAK2 and linear necessarily fit
        # similarly there even though the overall fit is excellent.
        # vermeulen1 NM23A.1500 (R²=0.9992 textbook sigmoid) was being
        # rejected by this gate because it wasn't classified "late".
        pf_late_bypass_2b = (
            pf_r2 is not None
            and pf_r2 >= gates.gate_2b_late_bypass_r2
        )
        if (not pf_reject and not pf_late_bypass_2b
                and pf_r.get('D0') is not None
                and not (isinstance(pf_r['D0'], float) and np.isnan(pf_r['D0']))
                and pf_r.get('fluor_data') is not None):
            try:
                pf_fs2b = pf_r.get('fit_start_cycle')
                pf_fe2b = pf_r.get('fit_end_cycle')
                if pf_fs2b is not None and pf_fe2b is not None:
                    pf_m2b = MAK2Model()
                    pf_c_full2b = cycles[:len(pf_r['fluor_data'])]
                    pf_pred2b = pf_m2b.simulate_to_cycle(
                        D0=pf_r['D0'], k=pf_r['k'], P0=pf_r['P0'],
                        cycles=pf_c_full2b,
                        F_bg_intercept=pf_r['F_bg_intercept'],
                        F_bg_slope=pf_r['F_bg_slope'],
                    )
                    pf_win2b = (pf_c_full2b >= pf_fs2b) & (pf_c_full2b <= pf_fe2b)
                    pf_cycles_win = pf_c_full2b[pf_win2b]
                    pf_pred_win2b = pf_pred2b[pf_win2b]
                    pf_fluor_full2b = np.asarray(pf_r['fluor_data'])
                    pf_fluor_win2b = pf_fluor_full2b[pf_win2b]

                    pf_d1_2b = np.gradient(pf_pred_win2b, pf_cycles_win)
                    pf_max_slope_idx = int(np.argmax(pf_d1_2b))
                    pf_max_slope_cycle = pf_cycles_win[pf_max_slope_idx]

                    pf_pre_mask = pf_cycles_win <= pf_max_slope_cycle
                    pf_fluor_pre = pf_fluor_win2b[pf_pre_mask]
                    pf_cycles_pre = pf_cycles_win[pf_pre_mask]

                    if len(pf_fluor_pre) >= 4:
                        pf_coeffs = np.polyfit(pf_cycles_pre, pf_fluor_pre, 1)
                        pf_lin_pred = np.polyval(pf_coeffs, pf_cycles_pre)
                        pf_ss_tot = float(np.sum((pf_fluor_pre - np.mean(pf_fluor_pre))**2))
                        if pf_ss_tot > 0:
                            pf_ss_res_lin = float(np.sum((pf_fluor_pre - pf_lin_pred)**2))
                            pf_r2_lin = 1.0 - pf_ss_res_lin / pf_ss_tot
                            pf_mak2_pre = pf_pred_win2b[pf_pre_mask]
                            pf_ss_res_mak = float(np.sum((pf_fluor_pre - pf_mak2_pre)**2))
                            pf_r2_mak = 1.0 - pf_ss_res_mak / pf_ss_tot
                            if pf_r2_mak - pf_r2_lin < gates.min_r2_gap_mak2_vs_linear:
                                pf_reject = True
                                pf_reason = (
                                    f'MAK2 not better than linear in growth region '
                                    f'(R²_MAK2={pf_r2_mak:.4f}, R²_lin={pf_r2_lin:.4f})'
                                )
            except Exception:
                pass

        # Gate 3: sigmoid shape
        pf_fit_width = (
            (pf_r.get('fit_end_cycle') or 0)
            - (pf_r.get('fit_start_cycle') or 0)
        )
        pf_high_r2 = (
            pf_r2 is not None
            and pf_r2 >= gates.gate_3_high_r2_bypass_r2
            and pf_fit_width >= gates.gate_3_high_r2_bypass_min_window
        )
        pf_late_bypass_3 = (
            pf_is_late and pf_r2 is not None
            and pf_r2 >= gates.gate_3_late_bypass_r2
        )
        if (not pf_reject and not pf_late_bypass_3 and not pf_high_r2
                and pf_r.get('D0') is not None
                and not (isinstance(pf_r['D0'], float) and np.isnan(pf_r['D0']))
                and pf_r.get('fluor_data') is not None):
            try:
                pf_fs3 = pf_r.get('fit_start_cycle')
                pf_fe3 = pf_r.get('fit_end_cycle')
                if pf_fs3 is not None and pf_fe3 is not None:
                    pf_m = MAK2Model()
                    pf_c_full = cycles[:len(pf_r['fluor_data'])]
                    pf_pred = pf_m.simulate_to_cycle(
                        D0=pf_r['D0'], k=pf_r['k'], P0=pf_r['P0'],
                        cycles=pf_c_full,
                        F_bg_intercept=pf_r['F_bg_intercept'],
                        F_bg_slope=pf_r['F_bg_slope'],
                    )
                    pf_win_mask = (pf_c_full >= pf_fs3) & (pf_c_full <= pf_fe3)
                    pf_pred_win = pf_pred[pf_win_mask]
                    if len(pf_pred_win) >= 5:
                        pf_d1 = np.gradient(pf_pred_win)
                        pf_d2 = np.gradient(pf_d1)
                        pf_pred_range = float(np.max(pf_pred_win) - np.min(pf_pred_win))
                        pf_d2_thresh = pf_pred_range * gates.inflection_threshold_pct_of_range
                        # Inflection criterion is asymmetric in magnitude:
                        # the positive (acceleration) side must clearly
                        # exceed noise via the relative threshold, but the
                        # negative (deceleration) side only needs to be
                        # negative — slow-k sigmoids decelerate gradually
                        # over many cycles, so |d²| stays small on the
                        # negative side even though d² is genuinely
                        # crossing zero. Requiring symmetric magnitudes
                        # rejected real slow amplifiers (k = 0.04–0.07
                        # range; observed on 4 of 7 kbqPCR false-FAILs).
                        pf_has_inflection = (
                            np.any(pf_d2 > pf_d2_thresh)
                            and np.any(pf_d2 < 0)
                        )
                        if not pf_has_inflection:
                            pf_reject = True
                            pf_reason = 'No inflection (monotone curve)'
            except Exception:
                pass

        if pf_reject:
            results_list[pf_idx]['Success'] = ''
            results_list[pf_idx]['error'] = f'No amplification detected ({pf_reason})'
            results_list[pf_idx]['D0'] = None
            results_list[pf_idx]['k'] = None
            results_list[pf_idx]['P0'] = None
            results_list[pf_idx]['Ct'] = None

    return results_list


def build_replicate_groups(results_list, sample_metadata, channels):
    """Aggregate per-well results into per-(channel, sample-name) replicate stats.

    Replicates are identified by matching ``Sample Name`` in the
    metadata: wells that share a name within the same channel are
    treated as technical replicates. For each replicate group, the
    function reports mean, SD, and CV% for D0 and Ct.

    The replicate-CV% number is what the UI surfaces as the headline
    precision metric — it's the empirical answer to "how
    reproducible is this assay?" for the user's particular plate.

    Args:
        results_list: Per-well results (post-quality-gates).
        sample_metadata: Per-well metadata dict (keyed by composite
            ``"{channel}_{well}"`` name); ``Sample Name`` is the
            grouping key.
        channels: List of channels present on the plate.

    Returns:
        DataFrame with one row per (channel, sample-name) group.
    """
    # Map sample keys to their sample names
    results_df = pd.DataFrame([{k: v for k, v in r.items() if k != 'fluor_data'} for r in results_list])
    if results_df.empty:
        return None

    # Add Group column based on sample name from metadata
    groups = {}
    for _, row in results_df.iterrows():
        sample_key = row['Sample']
        meta = sample_metadata.get(sample_key, {})
        sample_name = meta.get('Sample Name', sample_key)
        target = meta.get('Target Name', '')

        # For multi-target, prefix with target name
        if target and len(channels) > 1:
            group_key = f"{target} -- {sample_name}"
        else:
            group_key = sample_name

        if group_key not in groups:
            groups[group_key] = []
        groups[group_key].append(sample_key)

    # Only keep groups with >1 member
    groups = {k: v for k, v in groups.items() if len(v) > 1}

    if not groups:
        return None

    # Build group mapping: sample_key -> group_name
    sample_to_group = {}
    for group_name, members in groups.items():
        for m in members:
            sample_to_group[m] = group_name

    results_df['Group'] = results_df['Sample'].map(sample_to_group)

    # Filter to only grouped samples with valid results
    grouped_df = results_df[results_df['Group'].notna()].copy()
    if grouped_df.empty:
        return None

    return calculate_replicate_stats(grouped_df)


def _make_scatter_series(ws, x_col, y_col, data_rows, title="Data",
                         color="4472C4"):
    """Build an openpyxl scatter ``Series`` with marker styling.

    Markers only (no connecting line) — appropriate for standard-curve
    point plots where a separate trendline carries the regression.
    Also defined identically in ``app.py``; see CLAUDE.md.
    """
    from openpyxl.chart import Reference, Series as XlSeries
    from openpyxl.chart.marker import Marker

    x_vals = Reference(ws, min_col=x_col, min_row=2, max_row=data_rows + 1)
    y_vals = Reference(ws, min_col=y_col, min_row=2, max_row=data_rows + 1)
    series = XlSeries(y_vals, x_vals, title=title)
    series.marker = Marker(symbol='circle', size=7)
    series.marker.graphicalProperties.solidFill = color
    series.graphicalProperties.line.noFill = True
    return series


def _style_axis(axis, title, num_fmt="General"):
    """Apply title, number format, and tick/gridline styling to an openpyxl chart axis."""
    axis.title = title
    axis.delete = False
    axis.numFmt = num_fmt
    axis.majorTickMark = "out"
    axis.majorGridlines = None


def _add_std_curve_d0_chart(ws, df, data_rows):
    """Embed the D0 power-law standard-curve scatter+trendline into ``ws``.

    Companion to ``_add_std_curve_ct_chart``; the two charts let the
    user compare D0-vs-Ct calibration linearity directly inside the
    Excel output without re-running the app.
    """
    from openpyxl.chart import ScatterChart
    from openpyxl.chart.trendline import Trendline
    from openpyxl.utils import get_column_letter

    cols = list(df.columns)
    x_col = cols.index('log10_D0') + 1 if 'log10_D0' in cols else None
    y_col = cols.index('log10_Copies') + 1 if 'log10_Copies' in cols else None
    if not (x_col and y_col):
        return

    chart = ScatterChart()
    chart.title = "D0 Standard Curve (log-log)"
    chart.legend.position = 'b'
    _style_axis(chart.x_axis, "log10(D0)", "0.00")
    _style_axis(chart.y_axis, "log10(Known Copies)", "0.00")
    chart.width = 18
    chart.height = 12

    series = _make_scatter_series(ws, x_col, y_col, data_rows, "Standards")
    series.trendline = Trendline(trendlineType='linear',
                                 dispRSqr=True, dispEq=True)
    chart.series.append(series)
    ws.add_chart(chart, f"{get_column_letter(len(cols) + 2)}2")


def _add_std_curve_ct_chart(ws, df, data_rows):
    """Embed the Ct standard-curve scatter+trendline into ``ws``.

    Ct convention: x = Ct, y = log10(Known Copies). Companion to
    ``_add_std_curve_d0_chart``.
    """
    from openpyxl.chart import ScatterChart
    from openpyxl.chart.trendline import Trendline
    from openpyxl.utils import get_column_letter

    cols = list(df.columns)
    x_col = cols.index('Ct') + 1 if 'Ct' in cols else None
    y_col = cols.index('log10_Copies') + 1 if 'log10_Copies' in cols else None
    if not (x_col and y_col):
        return

    chart = ScatterChart()
    chart.title = "Ct Standard Curve"
    chart.legend.position = 'b'
    _style_axis(chart.x_axis, "Ct", "0.0")
    _style_axis(chart.y_axis, "log10(Known Copies)", "0.00")
    chart.width = 18
    chart.height = 12

    series = _make_scatter_series(ws, x_col, y_col, data_rows, "Standards")
    series.trendline = Trendline(trendlineType='linear',
                                 dispRSqr=True, dispEq=True)
    chart.series.append(series)
    ws.add_chart(chart, f"{get_column_letter(len(cols) + 2)}2")


def build_excel(results_list, cycles, all_samples, no_signal_samples,
                no_signal_fluor, sample_metadata, channels, batch_settings,
                replicate_stats_df=None, precision_comparison_df=None,
                std_curve_sheets=None, chart_sheets=None):
    """Write the multi-sheet Excel result file for one plate.

    The output is intentionally byte-comparable to what the
    Streamlit app produces — once written, the file can be loaded
    back into the app via "Load Previous Results" for visualization.
    Keeping the format identical between offline and interactive
    paths is a deliberate design choice; see CLAUDE.md.

    Sheets written (in order):

    - ``Batch Results``: per-well fit parameters + status (the
      authoritative results table).
    - ``No Signal Samples``: wells flagged before fitting + reason.
    - ``Replicate Statistics``: per-(channel, sample-name)
      replicate aggregates.
    - ``Standard Curve``: per-channel D0 calibration regressions.
    - ``Ct Standard Curve``: per-channel Ct calibration regressions
      (for comparison only).
    - ``Settings``: the configuration values used for this run
      (FIRST_FIT_CYCLE, CYCLES_BEFORE_MAX, …).
    - ``Metadata``: plate-level info (file paths, channels, well count).


    chart_sheets: dict of {sheet_name: {data: DataFrame, summary: dict,
                  chart_type: str}} for tabs with native Excel charts.
    """
    if std_curve_sheets is None:
        std_curve_sheets = {}
    if chart_sheets is None:
        chart_sheets = {}

    # Build results DataFrame (excluding internal-only keys)
    # KEEP 'error' column — the app needs it to show FAIL status on reload
    hidden = {'fluor_data', 'bg_slope_est', 'bg_intercept_est', 'retry_stage',
              'retry_attempts', 'retry_timed_out', 'retry_elapsed_s'}
    display_results = [{k: v for k, v in r.items() if k not in hidden} for r in results_list]
    results_df = pd.DataFrame(display_results)

    # Ensure empty strings in Success don't become NaN on Excel roundtrip
    results_df['Success'] = results_df['Success'].fillna('')

    # Add Channel and Well columns for multi-channel data
    if channels and len(channels) > 1 and results_df['Sample'].str.contains('_').any():
        results_df.insert(0, 'Channel', results_df['Sample'].apply(_ch))
        results_df.insert(1, 'Well', results_df['Sample'].apply(_get_well_pos))

    extra_sheets = {}

    # No Signal Samples sheet
    if no_signal_samples:
        no_signal_df = pd.DataFrame([
            {
                'Sample': name,
                'Reason': info['reason'],
                'Fluorescence Range': f"{info['F_range']:.4f}",
                '% of Max on Plate': f"{info['F_range_pct']:.1f}%"
            }
            for name, info in no_signal_samples.items()
        ])
        extra_sheets['No Signal Samples'] = no_signal_df

    # Replicate Statistics sheet
    if replicate_stats_df is not None and len(replicate_stats_df) > 0:
        extra_sheets['Replicate Statistics'] = replicate_stats_df

    # Precision Comparison sheet
    if precision_comparison_df is not None and len(precision_comparison_df) > 0:
        extra_sheets['Precision Comparison'] = precision_comparison_df

    # Standard curve variance sheets (plain data, no chart)
    for sheet_name, df in std_curve_sheets.items():
        if df is not None and len(df) > 0:
            extra_sheets[sheet_name] = df

    # Input Data sheet: raw fluorescence
    input_data = {'Cycle': cycles}
    for wn, wd in all_samples.items():
        arr = np.asarray(wd)
        if arr.ndim >= 1:
            input_data[wn] = arr[:len(cycles)]
    for ns_name in no_signal_samples:
        if ns_name not in input_data and ns_name in no_signal_fluor:
            arr = np.asarray(no_signal_fluor[ns_name])
            if arr.ndim >= 1:
                input_data[ns_name] = arr[:len(cycles)]
    for rl in results_list:
        rn = rl.get('Sample', '')
        if rn and rn not in input_data and rl.get('fluor_data') is not None:
            arr = np.asarray(rl['fluor_data'])
            if arr.ndim >= 1:
                input_data[rn] = arr[:len(cycles)]
    extra_sheets['Input Data'] = pd.DataFrame(input_data)

    # Metadata sheet
    if sample_metadata:
        meta_rows = []
        for mk, mv in sample_metadata.items():
            row = {'Well_Key': mk}
            if isinstance(mv, dict):
                row.update(mv)
            meta_rows.append(row)
        if meta_rows:
            extra_sheets['Metadata'] = pd.DataFrame(meta_rows)

    # Settings sheet
    if batch_settings:
        settings_rows = []
        for sk, sv in batch_settings.items():
            if isinstance(sv, dict):
                for sk2, sv2 in sv.items():
                    settings_rows.append({
                        'Setting': f'{sk}.{sk2}',
                        'Value': str(sv2) if sv2 is not None else '',
                    })
            else:
                settings_rows.append({
                    'Setting': sk,
                    'Value': str(sv) if sv is not None else '',
                })
        if settings_rows:
            extra_sheets['Settings'] = pd.DataFrame(settings_rows)

    # Build Excel file
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Batch Results', index=False)

        # Plain data sheets
        for sheet_name, df in extra_sheets.items():
            if df is not None and len(df) > 0:
                safe_name = sheet_name[:31]
                df.to_excel(writer, sheet_name=safe_name, index=False)

        # Chart sheets: data + summary stats + native Excel scatter chart
        for sheet_name, spec in chart_sheets.items():
            safe_name = sheet_name[:31]
            df = spec.get('data')
            summary = spec.get('summary', {})
            chart_type = spec.get('chart_type', '')

            if df is None or len(df) == 0:
                continue

            df.to_excel(writer, sheet_name=safe_name, index=False)
            ws = writer.sheets[safe_name]
            data_rows = len(df)

            # Write summary stats below data
            summary_start_row = data_rows + 3  # leave a blank row
            row = summary_start_row
            ws.cell(row=row, column=1, value='--- Summary ---')
            row += 1
            for k, v in summary.items():
                ws.cell(row=row, column=1, value=k)
                ws.cell(row=row, column=2, value=v)
                row += 1

            # Create scatter chart based on chart_type
            if chart_type == 'std_curve_d0':
                _add_std_curve_d0_chart(ws, df, data_rows)
            elif chart_type == 'std_curve_ct':
                _add_std_curve_ct_chart(ws, df, data_rows)

    buf.seek(0)
    return buf.getvalue()


def process_plate(mc_file, meta_file, plate_name):
    """Run the full per-plate pipeline and write the Excel result file.

    The end-to-end orchestration:

    ``load_plate_data`` → ``detect_no_signal_samples`` → ``run_pass1``
    → ``run_pass2`` → ``run_quality_gates`` → ``build_replicate_groups``
    → ``compare_precision`` → ``build_standard_curve`` (per channel)
    → ``apply_calibration`` (per channel) → ``build_ct_standard_curve``
    (per channel) → ``apply_ct_calibration`` (per channel) →
    ``build_excel``.

    Per-step timing is printed so the user can see which step
    dominates wall-clock for their plate. All exceptions in any
    step bubble up and terminate this plate; the top-level
    ``__main__`` driver catches them so a single bad plate doesn't
    abort the whole batch.

    Args:
        mc_file: Path to the Multicomponent CSV.
        meta_file: Path to the Sample Setup / Results CSV.
        plate_name: Display name and stem for the output file
            (``Results/{plate_name}_MAK2Plus_Results.xlsx``).

    Returns:
        ``Path`` to the written Excel file.
    """
    print(f"\n{'='*70}")
    print(f"  PLATE: {plate_name}")
    print(f"  Data:  {mc_file.name}")
    print(f"  Meta:  {meta_file.name}")
    print(f"{'='*70}")

    plate_t0 = time.perf_counter()

    # Load data
    print("\nLoading data...")
    cycles, all_samples, channels, rox_by_well, sample_metadata, abi_results_meta = \
        load_plate_data(mc_file, meta_file)
    print(f"  Loaded {len(all_samples)} wells, {len(cycles)} cycles, "
          f"channels: {channels}")

    # Phase 0: Signal detection (per-channel)
    print("\nPhase 0: Signal detection...")
    if len(channels) > 1:
        valid_samples = {}
        no_signal_samples = {}
        plate_stats = {'max_range': 0}
        for det_ch in channels:
            det_ch_samples = {
                name: fluor for name, fluor in all_samples.items()
                if _ch(name) == det_ch
            }
            if not det_ch_samples:
                continue
            v, ns, ps = detect_no_signal_samples(
                cycles, det_ch_samples,
                min_range_pct=2.0, min_r2=0.80, verbose=False
            )
            valid_samples.update(v)
            no_signal_samples.update(ns)
            plate_stats['max_range'] = max(plate_stats['max_range'], ps.get('max_range', 0))
    else:
        valid_samples, no_signal_samples, plate_stats = detect_no_signal_samples(
            cycles, all_samples, min_range_pct=2.0, min_r2=0.80, verbose=False
        )

    # Metadata rescue
    if sample_metadata:
        rescued = []
        for sname, info in list(no_signal_samples.items()):
            wm = sample_metadata.get(sname, {})
            if 'Ct_instrument' not in wm:
                continue
            inst_ct = wm.get('Ct_instrument')
            inst_noamp = wm.get('NOAMP', False)
            inst_undetermined = (
                inst_ct is None or (isinstance(inst_ct, float) and np.isnan(inst_ct))
            )
            if not inst_undetermined and not inst_noamp:
                valid_samples[sname] = all_samples[sname]
                del no_signal_samples[sname]
                rescued.append(sname)
        if rescued:
            print(f"  Rescued {len(rescued)} wells via instrument metadata")

    no_signal_fluor = {name: all_samples[name] for name in no_signal_samples if name in all_samples}
    all_samples_to_fit = valid_samples
    print(f"  Valid: {len(all_samples_to_fit)}, No signal: {len(no_signal_samples)}")

    # Phase 0b: Threshold computation
    print("\nComputing thresholds...")
    global_threshold = None
    global_baseline_mean = 0.0
    channel_thresholds = {}
    channel_baseline_means = {}

    all_fluorescence = list(all_samples_to_fit.values())
    if all_fluorescence:
        baseline_end = max(3, int(len(cycles) * 0.15))
        ch_arrays = {}
        for sname, fd in all_samples_to_fit.items():
            ch_arrays.setdefault(_ch(sname), []).append(fd)

        for ch, arrays in ch_arrays.items():
            ch_sds = []
            ch_means = []
            ch_early_all = []
            for fd in arrays:
                early = fd[:baseline_end]
                ch_early_all.extend(early)
                ch_sds.append(np.std(early))
                ch_means.append(np.mean(early))

            ch_sub = (np.mean(ch_early_all) < 0.1) or np.any(np.array(ch_early_all) < 0)
            ch_baseline_mean = 0.0 if ch_sub else np.median(ch_means)
            ch_median_sd = np.median(ch_sds)
            ch_max = np.max([np.max(fd) for fd in arrays])
            ch_dyn_range = ch_max - ch_baseline_mean
            ch_thresh = max(
                10 * ch_median_sd if ch_median_sd > 0 else 0.01,
                0.05 * ch_dyn_range
            )
            channel_thresholds[ch] = ch_thresh
            channel_baseline_means[ch] = ch_baseline_mean

        if len(channel_thresholds) == 1:
            global_threshold = next(iter(channel_thresholds.values()))
            global_baseline_mean = next(iter(channel_baseline_means.values()))
        else:
            all_sds = [s for arrays in ch_arrays.values()
                       for fd in arrays for s in [np.std(fd[:baseline_end])]]
            all_means = [np.mean(fd[:baseline_end])
                         for arrays in ch_arrays.values() for fd in arrays]
            already_sub = (np.mean([v for fd in all_fluorescence for v in fd[:baseline_end]]) < 0.1)
            global_baseline_mean = 0.0 if already_sub else np.median(all_means)
            plate_max = np.max([np.max(f) for f in all_fluorescence])
            global_threshold = max(
                10 * np.median(all_sds) if all_sds else 0.01,
                0.05 * (plate_max - global_baseline_mean)
            )

    # Override with instrument thresholds when ROX is available
    rox_norm_active = bool(rox_by_well)
    if abi_results_meta and rox_norm_active:
        inst_thresholds = abi_results_meta.get('channel_thresholds', {})
        if inst_thresholds:
            for ch, val in inst_thresholds.items():
                if ch in channel_thresholds:
                    channel_thresholds[ch] = val
            if len(inst_thresholds) == 1:
                global_threshold = next(iter(inst_thresholds.values()))

    # Pass 1
    print(f"\nPass 1: Fitting {len(all_samples_to_fit)} samples...")
    results_list = run_pass1(
        all_samples_to_fit, cycles, sample_metadata, rox_by_well,
        channel_thresholds, global_threshold, channel_baseline_means,
        global_baseline_mean
    )

    # Pass 2
    print(f"\nPass 2: Channel-aware retry...")
    results_list = run_pass2(
        results_list, cycles, sample_metadata, rox_by_well,
        channel_thresholds, global_threshold, channel_baseline_means,
        global_baseline_mean
    )

    # Pass 3: Quality gates
    print(f"\nPass 3: Quality gates...")
    results_list = run_quality_gates(results_list, cycles)

    # Count results
    n_success = sum(1 for r in results_list if str(r.get('Success', '')).startswith('✓'))
    n_failed = sum(1 for r in results_list if r.get('error') is not None)
    n_warn = len(results_list) - n_success - n_failed
    print(f"\n  Results: {n_success} ✓  {n_warn} ⚠️  {n_failed} ✗")

    # Replicate statistics
    print("\nComputing replicate statistics...")
    replicate_stats_df = build_replicate_groups(results_list, sample_metadata, channels)
    precision_comparison_df = None
    if replicate_stats_df is not None:
        print(f"  {len(replicate_stats_df)} replicate groups")
        # Precision comparison (Ct vs D0 CV%)
        try:
            precision_comparison_df = compare_precision(replicate_stats_df, efficiency=0.95)
            if precision_comparison_df is not None and len(precision_comparison_df) > 0:
                print(f"  Precision comparison: {len(precision_comparison_df)} groups")
        except Exception as e:
            print(f"  Precision comparison failed: {e}")
    else:
        print("  No replicate groups found")

    # Standard curve calibration
    print("\nComputing standard curves...")
    std_curve_sheets = {}   # variance sheets (plain data, no chart)
    chart_sheets = {}       # D0/Ct sheets (data + summary + Excel chart)
    has_standards = (
        sample_metadata is not None
        and any(m.get('Task') == 'STANDARD' for m in sample_metadata.values())
    )

    # Build results_df for calibration (same as what goes into Excel)
    hidden_cal = {'fluor_data', 'bg_slope_est', 'bg_intercept_est', 'retry_stage',
                  'retry_attempts', 'retry_timed_out', 'retry_elapsed_s'}
    cal_results = [{k: v for k, v in r.items() if k not in hidden_cal} for r in results_list]
    results_df_cal = pd.DataFrame(cal_results)

    if has_standards:
        # Determine per-channel calibration
        derived_ch = results_df_cal['Sample'].apply(_ch)
        unique_ch = derived_ch[derived_ch != 'default'].unique()
        if len(unique_ch) > 1:
            results_df_cal.insert(0, 'Channel', derived_ch)
            cal_channels = list(unique_ch)
        else:
            cal_channels = [None]

        for cal_ch in cal_channels:
            if cal_ch is not None:
                ch_mask = results_df_cal['Channel'] == cal_ch
                ch_df = results_df_cal[ch_mask].copy()
                ch_meta = {k: v for k, v in sample_metadata.items()
                           if k.startswith(f"{cal_ch}_")}
            else:
                ch_df = results_df_cal
                ch_meta = sample_metadata

            ch_label = f" ({cal_ch})" if cal_ch else ""

            calibration = build_standard_curve(ch_df, ch_meta)
            ct_calibration = build_ct_standard_curve(ch_df, ch_meta)

            if calibration is not None:
                print(f"  D0 standard curve{ch_label}: R²={calibration['r_squared']:.4f}, "
                      f"{calibration['n_standards']} wells, {calibration['n_concentrations']} levels")

                # D0 standard curve chart sheet (data + summary + Excel chart)
                chart_sheets[f'Std Curve D0{ch_label}'] = {
                    'data': calibration['per_point_data'].copy(),
                    'summary': {
                        'slope': calibration['slope'],
                        'intercept': calibration['intercept'],
                        'r_squared': calibration['r_squared'],
                        'n_standards': calibration['n_standards'],
                        'n_concentrations': calibration['n_concentrations'],
                        'median_cf': calibration.get('median_cf', np.nan),
                    },
                    'chart_type': 'std_curve_d0',
                }

                # Variance sheet (plain data)
                var_data = []
                for copies_val, var_info in sorted(calibration['replicate_variance'].items(), reverse=True):
                    row = {
                        'Known Copies': copies_val,
                        'N Replicates': var_info['n_replicates'],
                        'Mean D0': var_info['mean_D0'],
                        'SD D0': var_info['sd_D0'],
                        'D0 CV%': var_info['cv_D0_pct'],
                    }
                    if ct_calibration is not None:
                        ct_var = ct_calibration['replicate_variance'].get(copies_val)
                        if ct_var:
                            row['Mean Ct'] = ct_var['mean_Ct']
                            row['SD Ct'] = ct_var['sd_Ct']
                            row['Ct CV%'] = ct_var['cv_Ct_pct']
                    var_data.append(row)
                if var_data:
                    std_curve_sheets[f'Std Curve Variance{ch_label}'] = pd.DataFrame(var_data)

                # Apply calibration to results
                if cal_ch is not None:
                    ch_mask_apply = results_df_cal['Channel'] == cal_ch
                    cal_subset = apply_calibration(results_df_cal[ch_mask_apply].copy(), calibration=calibration)
                    results_df_cal.loc[ch_mask_apply, 'Copies_D0'] = cal_subset['Copies_D0']
                else:
                    results_df_cal = apply_calibration(results_df_cal, calibration=calibration)

            if ct_calibration is not None:
                print(f"  Ct standard curve{ch_label}: R²={ct_calibration['r_squared']:.4f}, "
                      f"efficiency={ct_calibration['efficiency']*100:.1f}%")

                # Ct standard curve chart sheet (data + summary + Excel chart)
                chart_sheets[f'Std Curve Ct{ch_label}'] = {
                    'data': ct_calibration['per_point_data'].copy(),
                    'summary': {
                        'slope': ct_calibration['slope'],
                        'intercept': ct_calibration['intercept'],
                        'r_squared': ct_calibration['r_squared'],
                        'efficiency': ct_calibration['efficiency'],
                        'n_standards': ct_calibration['n_standards'],
                        'n_concentrations': ct_calibration['n_concentrations'],
                    },
                    'chart_type': 'std_curve_ct',
                }

                if cal_ch is not None:
                    ch_mask_apply = results_df_cal['Channel'] == cal_ch
                    ct_subset = apply_ct_calibration(results_df_cal[ch_mask_apply].copy(), ct_calibration)
                    results_df_cal.loc[ch_mask_apply, 'Copies_Ct'] = ct_subset['Copies_Ct']
                else:
                    results_df_cal = apply_ct_calibration(results_df_cal, ct_calibration)

            if calibration is None and ct_calibration is None:
                print(f"  No standard curve{ch_label} (insufficient data)")

        # Update results_list with calibrated copies if computed
        if 'Copies_D0' in results_df_cal.columns or 'Copies_Ct' in results_df_cal.columns:
            for i, r in enumerate(results_list):
                if 'Copies_D0' in results_df_cal.columns:
                    results_list[i]['Copies_D0'] = results_df_cal.iloc[i].get('Copies_D0', np.nan)
                if 'Copies_Ct' in results_df_cal.columns:
                    results_list[i]['Copies_Ct'] = results_df_cal.iloc[i].get('Copies_Ct', np.nan)
    else:
        print("  No STANDARD wells found in metadata — skipping")

    # Build batch settings
    batch_settings = {
        'first_fit_cycle': FIRST_FIT_CYCLE,
        'cycles_before_max': CYCLES_BEFORE_MAX,
        'cycles_after_max': CYCLES_AFTER_MAX,
        'auto_truncate': AUTO_TRUNCATE,
        'truncate_cycle': TRUNCATE_CYCLE,
        'custom_bounds_dict': CUSTOM_BOUNDS_DICT,
        'global_threshold': global_threshold,
        'global_baseline_mean': global_baseline_mean,
        'channel_thresholds': channel_thresholds,
        'channel_baseline_means': channel_baseline_means,
    }

    # Build and save Excel
    print("\nBuilding Excel file...")
    excel_bytes = build_excel(
        results_list, cycles, all_samples, no_signal_samples,
        no_signal_fluor, sample_metadata, channels, batch_settings,
        replicate_stats_df, precision_comparison_df, std_curve_sheets,
        chart_sheets
    )

    output_path = OUTPUT_DIR / f"{plate_name}_MAK2Plus_Results.xlsx"
    with open(output_path, 'wb') as f:
        f.write(excel_bytes)

    elapsed = time.perf_counter() - plate_t0
    print(f"\n✅ {plate_name} complete in {elapsed/60:.1f} minutes")
    print(f"   Saved to: {output_path}")

    return output_path


# ── Main ────────���────────────────────���────────────────────────────────────────

if __name__ == '__main__':
    # Lazy mkdir — only when this module is invoked as a script.
    # Module imports (tests, FastAPI, etc.) must remain side-effect-free.
    OUTPUT_DIR.mkdir(exist_ok=True, parents=True)

    print("=" * 70)
    print("  MAK2+ Offline Batch Fitting")
    print(f"  {len(PLATES)} plates to process")
    print(f"  Output directory: {OUTPUT_DIR}")
    print("=" * 70)

    if not PLATES:
        print("\n❌ No plate files found! Check DATA_DIR path.")
        sys.exit(1)

    total_t0 = time.perf_counter()
    completed = []
    failed = []

    for mc_file, meta_file, plate_name in PLATES:
        try:
            output_path = process_plate(mc_file, meta_file, plate_name)
            completed.append((plate_name, output_path))
        except Exception as e:
            print(f"\n❌ {plate_name} FAILED: {e}")
            traceback.print_exc()
            failed.append((plate_name, str(e)))

    total_elapsed = time.perf_counter() - total_t0
    print(f"\n\n{'='*70}")
    print(f"  ALL DONE — {total_elapsed/60:.1f} minutes total")
    print(f"  Completed: {len(completed)}/{len(PLATES)}")
    for name, path in completed:
        print(f"    ✅ {name}: {path}")
    if failed:
        print(f"  Failed: {len(failed)}")
        for name, err in failed:
            print(f"    ❌ {name}: {err}")
    print("=" * 70)
